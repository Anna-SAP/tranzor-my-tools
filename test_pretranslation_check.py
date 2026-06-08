"""Unit tests for the Pre-Translation Coverage Check engine.

All classification tests inject stub coverage / mr-state / score callables, so
they run headless with no SQLite, no GitLab, and no Tranzor API.
"""
import openpyxl
import pytest

import pretranslation_check as ptc


# ---------------------------------------------------------------------------
# parse_delta_xlsx
# ---------------------------------------------------------------------------
def _write_xlsx(path, header, data_rows, sheet="Data"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(header)
    for r in data_rows:
        ws.append(r)
    wb.save(path)


def test_parse_basic_key_value(tmp_path):
    f = tmp_path / "delta.xlsx"
    _write_xlsx(f, ["Key", "Value"], [
        ["RingCentral.bui.abc123.CANCELLATION_REASONS.21", "Company grown"],
        ["RingCentral.bui.abc123.DISCOUNT_OFFER", "Keep your service"],
    ])
    rows = ptc.parse_delta_xlsx(str(f))
    assert len(rows) == 2
    assert rows[0]["opus_id"] == "RingCentral.bui.abc123.CANCELLATION_REASONS.21"
    assert rows[0]["source_text"] == "Company grown"


def test_parse_with_jira_columns(tmp_path):
    f = tmp_path / "delta_jira.xlsx"
    _write_xlsx(
        f,
        ["Key", "Value", "Feature Ticket", "JIRA Link / Commit", "Source (file:line)"],
        [["RingCentral.bui.x.K1", "Hello", "BUP-4285",
          "https://jira/browse/BUP-4285", "flows/x/src/lang/module/en.json:12"]],
    )
    rows = ptc.parse_delta_xlsx(str(f))
    assert rows[0]["feature_ticket"] == "BUP-4285"
    assert rows[0]["source_path"].startswith("flows/x")


def test_parse_skips_blank_keys_and_dedupes(tmp_path):
    f = tmp_path / "delta_dupe.xlsx"
    _write_xlsx(f, ["Key", "Value"], [
        ["RingCentral.a.b.K1", "one"],
        [None, "blank key skipped"],
        ["RingCentral.a.b.K1", "duplicate skipped"],
    ])
    rows = ptc.parse_delta_xlsx(str(f))
    assert [r["opus_id"] for r in rows] == ["RingCentral.a.b.K1"]


def test_parse_missing_key_column_raises(tmp_path):
    f = tmp_path / "bad.xlsx"
    _write_xlsx(f, ["Foo", "Bar"], [["a", "b"]])
    with pytest.raises(ValueError):
        ptc.parse_delta_xlsx(str(f))


def test_parse_missing_file_raises():
    with pytest.raises(FileNotFoundError):
        ptc.parse_delta_xlsx("does/not/exist.xlsx")


# ---------------------------------------------------------------------------
# check_delta classification
# ---------------------------------------------------------------------------
LANGS = ["de-DE", "fr-FR"]


def _row(opus_id, src="src"):
    return {"opus_id": opus_id, "source_text": src}


def _cov(target_language, *, kind="mr", text="译文", mr_iid=40273,
         task_id="t1", release="26.3", project_id="web/web"):
    return {"target_language": target_language, "translated_text": text,
            "source_kind": kind, "mr_iid": mr_iid, "task_id": task_id,
            "release": release, "project_id": project_id}


def _check(rows, coverage, *, mr_state=None, scores=None,
           expected_langs=LANGS, **kw):
    return ptc.check_delta(
        rows, expected_langs=expected_langs,
        coverage_fn=lambda ids: {k: v for k, v in coverage.items() if k in ids},
        mr_state_fn=lambda pid, iid: (mr_state or {}).get(int(iid)) if iid else None,
        score_fn=lambda tid: (scores or {}).get(tid, {}),
        **kw,
    )


def test_red_when_no_coverage():
    out = _check([_row("X")], coverage={})
    assert out["rows"][0]["verdict"] == ptc.RED
    assert out["rows"][0]["reason"] == "no_coverage"
    assert out["summary"]["red"] == 1


def test_green_when_all_langs_mr_merged_and_score_ok():
    cov = {"X": [_cov("de-DE"), _cov("fr-FR")]}
    out = _check(
        [_row("X")], cov,
        mr_state={40273: "merged"},
        scores={"t1": {("X", "de-DE"): 99, ("X", "fr-FR"): 100}},
    )
    r = out["rows"][0]
    assert r["verdict"] == ptc.GREEN, r
    assert out["summary"]["green"] == 1
    assert out["summary"]["skip_pct"] == 100.0


def test_amber_when_only_scan_coverage():
    cov = {"X": [_cov("de-DE", kind="scan"), _cov("fr-FR", kind="scan")]}
    out = _check([_row("X")], cov)
    assert out["rows"][0]["verdict"] == ptc.AMBER
    assert out["rows"][0]["reason"] == "non_mr_coverage"


def test_amber_when_some_langs_missing():
    cov = {"X": [_cov("de-DE")]}  # fr-FR not covered by MR
    out = _check([_row("X")], cov, mr_state={40273: "merged"})
    r = out["rows"][0]
    assert r["verdict"] == ptc.AMBER
    assert r["reason"] == "langs_missing"
    assert r["missing_langs"] == ["fr-FR"]


def test_amber_when_mr_not_merged():
    cov = {"X": [_cov("de-DE"), _cov("fr-FR")]}
    out = _check([_row("X")], cov, mr_state={40273: "opened"},
                 scores={"t1": {("X", "de-DE"): 99, ("X", "fr-FR"): 99}})
    assert out["rows"][0]["verdict"] == ptc.AMBER
    assert out["rows"][0]["reason"] == "mr_not_merged"


def test_amber_when_known_score_below_threshold():
    cov = {"X": [_cov("de-DE"), _cov("fr-FR")]}
    out = _check([_row("X")], cov, mr_state={40273: "merged"},
                 scores={"t1": {("X", "de-DE"): 80, ("X", "fr-FR"): 99}})
    r = out["rows"][0]
    assert r["verdict"] == ptc.AMBER
    assert r["reason"] == "score_below"
    assert r["min_score"] == 80


def test_unknown_score_passes_by_default_but_blocks_when_strict():
    # TM/ICE matches carry no GMG score: merged + all langs covered, no scores.
    cov = {"X": [_cov("de-DE"), _cov("fr-FR")]}
    lenient = _check([_row("X")], cov, mr_state={40273: "merged"}, scores={})
    assert lenient["rows"][0]["verdict"] == ptc.GREEN

    strict = _check([_row("X")], cov, mr_state={40273: "merged"}, scores={},
                    block_on_unknown_score=True)
    assert strict["rows"][0]["verdict"] == ptc.AMBER
    assert strict["rows"][0]["reason"] == "score_unknown"


def test_require_merged_false_skips_gitlab():
    cov = {"X": [_cov("de-DE"), _cov("fr-FR")]}
    called = {"n": 0}

    def _mr(pid, iid):
        called["n"] += 1
        return None

    out = ptc.check_delta(
        [_row("X")], expected_langs=LANGS,
        coverage_fn=lambda ids: cov,
        mr_state_fn=_mr, score_fn=lambda tid: {},
        require_merged=False,
    )
    assert out["rows"][0]["verdict"] == ptc.GREEN
    assert called["n"] == 0  # merge state never queried


def test_no_token_unconfirmable_merge_is_conservative_amber():
    cov = {"X": [_cov("de-DE"), _cov("fr-FR")]}
    out = _check([_row("X")], cov, mr_state={}, scores={"t1": {}})  # state None
    assert out["rows"][0]["verdict"] == ptc.AMBER
    assert out["rows"][0]["reason"] == "mr_not_merged"


def test_infer_expected_langs_from_coverage():
    cov = {
        "A": [_cov("de-DE"), _cov("fr-FR")],
        "B": [_cov("ja-JP", text="")],  # empty text ignored
    }
    out = ptc.check_delta(
        [_row("A"), _row("B")],
        coverage_fn=lambda ids: cov,
        mr_state_fn=lambda p, i: "merged",
        score_fn=lambda tid: {},
    )
    assert out["summary"]["expected_langs"] == ["de-DE", "fr-FR"]


def test_summary_counts_and_export(tmp_path):
    cov = {
        "G": [_cov("de-DE"), _cov("fr-FR")],
        "A": [_cov("de-DE")],            # missing fr-FR -> amber
        # "R" absent -> red
    }
    out = _check(
        [_row("G"), _row("A"), _row("R")], cov,
        mr_state={40273: "merged"},
        scores={"t1": {("G", "de-DE"): 99, ("G", "fr-FR"): 99}},
    )
    s = out["summary"]
    assert (s["green"], s["amber"], s["red"], s["total"]) == (1, 1, 1, 3)
    assert s["manual_needed"] == 2

    out_path = tmp_path / "manual.xlsx"
    n = ptc.export_manual_subset(out, str(out_path), include_amber=True)
    assert n == 2
    wb = openpyxl.load_workbook(str(out_path))
    ws = wb["Data"]
    keys = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert set(keys) == {"A", "R"}


# ---------------------------------------------------------------------------
# accept_kinds policy (default MR-only vs broadened)
# ---------------------------------------------------------------------------
def test_accept_kinds_default_mr_only_flags_file_coverage_amber():
    cov = {"F": [_cov("de-DE", kind="file"), _cov("fr-FR", kind="file")]}
    out = _check([_row("F")], cov)
    assert out["rows"][0]["verdict"] == ptc.AMBER
    assert out["rows"][0]["reason"] == "non_mr_coverage"
    assert out["rows"][0]["source_kind"] == "file"


def test_accept_kinds_broadened_counts_file_as_skippable():
    cov = {"F": [_cov("de-DE", kind="file"), _cov("fr-FR", kind="file")]}
    out = ptc.check_delta(
        [_row("F")], expected_langs=LANGS,
        coverage_fn=lambda ids: cov,
        mr_state_fn=lambda p, i: None,
        score_fn=lambda tid: {},
        accept_kinds=("mr", "file"),
    )
    assert out["rows"][0]["verdict"] == ptc.GREEN
