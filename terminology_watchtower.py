"""
Term Watchtower — terminology compliance checker for my-tools.

Pure logic module (no Tkinter imports). Responsibilities:

- Glossary CSV loading + validation
- Text/locale normalization
- Translation candidate model
- Term-rule matching and issue generation
- Issue status persistence
- HTML / Excel / Markdown evidence export

Phase 1 is deterministic: no LLM, no fuzzy matching. See Terminology_Watchtower_PRD.md.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import sys
import threading
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SEVERITIES = ("Critical", "High", "Medium", "Low")
_SEVERITY_RANK = {s: i for i, s in enumerate(SEVERITIES)}  # lower = more severe

STATUSES = ("New", "Reviewed", "Reported", "Ignored")
DEFAULT_STATUS = "New"

ISSUE_REQUIRED_MISSING = "RequiredTermMissing"
ISSUE_FORBIDDEN_USED = "ForbiddenVariantUsed"
ISSUE_MISSING_TARGET = "MissingTargetForTerm"
ISSUE_BOTH = "RequiredTermMissing+ForbiddenVariantUsed"

SOURCE_KIND_LEGACY = "legacy"
SOURCE_KIND_MR = "mr"
SOURCE_KIND_SCAN = "scan"
SOURCE_KIND_FILE = "file_translation"
SOURCE_KIND_UNKNOWN = "unknown"

SOURCE_LABELS = {
    SOURCE_KIND_LEGACY: "Legacy",
    SOURCE_KIND_MR: "MR Pipeline",
    SOURCE_KIND_SCAN: "Scan Tasks",
    SOURCE_KIND_FILE: "File Translation",
    SOURCE_KIND_UNKNOWN: "Unknown",
}

ENGLISH_REFERENCE_LOCALES = ("en_US", "en-US", "en_GB", "en-GB", "en")

# Sister-locale fallbacks. When a glossary covers the source locale but
# not the fallback locale, the fallback inherits the source's approved
# translation. This reflects standard L10n practice for closely-related
# variants — e.g. Hong Kong / Macau Traditional Chinese inherit Taiwan
# Traditional unless overridden, and Commonwealth English variants
# inherit British English. Conservative on purpose: pt-PT vs pt-BR are
# NOT linked here because their vocabulary diverges meaningfully.
LOCALE_FALLBACKS: Dict[str, str] = {
    "zh_hk": "zh_tw",
    "zh_mo": "zh_tw",
    "en_au": "en_gb",
    "en_nz": "en_gb",
    "en_ie": "en_gb",
}


def _canonical_locale_display(normalized: str) -> str:
    """``zh_hk`` → ``zh-HK``; ``en`` → ``en``. For synthetic fallback rules."""
    parts = normalized.split("_")
    if len(parts) == 1:
        return parts[0]
    return parts[0] + "-" + parts[1].upper()

# Local storage layout
DATA_DIR_NAME = ".tranzor_exporter"
SUBDIR = "terminology_watchtower"
GLOSSARY_FILENAME = "glossary.csv"
STATUSES_FILENAME = "issue_statuses.json"
LAST_SCAN_FILENAME = "last_scan.json"


def _data_dir() -> str:
    return os.path.join(os.path.expanduser("~"), DATA_DIR_NAME, SUBDIR)


def ensure_data_dir() -> str:
    path = _data_dir()
    os.makedirs(path, exist_ok=True)
    return path


def glossary_path() -> str:
    return os.path.join(ensure_data_dir(), GLOSSARY_FILENAME)


def statuses_path() -> str:
    return os.path.join(ensure_data_dir(), STATUSES_FILENAME)


def last_scan_path() -> str:
    return os.path.join(ensure_data_dir(), LAST_SCAN_FILENAME)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

@dataclass
class TermRule:
    rule_id: str
    source_term: str
    locale: str  # normalized
    approved_translation: str
    source_aliases: List[str] = field(default_factory=list)
    forbidden_translations: List[str] = field(default_factory=list)
    product_scope: List[str] = field(default_factory=list)
    severity: str = "High"
    case_sensitive: bool = False
    enabled: bool = True
    notes: str = ""
    locale_display: str = ""  # original style preserved for UI

    def to_dict(self) -> dict:
        d = asdict(self)
        return d


@dataclass
class TranslationCandidate:
    candidate_id: str
    product: str
    locale: str  # normalized
    key: str
    source_text: str
    target_text: str
    source_kind: str = SOURCE_KIND_UNKNOWN
    task_id: str = ""
    mr_id: str = ""
    score: Optional[float] = None
    error_category: str = ""
    reference_url: str = ""
    raw: Optional[dict] = None
    locale_display: str = ""


@dataclass
class TerminologyIssue:
    issue_id: str
    rule_id: str
    candidate_id: str
    issue_type: str
    severity: str
    status: str
    source_term: str
    locale: str
    locale_display: str
    expected: str
    actual: str
    source_text: str
    product: str
    source_kind: str
    source_label: str
    key: str
    task_id: str
    mr_id: str
    score: Optional[float]
    error_category: str
    reference_url: str
    forbidden_found: List[str]
    rule_notes: str
    last_seen: str  # ISO8601

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class ScanSummary:
    total_active_issues: int = 0
    critical: int = 0
    high: int = 0
    medium: int = 0
    low: int = 0
    affected_terms: int = 0
    affected_locales: int = 0
    affected_products: int = 0
    scanned_candidates: int = 0
    skipped_missing_source: int = 0
    skipped_no_match: int = 0
    failed_sources: List[str] = field(default_factory=list)
    sources_covered: List[str] = field(default_factory=list)
    last_scan_at: str = ""
    glossary_rules: int = 0
    glossary_enabled_rules: int = 0

    def to_dict(self) -> dict:
        return asdict(self)


# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

_WS_RE = re.compile(r"\s+", re.UNICODE)


def normalize_text(s: Optional[str]) -> str:
    """Trim + collapse whitespace. Does NOT strip accents or punctuation."""
    if s is None:
        return ""
    s = str(s).strip()
    if not s:
        return ""
    return _WS_RE.sub(" ", s)


def casefold_text(s: str) -> str:
    return s.casefold()


def normalize_locale(s: Optional[str]) -> str:
    if s is None:
        return ""
    return str(s).strip().lower().replace("-", "_")


def parse_bool(s: Optional[str], default: bool = False) -> bool:
    if s is None:
        return default
    v = str(s).strip().lower()
    if v == "":
        return default
    if v in ("true", "1", "yes", "y"):
        return True
    if v in ("false", "0", "no", "n"):
        return False
    return default


def parse_pipe_list(s: Optional[str]) -> List[str]:
    if not s:
        return []
    return [p.strip() for p in str(s).split("|") if p.strip()]


def normalize_severity(s: Optional[str], default: str = "High") -> Optional[str]:
    """Return canonical severity (`Critical`/`High`/`Medium`/`Low`) or None if invalid."""
    if s is None or str(s).strip() == "":
        return default
    v = str(s).strip()
    for canon in SEVERITIES:
        if v.lower() == canon.lower():
            return canon
    return None


def severity_max(*sevs: str) -> str:
    """Return the most severe of the given severities."""
    valid = [s for s in sevs if s in _SEVERITY_RANK]
    if not valid:
        return "Low"
    return min(valid, key=lambda s: _SEVERITY_RANK[s])


# ---------------------------------------------------------------------------
# Stable ID helpers
# ---------------------------------------------------------------------------

def _stable_hash(*parts: str) -> str:
    h = hashlib.sha1()
    for p in parts:
        h.update((p or "").encode("utf-8", errors="replace"))
        h.update(b"\x1f")
    return h.hexdigest()[:16]


def make_rule_id(source_term: str, locale: str, approved: str,
                 product_scope: Iterable[str]) -> str:
    return "R-" + _stable_hash(
        normalize_text(source_term).casefold(),
        normalize_locale(locale),
        normalize_text(approved),
        "|".join(sorted(product_scope or [])),
    )


def make_issue_id(rule_id: str, candidate_id: str, issue_type: str) -> str:
    return "ISS-" + _stable_hash(rule_id, candidate_id, issue_type)


# ---------------------------------------------------------------------------
# Glossary CSV import
# ---------------------------------------------------------------------------

GLOSSARY_REQUIRED_COLS = ("source_term", "locale", "approved_translation")
GLOSSARY_OPTIONAL_COLS = (
    "rule_id", "source_aliases", "forbidden_translations",
    "product_scope", "severity", "case_sensitive", "enabled", "notes",
)
GLOSSARY_ALL_COLS = GLOSSARY_REQUIRED_COLS + GLOSSARY_OPTIONAL_COLS


@dataclass
class GlossaryImportResult:
    rules: List[TermRule]
    total_rows: int
    imported: int
    skipped: int
    errors: List[Tuple[int, str]]  # (1-based row number, message)


def _open_csv_text(path_or_text: str, *, is_path: bool) -> str:
    if is_path:
        with open(path_or_text, "rb") as f:
            raw = f.read()
    else:
        if isinstance(path_or_text, bytes):
            raw = path_or_text
        else:
            raw = path_or_text.encode("utf-8")
    # UTF-8 with BOM tolerated
    if raw.startswith(b"\xef\xbb\xbf"):
        raw = raw[3:]
    return raw.decode("utf-8", errors="replace")


def import_glossary_csv(path: str) -> GlossaryImportResult:
    text = _open_csv_text(path, is_path=True)
    return _parse_glossary_csv(text)


# ---------------------------------------------------------------------------
# Glossary XLSX import (RingCentral standard format)
# ---------------------------------------------------------------------------
#
# Expected layout — one sheet (default `Terms`), with a header row like:
#
#   | Term ID | Source Term (EN) | Target Term (DE-DE) | Part of Speech |
#   | Definition | Context | Status | Notes | Last Mod Date |
#
# Locale comes from the `Target Term (XX-YY)` header. Multiple target columns
# in one sheet are supported — one TermRule per (row, target column).
# Status `VALID` → enabled=True, anything else → enabled=False but still
# imported so the user can see what was filtered out.

_TARGET_HEADER_RE = re.compile(
    r"^\s*Target\s*Term\s*\(\s*([A-Za-z]{2,3}[-_][A-Za-z]{2,4}|[A-Za-z]{2,3})\s*\)\s*$",
    re.IGNORECASE,
)
_FILENAME_LOCALE_RE = re.compile(
    r"[_\-]([A-Za-z]{2,3}[-_][A-Za-z]{2,4})(?:[_\-]|\.)",
)
_SOURCE_HEADER_TOKENS = ("source term", "source", "english", "en", "term (en)")


def _locale_from_filename(path: str) -> str:
    base = os.path.basename(path)
    m = _FILENAME_LOCALE_RE.search(base)
    return m.group(1) if m else ""


def _xlsx_norm_header(s: Any) -> str:
    return ("" if s is None else str(s)).strip().lower()


def import_glossary_xlsx(
    path: str,
    *,
    sheet: Optional[str] = None,
    default_severity: str = "High",
) -> GlossaryImportResult:
    """Import a RingCentral-style glossary XLSX.

    Heuristics:
      - First sheet (or named ``sheet``) is read.
      - Header row is the first non-empty row.
      - One TermRule is emitted per (row, target-language column).
      - Locale is taken from ``Target Term (XX-YY)`` header; if the header
        omits the locale (single-target sheets sometimes do), fall back to
        a locale token in the filename.
      - ``Status`` cell: ``VALID`` (case-insensitive) → enabled, else disabled.
      - Notes column aggregates ``Notes``, ``Definition``, ``Context``,
        ``Part of Speech`` (only those that are present and non-empty).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        return GlossaryImportResult(
            [], 0, 0, 0,
            [(0, "openpyxl is required to import XLSX glossaries")],
        )

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return GlossaryImportResult([], 0, 0, 0, [(0, f"Open failed: {e}")])

    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return GlossaryImportResult([], 0, 0, 0, [(0, "Sheet is empty")])

    # Find first non-empty row as header
    header_idx = 0
    for idx, r in enumerate(rows):
        if r and any((c is not None and str(c).strip()) for c in r):
            header_idx = idx
            break
    headers = [_xlsx_norm_header(c) for c in rows[header_idx]]

    # Map columns
    col_term_id: Optional[int] = None
    col_source: Optional[int] = None
    col_status: Optional[int] = None
    col_notes: Optional[int] = None
    col_definition: Optional[int] = None
    col_context: Optional[int] = None
    col_pos: Optional[int] = None
    target_cols: List[Tuple[int, str]] = []  # (col_idx, locale_raw)

    fallback_locale = _locale_from_filename(path)

    for c, h in enumerate(headers):
        if not h:
            continue
        if h in ("term id", "id", "rule id"):
            col_term_id = c
        elif any(t in h for t in ("source term", "english", "term (en)")) or h == "source":
            col_source = c
        elif h == "status":
            col_status = c
        elif h == "notes":
            col_notes = c
        elif h == "definition":
            col_definition = c
        elif h == "context":
            col_context = c
        elif "part of speech" in h or h == "pos":
            col_pos = c
        else:
            m = _TARGET_HEADER_RE.match(rows[header_idx][c] or "")
            if m:
                target_cols.append((c, m.group(1)))
            elif h.startswith("target term"):
                # Header has no locale — use filename fallback
                if fallback_locale:
                    target_cols.append((c, fallback_locale))

    errors: List[Tuple[int, str]] = []
    if col_source is None:
        errors.append((0, "Missing 'Source Term (EN)' column"))
    if not target_cols:
        errors.append((0, "No 'Target Term (XX-YY)' column found and "
                          "filename does not contain a locale token"))
    if errors:
        return GlossaryImportResult([], 0, 0, 0, errors)

    rules: List[TermRule] = []
    seen_ids: set = set()
    total = 0
    imported = 0
    skipped = 0

    for r_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(c in (None, "") for c in row):
            continue
        total += 1

        def _cell(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return "" if v is None else str(v).strip()

        source_term = normalize_text(_cell(col_source))
        if not source_term:
            errors.append((r_idx, "Empty source term"))
            skipped += 1
            continue

        status_raw = _cell(col_status)
        enabled = (status_raw.upper() == "VALID") if status_raw else True

        notes_parts: List[str] = []
        for idx, label in (
            (col_notes,      ""),
            (col_definition, "def"),
            (col_context,    "ctx"),
            (col_pos,        "pos"),
        ):
            v = _cell(idx)
            if v:
                notes_parts.append(f"{label}: {v}" if label else v)
        notes = " | ".join(notes_parts)
        if status_raw and status_raw.upper() != "VALID":
            notes = (f"status={status_raw}" + (" | " + notes if notes else ""))

        term_id_raw = _cell(col_term_id)

        for col_idx, locale_raw in target_cols:
            approved = normalize_text(_cell(col_idx))
            if not approved:
                # row has no target for this language — silently skip (very common)
                continue

            sev = default_severity
            rule_id = term_id_raw
            if len(target_cols) > 1 and rule_id:
                rule_id = f"{rule_id}::{normalize_locale(locale_raw)}"
            if not rule_id:
                rule_id = make_rule_id(source_term, locale_raw, approved, [])
            if rule_id in seen_ids:
                errors.append((r_idx, f"Duplicate rule_id '{rule_id}'"))
                skipped += 1
                continue
            seen_ids.add(rule_id)

            rules.append(TermRule(
                rule_id=rule_id,
                source_term=source_term,
                locale=normalize_locale(locale_raw),
                approved_translation=approved,
                source_aliases=[],
                forbidden_translations=[],
                product_scope=[],
                severity=sev,
                case_sensitive=False,
                enabled=enabled,
                notes=notes,
                locale_display=locale_raw,
            ))
            imported += 1

    return GlossaryImportResult(rules, total, imported, skipped, errors)


def import_glossary(path: str) -> GlossaryImportResult:
    """Import a glossary file by extension. Supports .csv and .xlsx."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return import_glossary_xlsx(path)
    return import_glossary_csv(path)


def expand_locale_fallbacks(
    rules: Iterable[TermRule],
    *,
    fallbacks: Optional[Dict[str, str]] = None,
) -> List[TermRule]:
    """Synthesize sister-locale rules where a source locale has a rule
    but the fallback target locale does not (per term).

    Each synthetic rule:
      - inherits approved/forbidden/severity/etc. from the source rule
      - gets a distinct ``rule_id`` (suffix ``::fallback::<locale>``) so
        status persistence stays separable from the original rule
      - has a note like ``"fallback from zh-TW"`` for traceability
      - keeps the source rule's ``case_sensitive`` and ``enabled`` flags

    Existing explicit rules for the fallback locale are NEVER overridden;
    fallback only kicks in when the fallback locale has no rule for that
    term. Disabled source rules are skipped.
    """
    rules = list(rules)
    fb_map = LOCALE_FALLBACKS if fallbacks is None else fallbacks
    if not fb_map or not rules:
        return rules

    by_term: Dict[str, List[TermRule]] = {}
    for r in rules:
        by_term.setdefault(r.source_term, []).append(r)

    extras: List[TermRule] = []
    for term, term_rules in by_term.items():
        have_locales = {r.locale for r in term_rules}
        for fb_loc, src_loc in fb_map.items():
            if fb_loc in have_locales:
                continue  # explicit rule wins
            if src_loc not in have_locales:
                continue
            src_rule = next(r for r in term_rules if r.locale == src_loc)
            if not src_rule.enabled:
                continue
            tag = f"fallback from {src_rule.locale_display or src_rule.locale}"
            new_notes = (src_rule.notes + " | " + tag) if src_rule.notes else tag
            extras.append(TermRule(
                rule_id=f"{src_rule.rule_id}::fallback::{fb_loc}",
                source_term=src_rule.source_term,
                locale=fb_loc,
                approved_translation=src_rule.approved_translation,
                source_aliases=list(src_rule.source_aliases),
                forbidden_translations=list(src_rule.forbidden_translations),
                product_scope=list(src_rule.product_scope),
                severity=src_rule.severity,
                case_sensitive=src_rule.case_sensitive,
                enabled=True,
                notes=new_notes,
                locale_display=_canonical_locale_display(fb_loc),
            ))
    return rules + extras


def merge_rules(
    existing: Iterable[TermRule], incoming: Iterable[TermRule],
) -> List[TermRule]:
    """Merge two glossaries by ``rule_id``. Incoming wins on collision."""
    by_id: Dict[str, TermRule] = {r.rule_id: r for r in existing}
    for r in incoming:
        by_id[r.rule_id] = r
    # Stable order: keep existing order first, then append new ones
    seen: set = set()
    out: List[TermRule] = []
    for r in existing:
        if r.rule_id in by_id and r.rule_id not in seen:
            out.append(by_id[r.rule_id])
            seen.add(r.rule_id)
    for r in incoming:
        if r.rule_id not in seen:
            out.append(by_id[r.rule_id])
            seen.add(r.rule_id)
    return out


def build_quick_check_rules(
    source_term: str,
    locale_to_approved: Iterable[Tuple[str, str]],
    *,
    severity: str = "High",
    forbidden_per_locale: Optional[Dict[str, List[str]]] = None,
    aliases: Optional[List[str]] = None,
    notes: str = "ad-hoc",
    apply_locale_fallbacks: bool = True,
) -> List[TermRule]:
    """Construct in-memory rules for an ad-hoc Quick Check (no persistence).

    When ``apply_locale_fallbacks`` is true (default), sister-locale
    fallback rules are appended automatically — e.g. a user-supplied
    zh-TW pair also produces a synthetic zh-HK rule.
    """
    out: List[TermRule] = []
    fb = forbidden_per_locale or {}
    al = list(aliases or [])
    for locale_raw, approved in locale_to_approved:
        approved = normalize_text(approved)
        if not approved:
            continue
        nloc = normalize_locale(locale_raw)
        rid = make_rule_id(source_term, locale_raw, approved, [])
        out.append(TermRule(
            rule_id="QC-" + rid[2:],
            source_term=normalize_text(source_term),
            locale=nloc,
            approved_translation=approved,
            source_aliases=al,
            forbidden_translations=fb.get(nloc, []) + fb.get(locale_raw, []),
            product_scope=[],
            severity=severity,
            case_sensitive=False,
            enabled=True,
            notes=notes,
            locale_display=locale_raw,
        ))
    if apply_locale_fallbacks:
        out = expand_locale_fallbacks(out)
    return out


_LOCALE_LINE_RE = re.compile(
    r"^[A-Za-z]{2,3}(?:[-_][A-Za-z0-9]{2,4})?$"
)

_PASTE_HEADER_TOKENS = {
    "translations", "translation", "translated",
    "locale", "locales", "language", "languages", "lang", "lng",
    "approved", "approved translation", "approved translations",
    "source term", "source", "source text", "term", "terms",
    "target", "target term", "value", "values",
}


def reorganize_paste(text: str) -> str:
    """Heuristically reformat free-form pasted text into one
    ``locale<TAB>approved_translation`` pair per line.

    Handles the common copy-paste shapes seen in the wild:

      1. Already paired (tab / comma / arrow separator) → returned as-is
         (only blank lines + header words are stripped).
      2. Alternating ``locale\\nvalue\\nlocale\\nvalue`` (the typical
         single-column copy from Excel / Confluence / Word).
      3. Same as (2) but with a leading "Translations" / "Locale" header.
      4. Multi-line translation values: lines between two locale tokens
         are concatenated with spaces.

    If no locale tokens can be detected at all, the input is returned
    unchanged so the user can fix it manually.
    """
    if not text:
        return ""

    raw_lines = [ln.rstrip() for ln in text.splitlines()]
    lines: List[str] = []
    for ln in raw_lines:
        s = ln.strip()
        if not s:
            continue
        if s.lower() in _PASTE_HEADER_TOKENS:
            continue
        lines.append(s)
    if not lines:
        return ""

    # If most lines already contain a separator, treat as paired and just
    # normalize. Pairs already in the canonical form survive untouched.
    def _has_sep(ln: str) -> bool:
        if "\t" in ln or "→" in ln or "->" in ln:
            return True
        if "," in ln:
            head = ln.split(",", 1)[0].strip()
            if 1 <= len(head) <= 12 and _LOCALE_LINE_RE.match(head):
                return True
        # locale + whitespace + value (must start with a locale token)
        m = re.match(r"^([A-Za-z]{2,3}(?:[-_][A-Za-z0-9]{2,4})?)\s+\S", ln)
        if m:
            return True
        return False

    sep_count = sum(1 for ln in lines if _has_sep(ln))
    if sep_count >= max(1, len(lines) // 2):
        return "\n".join(lines)

    # Alternating-pair extraction.
    pairs: List[Tuple[str, str]] = []
    i = 0
    n = len(lines)
    while i < n:
        if _LOCALE_LINE_RE.match(lines[i]):
            locale = lines[i]
            i += 1
            value_parts: List[str] = []
            while i < n and not _LOCALE_LINE_RE.match(lines[i]):
                value_parts.append(lines[i])
                i += 1
            if value_parts:
                pairs.append((locale, " ".join(value_parts).strip()))
        else:
            # Skip stray non-locale leading line(s); user might have left
            # an extra header that didn't match our token set.
            i += 1

    if not pairs:
        return "\n".join(lines)

    return "\n".join(f"{loc}\t{val}" for loc, val in pairs)


def parse_clipboard_pairs(text: str) -> List[Tuple[str, str]]:
    """Parse a 2-column clipboard paste into (locale, approved) pairs.

    Accepts:
      - tab-separated (Excel default)
      - 2+ space-separated with locale token in the first ~12 chars
      - comma-separated
      - arrow ``->`` / ``→`` between the columns
    Empty lines and lines starting with ``#`` are ignored.
    """
    out: List[Tuple[str, str]] = []
    if not text:
        return out
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts: List[str]
        if "\t" in line:
            parts = [p.strip() for p in line.split("\t", 1)]
        elif "→" in line:
            parts = [p.strip() for p in line.split("→", 1)]
        elif "->" in line:
            parts = [p.strip() for p in line.split("->", 1)]
        elif "," in line and len(line.split(",", 1)[0]) <= 12:
            parts = [p.strip() for p in line.split(",", 1)]
        else:
            # Fall back to first run of whitespace, but only if locale token at
            # the very start (e.g. "fr_FR  Réceptionniste IA").
            m = re.match(r"^([A-Za-z]{2,3}[-_]?[A-Za-z]{0,4})\s+(.+)$", line)
            if not m:
                continue
            parts = [m.group(1).strip(), m.group(2).strip()]
        if len(parts) == 2 and parts[0] and parts[1]:
            out.append((parts[0], parts[1]))
    return out


def import_glossary_csv_text(text: str) -> GlossaryImportResult:
    if text and text.startswith("﻿"):
        text = text[1:]
    return _parse_glossary_csv(text)


def _parse_glossary_csv(text: str) -> GlossaryImportResult:
    rules: List[TermRule] = []
    errors: List[Tuple[int, str]] = []
    total = 0
    imported = 0
    skipped = 0

    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        return GlossaryImportResult([], 0, 0, 0, [(0, "CSV is empty")])

    fields = [f.strip() for f in reader.fieldnames]
    missing = [c for c in GLOSSARY_REQUIRED_COLS if c not in fields]
    if missing:
        return GlossaryImportResult(
            [], 0, 0, 0,
            [(0, f"Missing required columns: {', '.join(missing)}")],
        )

    seen_ids: set = set()

    for idx, row in enumerate(reader, start=2):  # row 1 = header
        total += 1
        # Strip whitespace from each cell key + value
        norm_row = {(k or "").strip(): (v if v is not None else "")
                    for k, v in row.items() if k is not None}

        source_term = normalize_text(norm_row.get("source_term", ""))
        locale_raw = (norm_row.get("locale", "") or "").strip()
        approved = normalize_text(norm_row.get("approved_translation", ""))

        if not source_term:
            errors.append((idx, "Empty source_term"))
            skipped += 1
            continue
        if not locale_raw:
            errors.append((idx, "Empty locale"))
            skipped += 1
            continue
        if not approved:
            errors.append((idx, "Empty approved_translation"))
            skipped += 1
            continue

        sev = normalize_severity(norm_row.get("severity"))
        if sev is None:
            errors.append((idx, f"Invalid severity '{norm_row.get('severity')}'; "
                                f"expected one of {', '.join(SEVERITIES)}"))
            skipped += 1
            continue

        product_scope = parse_pipe_list(norm_row.get("product_scope"))
        aliases = parse_pipe_list(norm_row.get("source_aliases"))
        forbidden = parse_pipe_list(norm_row.get("forbidden_translations"))
        case_sensitive = parse_bool(norm_row.get("case_sensitive"), default=False)
        enabled = parse_bool(norm_row.get("enabled"), default=True)
        notes = (norm_row.get("notes") or "").strip()

        rule_id = (norm_row.get("rule_id") or "").strip()
        if not rule_id:
            rule_id = make_rule_id(source_term, locale_raw, approved, product_scope)
        if rule_id in seen_ids:
            errors.append((idx, f"Duplicate rule_id '{rule_id}' in CSV"))
            skipped += 1
            continue
        seen_ids.add(rule_id)

        rule = TermRule(
            rule_id=rule_id,
            source_term=source_term,
            locale=normalize_locale(locale_raw),
            approved_translation=approved,
            source_aliases=aliases,
            forbidden_translations=forbidden,
            product_scope=product_scope,
            severity=sev,
            case_sensitive=case_sensitive,
            enabled=enabled,
            notes=notes,
            locale_display=locale_raw,
        )
        rules.append(rule)
        imported += 1

    return GlossaryImportResult(rules, total, imported, skipped, errors)


def export_glossary_csv(rules: Iterable[TermRule], path: str) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(GLOSSARY_ALL_COLS))
        w.writeheader()
        for r in rules:
            w.writerow({
                "rule_id": r.rule_id,
                "source_term": r.source_term,
                "locale": r.locale_display or r.locale,
                "approved_translation": r.approved_translation,
                "source_aliases": "|".join(r.source_aliases),
                "forbidden_translations": "|".join(r.forbidden_translations),
                "product_scope": "|".join(r.product_scope),
                "severity": r.severity,
                "case_sensitive": "true" if r.case_sensitive else "false",
                "enabled": "true" if r.enabled else "false",
                "notes": r.notes,
            })


def export_glossary_template(path: str) -> None:
    """Write a blank CSV template (header + one example row)."""
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(GLOSSARY_ALL_COLS))
        w.writeheader()
        w.writerow({
            "rule_id": "",
            "source_term": "AI receptionist",
            "locale": "fr_FR",
            "approved_translation": "Réceptionniste IA",
            "source_aliases": "AI receptionist",
            "forbidden_translations": "Standard IA",
            "product_scope": "Voice|AI",
            "severity": "Critical",
            "case_sensitive": "false",
            "enabled": "true",
            "notes": "Approved product term",
        })


def save_glossary(rules: Iterable[TermRule]) -> str:
    p = glossary_path()
    export_glossary_csv(rules, p)
    return p


def load_glossary() -> List[TermRule]:
    p = glossary_path()
    if not os.path.exists(p):
        return []
    res = import_glossary_csv(p)
    return res.rules


# ---------------------------------------------------------------------------
# Status persistence
# ---------------------------------------------------------------------------

class StatusStore:
    """JSON-backed map of issue_id -> status. Survives across scans."""

    def __init__(self, path: Optional[str] = None) -> None:
        self.path = path or statuses_path()
        self._lock = threading.Lock()
        self._map: Dict[str, str] = {}
        self._load()

    def _load(self) -> None:
        if not os.path.exists(self.path):
            return
        try:
            with open(self.path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                self._map = {str(k): str(v) for k, v in data.items()
                             if v in STATUSES}
        except Exception:
            self._map = {}

    def get(self, issue_id: str) -> str:
        with self._lock:
            return self._map.get(issue_id, DEFAULT_STATUS)

    def set(self, issue_id: str, status: str) -> None:
        if status not in STATUSES:
            raise ValueError(f"Invalid status: {status}")
        with self._lock:
            self._map[issue_id] = status
        self._save()

    def all(self) -> Dict[str, str]:
        with self._lock:
            return dict(self._map)

    def _save(self) -> None:
        ensure_data_dir()
        tmp = self.path + ".tmp"
        with self._lock:
            data = dict(self._map)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, sort_keys=True)
        os.replace(tmp, self.path)


def save_last_scan(summary: ScanSummary) -> str:
    p = last_scan_path()
    ensure_data_dir()
    tmp = p + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(summary.to_dict(), f, ensure_ascii=False, indent=2)
    os.replace(tmp, p)
    return p


def load_last_scan() -> Optional[ScanSummary]:
    p = last_scan_path()
    if not os.path.exists(p):
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            d = json.load(f)
        return ScanSummary(**d)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Scanner
# ---------------------------------------------------------------------------

def _contains(haystack: str, needle: str, case_sensitive: bool) -> bool:
    if not needle:
        return False
    if case_sensitive:
        return needle in haystack
    return casefold_text(needle) in casefold_text(haystack)


@dataclass
class _PreparedRule:
    """Internal: rule with pre-casefolded matching strings."""
    rule: TermRule
    source_terms_cmp: List[str]      # source_term + aliases (casefolded if !case_sensitive)
    approved_cmp: str
    forbidden_cmp: List[Tuple[str, str]]  # (display, cmp)
    case_sensitive: bool


def _prepare_rule(rule: TermRule) -> _PreparedRule:
    cs = rule.case_sensitive

    def _conv(s: str) -> str:
        return s if cs else s.casefold()

    terms = [rule.source_term] + [a for a in rule.source_aliases if a]
    return _PreparedRule(
        rule=rule,
        source_terms_cmp=[_conv(t) for t in terms if t],
        approved_cmp=_conv(rule.approved_translation),
        forbidden_cmp=[(f, _conv(f)) for f in rule.forbidden_translations if f],
        case_sensitive=cs,
    )


def _evaluate_prepared(
    pr: _PreparedRule, src_cmp: str, src_cs_cmp: str,
    tgt_cmp: str, tgt_cs_cmp: str, raw_target: str,
) -> Optional[Tuple[str, str, List[str]]]:
    """Match a prepared rule against pre-casefolded candidate strings.

    src_cmp / tgt_cmp = casefolded form (for case-insensitive rules).
    src_cs_cmp / tgt_cs_cmp = raw form (for case-sensitive rules).
    raw_target = normalized target text used for the empty check.
    """
    src = src_cs_cmp if pr.case_sensitive else src_cmp
    if not any(t in src for t in pr.source_terms_cmp):
        return None

    if not raw_target:
        sev = severity_max(pr.rule.severity, "High")
        return (ISSUE_MISSING_TARGET, sev, [])

    tgt = tgt_cs_cmp if pr.case_sensitive else tgt_cmp
    approved_ok = pr.approved_cmp in tgt
    forbidden_found = [disp for disp, cmp in pr.forbidden_cmp if cmp in tgt]

    has_forbidden = bool(forbidden_found)
    missing_required = not approved_ok

    if missing_required and has_forbidden:
        return (ISSUE_BOTH, "Critical", forbidden_found)
    if missing_required:
        return (ISSUE_REQUIRED_MISSING, pr.rule.severity, forbidden_found)
    if has_forbidden:
        sev = severity_max(pr.rule.severity, "Medium")
        return (ISSUE_FORBIDDEN_USED, sev, forbidden_found)
    return None


def scan_candidates(
    rules: Iterable[TermRule],
    candidates: Iterable[TranslationCandidate],
    status_store: Optional[StatusStore] = None,
    progress_cb: Optional[Callable[[int, int], None]] = None,
) -> Tuple[List[TerminologyIssue], ScanSummary]:
    """Run a deterministic terminology scan.

    progress_cb(scanned, total_candidates) — called periodically. The total
    is None when ``candidates`` is a generator (we don't materialize it).
    """
    rules = [r for r in rules if r.enabled]
    enabled_count = len(rules)

    # Pre-compile rules per locale (casefold once, not per candidate).
    prepared_by_locale: Dict[str, List[_PreparedRule]] = {}
    for r in rules:
        prepared_by_locale.setdefault(r.locale, []).append(_prepare_rule(r))

    # Materialize candidates lazily but allow progress reporting on a known total
    cand_list = candidates if isinstance(candidates, list) else list(candidates)
    total = len(cand_list)
    progress_step = max(500, total // 50) if total else 1

    issues: List[TerminologyIssue] = []
    summary = ScanSummary()
    summary.glossary_enabled_rules = enabled_count

    affected_terms: set = set()
    affected_locales: set = set()
    affected_products: set = set()

    scanned = 0
    skipped_no_src = 0
    skipped_no_match = 0

    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    for cand in cand_list:
        scanned += 1
        if progress_cb is not None and scanned % progress_step == 0:
            try:
                progress_cb(scanned, total)
            except Exception:
                pass

        applicable = prepared_by_locale.get(cand.locale)
        if not applicable:
            skipped_no_match += 1
            continue
        src_norm = normalize_text(cand.source_text)
        if not src_norm:
            skipped_no_src += 1
            continue

        tgt_norm = normalize_text(cand.target_text)
        # Pre-casefold candidate strings ONCE per candidate.
        src_cf = src_norm.casefold()
        tgt_cf = tgt_norm.casefold()

        # Product-scope pre-filter
        cand_product = cand.product

        any_applied = False
        for pr in applicable:
            scope = pr.rule.product_scope
            if scope and cand_product not in scope:
                continue
            verdict = _evaluate_prepared(pr, src_cf, src_norm, tgt_cf, tgt_norm, tgt_norm)
            if verdict is None:
                continue
            any_applied = True
            issue_type, sev, forbidden_found = verdict
            rule = pr.rule
            iid = make_issue_id(rule.rule_id, cand.candidate_id, issue_type)
            status = (status_store.get(iid) if status_store else DEFAULT_STATUS)
            issues.append(TerminologyIssue(
                issue_id=iid,
                rule_id=rule.rule_id,
                candidate_id=cand.candidate_id,
                issue_type=issue_type,
                severity=sev,
                status=status,
                source_term=rule.source_term,
                locale=rule.locale,
                locale_display=cand.locale_display or rule.locale_display or rule.locale,
                expected=rule.approved_translation,
                actual=cand.target_text or "",
                source_text=cand.source_text,
                product=cand.product,
                source_kind=cand.source_kind,
                source_label=SOURCE_LABELS.get(cand.source_kind, cand.source_kind),
                key=cand.key,
                task_id=cand.task_id,
                mr_id=cand.mr_id,
                score=cand.score,
                error_category=cand.error_category,
                reference_url=cand.reference_url,
                forbidden_found=forbidden_found,
                rule_notes=rule.notes,
                last_seen=now_iso,
            ))
            affected_terms.add(rule.source_term)
            affected_locales.add(rule.locale)
            affected_products.add(cand.product)

        if not any_applied:
            # locale matched at least one rule but source didn't contain term
            skipped_no_match += 1

    if progress_cb is not None:
        try:
            progress_cb(scanned, total)
        except Exception:
            pass

    # Filter: drop Ignored issues from active list (FR-13: keep status, hide).
    active = [i for i in issues if i.status != "Ignored"]

    summary.total_active_issues = len(active)
    summary.scanned_candidates = scanned
    summary.skipped_missing_source = skipped_no_src
    summary.skipped_no_match = skipped_no_match
    summary.affected_terms = len(affected_terms)
    summary.affected_locales = len(affected_locales)
    summary.affected_products = len(affected_products)
    summary.last_scan_at = now_iso

    for i in active:
        if i.severity == "Critical":
            summary.critical += 1
        elif i.severity == "High":
            summary.high += 1
        elif i.severity == "Medium":
            summary.medium += 1
        elif i.severity == "Low":
            summary.low += 1

    return active, summary


# ---------------------------------------------------------------------------
# Adapter: build candidates from FullTranslationInventory
# ---------------------------------------------------------------------------

def _pick_source_text(prod_data: Dict[str, Dict[str, str]],
                      opus_id: str) -> Optional[str]:
    for ref in ENGLISH_REFERENCE_LOCALES:
        loc_map = prod_data.get(ref)
        if loc_map:
            v = loc_map.get(opus_id)
            if v:
                return v
        # also try normalized form
        for key, lm in prod_data.items():
            if normalize_locale(key) == normalize_locale(ref):
                v = lm.get(opus_id)
                if v:
                    return v
    return None


_FT_SOURCE_KIND_MAP = {
    "Legacy": SOURCE_KIND_LEGACY,
    "MR": SOURCE_KIND_MR,
    "Scan": SOURCE_KIND_SCAN,
    "File": SOURCE_KIND_FILE,
}


def candidates_from_full_inventory(inv: Any) -> List[TranslationCandidate]:
    """Convert a FullTranslationInventory into TranslationCandidate records.

    `inv` must expose ``data[product][locale][opus_id] = translated_text``
    and optionally ``sources[product][locale][opus_id] = meta``.

    English reference locales are skipped as targets (used as source text).
    """
    out: List[TranslationCandidate] = []
    data = getattr(inv, "data", {}) or {}
    sources = getattr(inv, "sources", {}) or {}

    eng_set = {normalize_locale(x) for x in ENGLISH_REFERENCE_LOCALES}

    for product, loc_map in data.items():
        for locale_raw, kv in loc_map.items():
            nloc = normalize_locale(locale_raw)
            if nloc in eng_set:
                continue
            for opus_id, target in kv.items():
                src_text = _pick_source_text(loc_map, opus_id) or ""
                meta = (((sources.get(product) or {}).get(locale_raw)) or {}).get(opus_id) or {}
                kind = _FT_SOURCE_KIND_MAP.get(meta.get("source") or "", SOURCE_KIND_UNKNOWN)
                cand_id = f"{product}::{nloc}::{opus_id}"
                out.append(TranslationCandidate(
                    candidate_id=cand_id,
                    product=product,
                    locale=nloc,
                    locale_display=locale_raw,
                    key=opus_id,
                    source_text=src_text,
                    target_text=target or "",
                    source_kind=kind,
                    task_id=str(meta.get("task_id") or ""),
                    mr_id=str(meta.get("merge_request_iid") or ""),
                    raw=meta,
                ))
    return out


# ---------------------------------------------------------------------------
# Filtering helpers (used by GUI; pure for testability)
# ---------------------------------------------------------------------------

def filter_issues(
    issues: Iterable[TerminologyIssue],
    *,
    search: str = "",
    severity: str = "",
    locale: str = "",
    product: str = "",
    source_kind: str = "",
    status: str = "",
) -> List[TerminologyIssue]:
    q = (search or "").strip().casefold()
    out = []
    for i in issues:
        if severity and i.severity != severity:
            continue
        if locale and i.locale != normalize_locale(locale):
            continue
        if product and i.product != product:
            continue
        if source_kind and i.source_kind != source_kind:
            continue
        if status and i.status != status:
            continue
        if q:
            hay = " ".join([
                i.source_term, i.source_text, i.expected, i.actual,
                i.product, i.key,
            ]).casefold()
            if q not in hay:
                continue
        out.append(i)
    return out


# ---------------------------------------------------------------------------
# Evidence export
# ---------------------------------------------------------------------------

ISSUE_FIELDS_FOR_EXPORT = (
    "issue_id", "severity", "status", "issue_type",
    "source_term", "locale", "locale_display",
    "product", "source_kind", "source_label",
    "key", "task_id", "mr_id",
    "source_text", "actual", "expected",
    "forbidden_found", "rule_notes", "score", "error_category",
    "reference_url", "last_seen",
)

_ISSUE_FIELD_LABELS = {
    "issue_id": "Issue ID",
    "severity": "Severity",
    "status": "Status",
    "issue_type": "Issue Type",
    "source_term": "Source Term",
    "locale": "Locale",
    "locale_display": "Locale (display)",
    "product": "Product",
    "source_kind": "Source Kind",
    "source_label": "Source",
    "key": "Key/Reference",
    "task_id": "Task ID",
    "mr_id": "MR ID",
    "source_text": "Source Text",
    "actual": "Actual Translation",
    "expected": "Expected Translation",
    "forbidden_found": "Forbidden Variants Found",
    "rule_notes": "Rule Notes",
    "score": "Score",
    "error_category": "Error Category",
    "reference_url": "Reference URL",
    "last_seen": "Last Seen",
}


def _value_for_export(issue: TerminologyIssue, field_name: str) -> str:
    v = getattr(issue, field_name, "")
    if v is None:
        return ""
    if isinstance(v, list):
        return "|".join(str(x) for x in v)
    return str(v)


def export_evidence_markdown(
    issues: List[TerminologyIssue],
    summary: ScanSummary,
    *,
    title: str = "Term Watchtower — Evidence Report",
    filters_applied: Optional[Dict[str, str]] = None,
    glossary_timestamp: str = "",
) -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines: List[str] = []
    lines.append(f"# {title}")
    lines.append("")
    lines.append(f"- Generated: {now}")
    lines.append(f"- Scan: {summary.last_scan_at or '-'}")
    if glossary_timestamp:
        lines.append(f"- Glossary updated: {glossary_timestamp}")
    lines.append(f"- Issues in this report: {len(issues)}")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- Total active issues: {summary.total_active_issues}")
    lines.append(f"- Critical: {summary.critical}  High: {summary.high}  "
                 f"Medium: {summary.medium}  Low: {summary.low}")
    lines.append(f"- Affected terms: {summary.affected_terms}")
    lines.append(f"- Affected locales: {summary.affected_locales}")
    lines.append(f"- Affected products: {summary.affected_products}")
    lines.append(f"- Scanned candidates: {summary.scanned_candidates}")
    lines.append(f"- Skipped (missing source): {summary.skipped_missing_source}")
    lines.append(f"- Failed sources: "
                 f"{', '.join(summary.failed_sources) if summary.failed_sources else '-'}")
    lines.append("")
    if filters_applied:
        lines.append("## Filters")
        lines.append("")
        for k, v in filters_applied.items():
            if v:
                lines.append(f"- {k}: `{v}`")
        lines.append("")

    lines.append("## Issues")
    lines.append("")
    lines.append("| Severity | Term | Locale | Expected | Actual | Product | Source | Key |")
    lines.append("|---|---|---|---|---|---|---|---|")
    for i in issues:
        lines.append("| {sev} | {term} | {loc} | {exp} | {act} | {prod} | {src} | {key} |".format(
            sev=i.severity,
            term=_md_escape(i.source_term),
            loc=i.locale_display or i.locale,
            exp=_md_escape(i.expected),
            act=_md_escape(i.actual),
            prod=_md_escape(i.product),
            src=i.source_label or i.source_kind,
            key=_md_escape(i.key),
        ))
    lines.append("")

    lines.append("## Detailed Evidence")
    lines.append("")
    for i in issues:
        lines.append(f"### [{i.severity}] {i.source_term} → {i.locale_display or i.locale}")
        lines.append("")
        lines.append(f"- **Issue type:** `{i.issue_type}`")
        lines.append(f"- **Status:** {i.status}")
        lines.append(f"- **Product:** {i.product}")
        lines.append(f"- **Source:** {i.source_label} (`{i.source_kind}`)")
        lines.append(f"- **Key/Reference:** `{i.key}`")
        if i.task_id:
            lines.append(f"- **Task ID:** `{i.task_id}`")
        if i.mr_id:
            lines.append(f"- **MR ID:** `{i.mr_id}`")
        lines.append("")
        lines.append("**Source text:**")
        lines.append("")
        lines.append("> " + (i.source_text or "(empty)").replace("\n", "\n> "))
        lines.append("")
        lines.append(f"**Expected:** {i.expected or '(empty)'}")
        lines.append("")
        lines.append(f"**Actual:** {i.actual or '(empty)'}")
        lines.append("")
        if i.forbidden_found:
            lines.append(f"**Forbidden variants found:** {', '.join(i.forbidden_found)}")
            lines.append("")
        if i.rule_notes:
            lines.append(f"**Rule notes:** {i.rule_notes}")
            lines.append("")
        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def _md_escape(s: str) -> str:
    if not s:
        return ""
    return str(s).replace("|", "\\|").replace("\n", " ").replace("\r", " ")


def _html_escape(s: str) -> str:
    if s is None:
        return ""
    return (str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


_HTML_SEVERITY_COLORS = {
    "Critical": "#ef4444",
    "High":     "#f59e0b",
    "Medium":   "#3b82f6",
    "Low":      "#6b7280",
}


def export_evidence_html(
    issues: List[TerminologyIssue],
    summary: ScanSummary,
    *,
    title: str = "Term Watchtower — Evidence Report",
    filters_applied: Optional[Dict[str, str]] = None,
    glossary_timestamp: str = "",
) -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows_html: List[str] = []
    detail_html: List[str] = []
    for i in issues:
        sev_color = _HTML_SEVERITY_COLORS.get(i.severity, "#6b7280")
        rows_html.append(
            "<tr>"
            f"<td><span class='sev' style='background:{sev_color}'>{_html_escape(i.severity)}</span></td>"
            f"<td>{_html_escape(i.source_term)}</td>"
            f"<td>{_html_escape(i.locale_display or i.locale)}</td>"
            f"<td>{_html_escape(i.expected)}</td>"
            f"<td class='actual'>{_html_escape(i.actual)}</td>"
            f"<td>{_html_escape(i.product)}</td>"
            f"<td>{_html_escape(i.source_label)}</td>"
            f"<td><code>{_html_escape(i.key)}</code></td>"
            f"<td>{_html_escape(i.status)}</td>"
            "</tr>"
        )

        detail_html.append(
            "<div class='card'>"
            f"<div class='card-head'>"
            f"<span class='sev' style='background:{sev_color}'>{_html_escape(i.severity)}</span>"
            f" <strong>{_html_escape(i.source_term)}</strong> → "
            f" <em>{_html_escape(i.locale_display or i.locale)}</em>"
            f" <span class='muted'>· {_html_escape(i.issue_type)} · {_html_escape(i.status)}</span>"
            "</div>"
            f"<div class='kv'><span>Product</span><span>{_html_escape(i.product)}</span></div>"
            f"<div class='kv'><span>Source</span><span>{_html_escape(i.source_label)} "
            f"(<code>{_html_escape(i.source_kind)}</code>)</span></div>"
            f"<div class='kv'><span>Key</span><span><code>{_html_escape(i.key)}</code></span></div>"
            + (f"<div class='kv'><span>Task ID</span><span><code>{_html_escape(i.task_id)}</code></span></div>" if i.task_id else "")
            + (f"<div class='kv'><span>MR ID</span><span><code>{_html_escape(i.mr_id)}</code></span></div>" if i.mr_id else "")
            + f"<div class='block'><div class='label'>Source text</div><div class='src'>{_html_escape(i.source_text)}</div></div>"
            + f"<div class='diff'>"
              f"<div class='exp'><div class='label'>Expected</div><div>{_html_escape(i.expected)}</div></div>"
              f"<div class='act'><div class='label'>Actual</div><div>{_html_escape(i.actual)}</div></div>"
              f"</div>"
            + (f"<div class='kv'><span>Forbidden found</span><span>{_html_escape(', '.join(i.forbidden_found))}</span></div>" if i.forbidden_found else "")
            + (f"<div class='kv'><span>Rule notes</span><span>{_html_escape(i.rule_notes)}</span></div>" if i.rule_notes else "")
            + "</div>"
        )

    filters_html = ""
    if filters_applied:
        items = "".join(
            f"<li><code>{_html_escape(k)}</code>: {_html_escape(v)}</li>"
            for k, v in filters_applied.items() if v
        )
        if items:
            filters_html = f"<ul class='filters'>{items}</ul>"

    sources_failed = (
        f"<li>Failed sources: {_html_escape(', '.join(summary.failed_sources))}</li>"
        if summary.failed_sources else ""
    )
    glossary_line = (
        f"<li>Glossary updated: {_html_escape(glossary_timestamp)}</li>"
        if glossary_timestamp else ""
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>{_html_escape(title)}</title>
<style>
  :root {{
    --bg:#0f172a; --card:#16213e; --fg:#e5e7eb; --muted:#94a3b8;
    --border:#2a2a4a; --accent:#e94560;
  }}
  body {{ background:var(--bg); color:var(--fg); font-family:-apple-system,Segoe UI,Helvetica,Arial,sans-serif; margin:0; padding:32px; }}
  h1 {{ margin:0 0 8px; }}
  h2 {{ margin-top:32px; border-bottom:1px solid var(--border); padding-bottom:6px; }}
  ul {{ line-height:1.6; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th, td {{ padding:8px 10px; border-bottom:1px solid var(--border); text-align:left; vertical-align:top; }}
  th {{ background:#0a0a1a; position:sticky; top:0; }}
  td.actual {{ color:#fda4af; }}
  .sev {{ display:inline-block; padding:2px 8px; border-radius:10px; color:#fff; font-size:11px; font-weight:600; }}
  .filters code {{ background:#0a0a1a; padding:2px 6px; border-radius:4px; }}
  .card {{ background:var(--card); border:1px solid var(--border); border-radius:8px; padding:14px 16px; margin:12px 0; }}
  .card-head {{ font-size:14px; margin-bottom:8px; }}
  .muted {{ color:var(--muted); }}
  .kv {{ display:grid; grid-template-columns:160px 1fr; gap:8px; padding:2px 0; font-size:13px; }}
  .kv span:first-child {{ color:var(--muted); }}
  .block {{ margin-top:8px; }}
  .block .label {{ color:var(--muted); font-size:12px; margin-bottom:4px; }}
  .src {{ background:#0a0a1a; border-radius:6px; padding:8px 10px; font-size:13px; }}
  .diff {{ display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-top:8px; }}
  .diff .label {{ color:var(--muted); font-size:12px; margin-bottom:4px; }}
  .exp {{ background:#10271b; border-left:3px solid #10b981; padding:8px 10px; border-radius:4px; }}
  .act {{ background:#3b1115; border-left:3px solid var(--accent); padding:8px 10px; border-radius:4px; }}
  code {{ background:#0a0a1a; padding:1px 6px; border-radius:3px; font-size:12px; }}
</style>
</head>
<body>
<h1>{_html_escape(title)}</h1>
<ul>
  <li>Generated: {_html_escape(now)}</li>
  <li>Scan: {_html_escape(summary.last_scan_at or '-')}</li>
  {glossary_line}
  <li>Issues in report: {len(issues)}</li>
  <li>Total active issues at scan: {summary.total_active_issues}
      (Critical {summary.critical} · High {summary.high}
       · Medium {summary.medium} · Low {summary.low})</li>
  <li>Scanned candidates: {summary.scanned_candidates},
      skipped missing source: {summary.skipped_missing_source}</li>
  {sources_failed}
</ul>
{filters_html}

<h2>Issues</h2>
<table>
  <thead>
    <tr>
      <th>Severity</th><th>Term</th><th>Locale</th>
      <th>Expected</th><th>Actual</th><th>Product</th>
      <th>Source</th><th>Key</th><th>Status</th>
    </tr>
  </thead>
  <tbody>{''.join(rows_html)}</tbody>
</table>

<h2>Detailed Evidence</h2>
{''.join(detail_html)}
</body>
</html>
"""


def export_evidence_xlsx(
    issues: List[TerminologyIssue],
    summary: ScanSummary,
    out_path: str,
    *,
    filters_applied: Optional[Dict[str, str]] = None,
) -> str:
    """Write evidence report to .xlsx. Requires openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel export") from e

    wb = Workbook()

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "Term Watchtower — Evidence Report"
    ws_sum["A1"].font = Font(size=14, bold=True)
    rows = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Scan time", summary.last_scan_at or "-"),
        ("Total active issues", summary.total_active_issues),
        ("Critical", summary.critical),
        ("High", summary.high),
        ("Medium", summary.medium),
        ("Low", summary.low),
        ("Affected terms", summary.affected_terms),
        ("Affected locales", summary.affected_locales),
        ("Affected products", summary.affected_products),
        ("Scanned candidates", summary.scanned_candidates),
        ("Skipped (missing source)", summary.skipped_missing_source),
        ("Skipped (no rule match)", summary.skipped_no_match),
        ("Failed sources", ", ".join(summary.failed_sources) or "-"),
        ("Issues in report", len(issues)),
    ]
    for r_idx, (k, v) in enumerate(rows, start=3):
        ws_sum.cell(row=r_idx, column=1, value=k).font = Font(bold=True)
        ws_sum.cell(row=r_idx, column=2, value=v)

    if filters_applied:
        start = 3 + len(rows) + 1
        ws_sum.cell(row=start, column=1, value="Filters").font = Font(bold=True)
        offset = 1
        for k, v in filters_applied.items():
            if not v:
                continue
            ws_sum.cell(row=start + offset, column=1, value=k)
            ws_sum.cell(row=start + offset, column=2, value=v)
            offset += 1

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 60

    # Sheet 2: Issues
    ws = wb.create_sheet("Issues")
    headers = [_ISSUE_FIELD_LABELS[f] for f in ISSUE_FIELDS_FOR_EXPORT]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    sev_fills = {
        "Critical": PatternFill("solid", fgColor="FCA5A5"),
        "High":     PatternFill("solid", fgColor="FCD34D"),
        "Medium":   PatternFill("solid", fgColor="93C5FD"),
        "Low":      PatternFill("solid", fgColor="D1D5DB"),
    }
    for r_idx, issue in enumerate(issues, start=2):
        for c_idx, fld in enumerate(ISSUE_FIELDS_FOR_EXPORT, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_value_for_export(issue, fld))
            if fld == "severity":
                cell.fill = sev_fills.get(issue.severity, PatternFill())
                cell.font = Font(bold=True)

    # Reasonable column widths
    widths = {
        "issue_id": 22, "severity": 10, "status": 10, "issue_type": 28,
        "source_term": 22, "locale": 10, "locale_display": 14,
        "product": 22, "source_kind": 14, "source_label": 16,
        "key": 32, "task_id": 22, "mr_id": 14,
        "source_text": 60, "actual": 40, "expected": 40,
        "forbidden_found": 28, "rule_notes": 30, "score": 8,
        "error_category": 18, "reference_url": 32, "last_seen": 22,
    }
    for c_idx, fld in enumerate(ISSUE_FIELDS_FOR_EXPORT, start=1):
        ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = widths.get(fld, 18)

    # Sheet 3: Locale × Term distribution
    ws_dist = wb.create_sheet("Distribution")
    ws_dist.cell(row=1, column=1, value="Locale \\ Term").font = Font(bold=True)
    terms = sorted({i.source_term for i in issues})
    locales = sorted({i.locale_display or i.locale for i in issues})
    for c, t in enumerate(terms, start=2):
        cell = ws_dist.cell(row=1, column=c, value=t)
        cell.font = Font(bold=True)
    for r, loc in enumerate(locales, start=2):
        ws_dist.cell(row=r, column=1, value=loc).font = Font(bold=True)
        for c, t in enumerate(terms, start=2):
            n = sum(1 for i in issues
                    if (i.locale_display or i.locale) == loc and i.source_term == t)
            if n:
                ws_dist.cell(row=r, column=c, value=n)

    wb.save(out_path)
    return out_path


def write_evidence_html(
    issues: List[TerminologyIssue],
    summary: ScanSummary,
    out_path: str,
    *,
    title: str = "Term Watchtower — Evidence Report",
    filters_applied: Optional[Dict[str, str]] = None,
    glossary_timestamp: str = "",
) -> str:
    html = export_evidence_html(
        issues, summary,
        title=title, filters_applied=filters_applied,
        glossary_timestamp=glossary_timestamp,
    )
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    return out_path


def write_evidence_markdown(
    issues: List[TerminologyIssue],
    summary: ScanSummary,
    out_path: str,
    *,
    title: str = "Term Watchtower — Evidence Report",
    filters_applied: Optional[Dict[str, str]] = None,
    glossary_timestamp: str = "",
) -> str:
    md = export_evidence_markdown(
        issues, summary,
        title=title, filters_applied=filters_applied,
        glossary_timestamp=glossary_timestamp,
    )
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(md)
    return out_path


def issue_summary_text(i: TerminologyIssue) -> str:
    """Plain-text one-issue summary for clipboard."""
    parts = [
        f"[{i.severity}] {i.source_term} ({i.locale_display or i.locale})",
        f"Issue: {i.issue_type}  Status: {i.status}",
        f"Product: {i.product}  Source: {i.source_label}  Key: {i.key}",
        f"Expected: {i.expected}",
        f"Actual:   {i.actual}",
        f"Source text: {i.source_text}",
    ]
    if i.forbidden_found:
        parts.append(f"Forbidden found: {', '.join(i.forbidden_found)}")
    if i.rule_notes:
        parts.append(f"Rule notes: {i.rule_notes}")
    return "\n".join(parts)
