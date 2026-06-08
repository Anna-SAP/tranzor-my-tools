"""
Pre-Translation Coverage Check — engine (no GUI; unit-testable headless).

Problem this solves
-------------------
The l10n portal "Pull Release Monitoring" computes its Purchase delta as a
pure *source-language* diff between two commits on the monitored branch — it
is translation-agnostic, so a string Tranzor's MR pipeline already translated
STILL shows up in Purchase. Anna (the PM) pulls the trigger on manual File
translation, so before she does she needs a check: of the strings in this
Purchase delta, which ones has Tranzor already translated? Those can be
skipped (avoids redundant work and overwriting Tranzor's translations); the
rest genuinely need manual translation.

Input  : a Purchase delta XLSX exported from the l10n portal — sheet ``Data``
         with columns ``Key`` (= the full OPUS ID) and ``Value`` (= source EN
         text). An optional "_with_JIRA" variant adds Feature Ticket / JIRA
         Link / Source (file:line) columns, which we keep for display.
Join    : the XLSX ``Key`` is byte-identical to ``opus_index.opus_id``
         (``RingCentral.{alias}.{md5(path)}.{logicalKey}``), so coverage is a
         direct equality join against the local OPUS-ID index that
         ``opus_id_monitor`` syncs from the Tranzor API.

Verdict (conservative — only declare GREEN when high-confidence, because a
wrong "skip" ships untranslated text, while a wrong "needs manual" only costs
a redundant translation):

    GREEN  可跳过  : opus_id covered by an MR (source_kind='mr') for ALL needed
                     target languages with non-empty text, that MR is MERGED,
                     and no known score is below threshold.
    AMBER  需复核  : has coverage but with a reservation — only scan/file (not
                     MR) coverage / some needed languages missing / MR not
                     merged (or unconfirmable) / a known score below threshold.
    RED    待人工  : opus_id absent from the index — no Tranzor coverage at all.

The expensive per-MR (GitLab) and per-task (Tranzor results) lookups run ONLY
for rows that already pass the cheap local checks, so a typical drop costs a
handful of API calls, not one per string.
"""
from __future__ import annotations

import os

import openpyxl

# Verdict codes (stable identifiers; the GUI maps them to localized labels).
GREEN = "green"   # 可跳过
AMBER = "amber"   # 需复核
RED = "red"       # 待人工

DEFAULT_SCORE_THRESHOLD = 98  # mirrors quality_overview.DEFAULT_THRESHOLD

# Header aliases (lower-cased, trimmed) → logical field. Tolerant of small
# naming drifts in the l10n portal export.
_ID_HEADERS = {"key", "opus id", "opus_id", "opusid", "id"}
_SRC_HEADERS = {"value", "source", "source text", "en", "en-us", "source (en)"}
_TICKET_HEADERS = {"feature ticket", "ticket", "jira"}
_JIRA_HEADERS = {"jira link / commit", "jira link", "link", "jira link/commit"}
_PATH_HEADERS = {"source (file:line)", "source file", "file", "path", "source path"}


# ---------------------------------------------------------------------------
# 1. Parse the Purchase delta XLSX
# ---------------------------------------------------------------------------
def parse_delta_xlsx(path: str) -> list[dict]:
    """Read a Purchase delta XLSX into rows.

    Returns ``[{opus_id, source_text, feature_ticket, jira_link, source_path},
    ...]`` for every data row that has a non-blank Key. Prefers a sheet named
    ``Data`` but falls back to the first sheet. Raises ``ValueError`` if no
    recognizable Key column is found.
    """
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f"Delta file not found: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return []

    header = rows[0]
    col = {}  # logical field -> column index
    for idx, name in enumerate(header):
        if name is None:
            continue
        h = str(name).strip().lower()
        if h in _ID_HEADERS and "id" not in col:
            col["id"] = idx
        elif h in _SRC_HEADERS and "src" not in col:
            col["src"] = idx
        elif h in _TICKET_HEADERS and "ticket" not in col:
            col["ticket"] = idx
        elif h in _JIRA_HEADERS and "jira" not in col:
            col["jira"] = idx
        elif h in _PATH_HEADERS and "path" not in col:
            col["path"] = idx

    if "id" not in col:
        raise ValueError(
            "Could not find an OPUS ID column (expected a 'Key' header). "
            f"Saw headers: {[str(h) for h in header]}")

    def _cell(row, field):
        i = col.get(field)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else str(v).strip()

    out = []
    seen = set()
    for row in rows[1:]:
        opus_id = _cell(row, "id")
        if not opus_id or opus_id in seen:
            continue
        seen.add(opus_id)
        out.append({
            "opus_id": opus_id,
            "source_text": _cell(row, "src"),
            "feature_ticket": _cell(row, "ticket"),
            "jira_link": _cell(row, "jira"),
            "source_path": _cell(row, "path"),
        })
    return out


# ---------------------------------------------------------------------------
# 2. Default dependency providers (live data). Injected for tests.
# ---------------------------------------------------------------------------
def _default_coverage_fn(db_path=None):
    import opus_id_monitor as om

    def _fn(opus_ids):
        return om.get_coverage_rows(opus_ids, db_path=db_path)

    return _fn


def _default_mr_state_fn():
    """Return ``project_id, mr_iid -> state`` ('merged'/'opened'/... or None).

    None means "could not confirm" (no GitLab token, API error, or missing
    coordinates) — the classifier treats that conservatively as not-merged.
    """
    import gitlab_client

    client = gitlab_client.GitLabClient()
    if not client.has_token():
        return lambda project_id, mr_iid: None

    cache: dict = {}

    def _fn(project_id, mr_iid):
        if project_id in (None, "") or mr_iid in (None, ""):
            return None
        try:
            key = (str(project_id), int(mr_iid))
        except (TypeError, ValueError):
            return None
        if key in cache:
            return cache[key]
        try:
            mr = client.get_merge_request(project_id, mr_iid)
            state = (mr or {}).get("state")
        except Exception:
            state = None
        cache[key] = state
        return state

    return _fn


def _default_score_fn():
    """Return ``task_id -> {(opus_id, target_language): final_score}``."""
    import export_mr_pipeline as mr_api

    cache: dict = {}

    def _fn(task_id):
        if not task_id:
            return {}
        if task_id in cache:
            return cache[task_id]
        m: dict = {}
        try:
            res = mr_api.fetch_mr_results(task_id)
            for t in res.get("translations", []) or []:
                oid = t.get("opus_id")
                lg = t.get("target_language")
                sc = t.get("final_score")
                if oid and lg and sc is not None:
                    m[(oid, lg)] = sc
        except Exception:
            pass
        cache[task_id] = m
        return m

    return _fn


# ---------------------------------------------------------------------------
# 3. The classifier
# ---------------------------------------------------------------------------
def _non_empty(rec):
    return bool((rec.get("translated_text") or "").strip())


def infer_expected_langs(coverage: dict) -> list[str]:
    """Best-effort "languages this release ships" = the union of languages
    that have any translation across the matched opus_ids. A pragmatic default
    when the caller doesn't pass an explicit language set; the UI lets the PM
    override it. (Blind spot: a language Tranzor missed for *every* string in
    the delta won't appear here — documented; override to be exhaustive.)"""
    langs = set()
    for recs in coverage.values():
        for rec in recs:
            if _non_empty(rec):
                lg = rec.get("target_language")
                if lg:
                    langs.add(lg)
    return sorted(langs)


def check_delta(
    rows,
    expected_langs=None,
    *,
    db_path=None,
    coverage_fn=None,
    mr_state_fn=None,
    score_fn=None,
    accept_kinds=("mr",),
    score_threshold=DEFAULT_SCORE_THRESHOLD,
    require_merged=True,
    block_on_unknown_score=False,
):
    """Classify each delta row into GREEN / AMBER / RED.

    Parameters mirror the "strict 🟢" policy the PM chose: a row is GREEN only
    when every needed language is MR-covered, the MR is merged
    (``require_merged``), and no *known* score is below ``score_threshold``.
    A missing score does not by itself block GREEN unless
    ``block_on_unknown_score`` is set (TM/ICE matches legitimately carry no
    GMG score). All three dependency callables are injectable for testing.
    """
    if coverage_fn is None:
        coverage_fn = _default_coverage_fn(db_path)

    opus_ids = [r["opus_id"] for r in rows]
    coverage = coverage_fn(opus_ids) or {}

    if expected_langs:
        expected = set(expected_langs)
    else:
        expected = set(infer_expected_langs(coverage))
    accept = set(accept_kinds)

    # Lazily build the expensive providers only if a row reaches the gate.
    _holder = {"mr": mr_state_fn, "score": score_fn}

    def mr_fn(pid, iid):
        if _holder["mr"] is None:
            _holder["mr"] = _default_mr_state_fn()
        return _holder["mr"](pid, iid)

    def score_fn_(tid):
        if _holder["score"] is None:
            _holder["score"] = _default_score_fn()
        return _holder["score"](tid)

    results = []
    counts = {GREEN: 0, AMBER: 0, RED: 0}

    for r in rows:
        oid = r["opus_id"]
        recs = coverage.get(oid, [])
        res = {
            "opus_id": oid,
            "source_text": r.get("source_text", ""),
            "feature_ticket": r.get("feature_ticket", ""),
            "jira_link": r.get("jira_link", ""),
            "source_path": r.get("source_path", ""),
            "needed_langs": sorted(expected),
            "covered_langs": [],
            "missing_langs": sorted(expected),
            "mr_iid": None,
            "release": None,
            "mr_state": None,
            "min_score": None,
            "source_kind": None,
            "verdict": RED,
            "reason": "no_coverage",
        }

        if not recs:
            counts[RED] += 1
            results.append(res)
            continue

        live = [x for x in recs if _non_empty(x)]
        any_langs = sorted({x.get("target_language") for x in live})
        pipeline_recs = [x for x in live if x.get("source_kind") in accept]
        pipeline_langs = sorted({x.get("target_language") for x in pipeline_recs})
        res["covered_langs"] = pipeline_langs or any_langs
        res["source_kind"] = "+".join(
            sorted({x.get("source_kind") for x in live if x.get("source_kind")})) or None
        res["release"] = next((x.get("release") for x in pipeline_recs if x.get("release")),
                              recs[0].get("release"))

        # Covered, but not by an accepted pipeline kind (e.g. only File/Scan
        # while the policy accepts MR only) → review, don't auto-skip.
        if not pipeline_recs:
            res["verdict"] = AMBER
            res["reason"] = "non_mr_coverage"
            counts[AMBER] += 1
            results.append(res)
            continue

        missing = sorted(expected - set(pipeline_langs))
        res["missing_langs"] = missing
        if missing:
            res["verdict"] = AMBER
            res["reason"] = "langs_missing"
            counts[AMBER] += 1
            results.append(res)
            continue

        # MR-kind sub-records carry the merge state + scores (File/Scan don't).
        mr_recs = [x for x in pipeline_recs if x.get("source_kind") == "mr"]
        mr_iids = sorted({x.get("mr_iid") for x in mr_recs if x.get("mr_iid")})
        task_ids = sorted({x.get("task_id") for x in mr_recs if x.get("task_id")})
        proj = next((x.get("project_id") for x in mr_recs if x.get("project_id")), None)
        res["mr_iid"] = mr_iids[0] if mr_iids else None

        # All needed languages covered → gate on merge state (MR only) + score.
        if require_merged and mr_recs:
            states = [mr_fn(proj, iid) for iid in mr_iids] if mr_iids else [None]
            res["mr_state"] = ",".join(str(s) for s in states)
            if not all(s == "merged" for s in states):
                res["verdict"] = AMBER
                res["reason"] = "mr_not_merged"
                counts[AMBER] += 1
                results.append(res)
                continue

        score_map = {}
        for tid in task_ids:
            score_map.update(score_fn_(tid))
        known = [score_map[(oid, lg)] for lg in expected if (oid, lg) in score_map]
        unknown = len(known) < len(expected)
        res["min_score"] = min(known) if known else None
        if known and min(known) < score_threshold:
            res["verdict"] = AMBER
            res["reason"] = "score_below"
            counts[AMBER] += 1
            results.append(res)
            continue
        if unknown and block_on_unknown_score:
            res["verdict"] = AMBER
            res["reason"] = "score_unknown"
            counts[AMBER] += 1
            results.append(res)
            continue

        res["verdict"] = GREEN
        res["reason"] = "ok"
        counts[GREEN] += 1
        results.append(res)

    total = len(rows)
    summary = {
        "total": total,
        "green": counts[GREEN],
        "amber": counts[AMBER],
        "red": counts[RED],
        "skip_pct": round(100.0 * counts[GREEN] / total, 1) if total else 0.0,
        "expected_langs": sorted(expected),
        "manual_needed": counts[AMBER] + counts[RED],
    }
    return {"rows": results, "summary": summary}


# ---------------------------------------------------------------------------
# 4. Export the "still needs manual" subset back to XLSX
# ---------------------------------------------------------------------------
def export_manual_subset(result: dict, path: str, include_amber: bool = True) -> int:
    """Write the RED (and, by default, AMBER) rows to an XLSX shaped like the
    input (Key/Value + why), so the PM can hand exactly the genuine gaps to
    manual File translation. Returns the number of rows written."""
    wanted = {RED, AMBER} if include_amber else {RED}
    rows = [r for r in result.get("rows", []) if r["verdict"] in wanted]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Key", "Value", "Verdict", "Reason", "Missing Langs",
               "Feature Ticket", "Source (file:line)"])
    for r in rows:
        ws.append([
            r["opus_id"], r.get("source_text", ""), r["verdict"], r.get("reason", ""),
            ", ".join(r.get("missing_langs", [])),
            r.get("feature_ticket", ""), r.get("source_path", ""),
        ])
    wb.save(path)
    return len(rows)
