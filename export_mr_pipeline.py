"""
MR Pipeline API 数据获取与导出模块
===================================
通过 Tranzor MR Pipeline API 获取 MR 触发的翻译任务数据，
并导出为 HTML / Excel 格式。

API 基础路径: /api/v1
"""

import difflib
import html as html_mod
import json
import os
import sys
import time
import webbrowser
from collections import OrderedDict
from datetime import date

try:
    import requests
except ImportError:
    requests = None

# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------
TRANZOR_URL = "http://tranzor-platform.int.rclabenv.com"
MR_API = f"{TRANZOR_URL}/api/v1"

# HTTP session & retry config
_session = requests.Session() if requests else None
MAX_RETRIES = 3


def _api_get(url, **kwargs):
    """带重试的 GET 请求"""
    if _session is None:
        raise RuntimeError("requests package not available")
    kwargs.setdefault("timeout", 30)
    for attempt in range(MAX_RETRIES):
        try:
            return _session.get(url, **kwargs)
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError) as e:
            if attempt < MAX_RETRIES - 1:
                wait = 2 ** attempt
                print(f"    ⚠ 请求超时，{wait}s 后重试 ({attempt+1}/{MAX_RETRIES})...")
                time.sleep(wait)
            else:
                raise


# ---------------------------------------------------------------------------
# 1) Dashboard filters — 供下拉框使用
# ---------------------------------------------------------------------------
def fetch_mr_filters():
    """GET /dashboard/filters → { project_ids: [...], releases: [...] }"""
    resp = _api_get(f"{MR_API}/dashboard/filters")
    resp.raise_for_status()
    data = resp.json()
    return {
        "project_ids": data.get("project_ids", []),
        "releases": data.get("releases", []),
    }


def fetch_languages():
    """从最近 completed task 的翻译结果中提取所有可用的 target language 列表"""
    try:
        _, tasks = fetch_mr_tasks(status="completed", limit=10)
        langs = set()
        for t in tasks:
            tid = t.get("task_id")
            if not tid:
                continue
            try:
                results = fetch_mr_results(tid)
                for tr in results.get("translations", []):
                    lang = tr.get("target_language", "")
                    if lang:
                        langs.add(lang)
                if langs:
                    break
            except Exception:
                continue
        return sorted(langs)
    except Exception:
        return []


# ---------------------------------------------------------------------------
# 2) 任务列表
# ---------------------------------------------------------------------------
def fetch_mr_tasks(project_id=None, release=None, status=None,
                   limit=50, offset=0):
    """GET /tasks?... → { total, tasks: [...] }"""
    params = {"limit": limit, "offset": offset}
    if project_id:
        params["project_id"] = project_id
    if release:
        params["release"] = release
    if status:
        params["status"] = status

    resp = _api_get(f"{MR_API}/tasks", params=params)
    resp.raise_for_status()
    data = resp.json()
    return data.get("total", 0), data.get("tasks", [])


# ---------------------------------------------------------------------------
# 3) 任务详情
# ---------------------------------------------------------------------------
def fetch_mr_task_detail(task_id):
    """GET /tasks/{task_id}"""
    resp = _api_get(f"{MR_API}/tasks/{task_id}")
    resp.raise_for_status()
    return resp.json()


# ---------------------------------------------------------------------------
# 4) 翻译结果（含评估数据）
# ---------------------------------------------------------------------------
def fetch_mr_results(task_id, target_language=None,
                     min_score=None, max_score=None):
    """GET /tasks/{task_id}/results → { task_id, translations: [...], summary }"""
    params = {}
    if target_language:
        params["target_language"] = target_language
    if min_score is not None:
        params["min_score"] = min_score
    if max_score is not None:
        params["max_score"] = max_score

    resp = _api_get(f"{MR_API}/tasks/{task_id}/results", params=params)
    resp.raise_for_status()
    return resp.json()


# ---------------------------------------------------------------------------
# 4b) 聚合所有 completed 任务的翻译结果
# ---------------------------------------------------------------------------
MAX_WORKERS = 4


def collect_all_mr_results(progress_callback=None):
    """遍历所有 completed 状态的 MR Pipeline 任务，聚合翻译结果。

    Args:
        progress_callback: 可选回调 (msg: str) 用于输出进度日志

    Returns:
        与 fetch_mr_results 相同的结构: { "translations": [...], "summary": {} }
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    log = progress_callback or print

    # Step 1: 分页获取所有 completed 任务
    log("  正在获取 MR Pipeline 任务列表...")
    all_tasks = []
    offset = 0
    batch_size = 100
    while True:
        total, batch = fetch_mr_tasks(status="completed",
                                      limit=batch_size, offset=offset)
        all_tasks.extend(batch)
        if not batch or offset + batch_size >= total:
            break
        offset += batch_size
    log(f"  找到 {len(all_tasks)} 个已完成的 MR 任务")

    if not all_tasks:
        return {"translations": [], "summary": {}}

    # Step 2: 并发获取每个任务的翻译结果
    all_translations = []
    total_count = len(all_tasks)

    def _fetch_one(task_info):
        idx, task = task_info
        tid = task.get("task_id")
        if not tid:
            return []
        try:
            results = fetch_mr_results(tid)
            trs = results.get("translations", [])
            if trs:
                pid = task.get("project_id", "")
                mr_iid = task.get("merge_request_iid", "")
                log(f"  [{idx}/{total_count}] Task {tid[:8]}… "
                    f"(MR#{mr_iid}, {pid}) — {len(trs)} 条翻译")
            return trs
        except Exception as e:
            log(f"  ⚠ [{idx}/{total_count}] Task {tid[:8]}… 获取失败: {e}")
            return []

    task_infos = [(i + 1, t) for i, t in enumerate(all_tasks)]

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_fetch_one, info): info for info in task_infos}
        for f in as_completed(futures):
            all_translations.extend(f.result())

    log(f"\n  ✓ 共聚合 {len(all_translations)} 条翻译 (来自 {total_count} 个任务)")

    return {"translations": all_translations, "summary": {}}


# ---------------------------------------------------------------------------
# 4c) 检测同一 MR 下所有 task 之间的翻译变更
# ---------------------------------------------------------------------------
def detect_mr_changes(task_id, progress_callback=None):
    """检测给定 task 所属 MR 的全生命周期翻译变更。

    算法：找到同一 MR 的全部 completed task，按时间排序，
    逐对相邻 task 比较 (opus_id, target_language) 的 translated_text。

    Returns:
        list[dict]: 每条记录代表一次翻译文本变更，包含 prev_translated_text 等字段
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    log = progress_callback or print

    # Step 1: 获取当前 task 详情，提取 project_id 和 merge_request_iid
    log("  正在获取任务详情...")
    detail = fetch_mr_task_detail(task_id)
    project_id = detail.get("project_id")
    mr_iid = detail.get("merge_request_iid")
    if not project_id or not mr_iid:
        log("  ⚠ 无法获取 project_id 或 merge_request_iid")
        return []

    # Step 2: 获取同一 MR 的全部 completed task
    log(f"  正在获取 {project_id} MR#{mr_iid} 的所有任务...")
    all_tasks = []
    offset = 0
    batch_size = 100
    while True:
        total, batch = fetch_mr_tasks(project_id=project_id,
                                      status="completed",
                                      limit=batch_size, offset=offset)
        # 客户端按 mr_iid 过滤（API 不支持 mr_id 参数）
        for t in batch:
            if t.get("merge_request_iid") == mr_iid:
                all_tasks.append(t)
        if not batch or offset + batch_size >= total:
            break
        offset += batch_size

    # 按 created_at 升序排列
    all_tasks.sort(key=lambda t: t.get("created_at", ""))
    log(f"  找到 {len(all_tasks)} 个同 MR 的已完成任务")

    if len(all_tasks) < 2:
        log("  仅有 1 个任务，无法检测变更")
        return []

    # Step 3: 并发获取每个 task 的翻译结果
    log("  正在获取各任务的翻译结果...")
    task_results = {}  # task_id -> translations list
    total_count = len(all_tasks)

    def _fetch_one(task_info):
        idx, task = task_info
        tid = task.get("task_id")
        if not tid:
            return tid, []
        try:
            results = fetch_mr_results(tid)
            trs = results.get("translations", [])
            log(f"  [{idx}/{total_count}] Task {tid[:8]}… — {len(trs)} 条翻译")
            return tid, trs
        except Exception as e:
            log(f"  ⚠ [{idx}/{total_count}] Task {tid[:8]}… 获取失败: {e}")
            return tid, []

    task_infos = [(i + 1, t) for i, t in enumerate(all_tasks)]
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_fetch_one, info): info for info in task_infos}
        for f in as_completed(futures):
            tid, trs = f.result()
            if tid:
                task_results[tid] = trs

    # Step 4: 尝试从 Dashboard Cases API 获取 fixed_by_lead 信息
    log("  正在获取 Language Lead 修改记录...")
    fixed_by_map = {}  # (opus_id, target_language) -> fixed_by_lead
    try:
        cases_data = fetch_dashboard_cases(
            project_id=project_id, mr_limit=50)
        for mr_item in cases_data.get("mrs", []):
            if mr_item.get("mr_iid") != mr_iid:
                continue
            for case in mr_item.get("cases", []):
                fixer = case.get("fixed_by_lead")
                if fixer:
                    key = (case.get("opus_id", ""),
                           case.get("target_language", ""))
                    fixed_by_map[key] = fixer
    except Exception as e:
        log(f"  ⚠ 获取 Language Lead 信息失败: {e}")

    # MR-level metadata from task detail
    mr_meta = {
        "mr_link": detail.get("mr_link", ""),
        "project_id": project_id,
        "mr_iid": mr_iid,
        "release": detail.get("release", ""),
        "jira_ticket_id": detail.get("jira_ticket_id", ""),
    }

    # Step 5: 逐对相邻 task 做 diff
    changes = []
    for i in range(len(all_tasks) - 1):
        prev_task = all_tasks[i]
        curr_task = all_tasks[i + 1]
        prev_tid = prev_task.get("task_id")
        curr_tid = curr_task.get("task_id")

        prev_trs = task_results.get(prev_tid, [])
        curr_trs = task_results.get(curr_tid, [])

        # Build prev text map: (opus_id, target_language) -> translated_text
        prev_map = {}
        for t in prev_trs:
            key = (t.get("opus_id", ""), t.get("target_language", ""))
            prev_map[key] = t.get("translated_text", "")

        # Compare current with previous
        for t in curr_trs:
            key = (t.get("opus_id", ""), t.get("target_language", ""))
            curr_text = t.get("translated_text", "")
            prev_text = prev_map.get(key)

            if prev_text is not None and prev_text != curr_text:
                changes.append({
                    **mr_meta,
                    "opus_id": t.get("opus_id", ""),
                    "source_text": t.get("source_text", ""),
                    "target_language": t.get("target_language", ""),
                    "prev_translated_text": prev_text,
                    "translated_text": curr_text,
                    "prev_task_id": prev_tid,
                    "task_id": curr_tid,
                    "prev_task_created": prev_task.get("created_at", ""),
                    "task_created": curr_task.get("created_at", ""),
                    "fixed_by": fixed_by_map.get(key, ""),
                    "final_score": t.get("final_score"),
                    "error_category": t.get("error_category"),
                    "reason": t.get("reason"),
                    "iteration": t.get("iteration"),
                })

    log(f"\n  ✓ 检测到 {len(changes)} 条翻译变更 "
        f"(跨 {len(all_tasks)} 个任务)")
    return changes


# ---------------------------------------------------------------------------
# 5) Dashboard 概览
# ---------------------------------------------------------------------------
def fetch_dashboard_overview(project_id=None, release=None,
                             start_time=None, end_time=None):
    """GET /dashboard/overview"""
    params = {}
    if project_id:
        params["project_id"] = project_id
    if release:
        params["release"] = release
    if start_time:
        params["start_time"] = start_time
    if end_time:
        params["end_time"] = end_time

    resp = _api_get(f"{MR_API}/dashboard/overview", params=params)
    resp.raise_for_status()
    return resp.json()


# ---------------------------------------------------------------------------
# 6) Dashboard Cases（按 MR 分组）
# ---------------------------------------------------------------------------
def fetch_dashboard_cases(project_id=None, release=None, language=None,
                          min_score=None, max_score=None,
                          mr_limit=100, mr_offset=0):
    """GET /dashboard/cases"""
    params = {"mr_limit": mr_limit, "mr_offset": mr_offset}
    if project_id:
        params["project_id"] = project_id
    if release:
        params["release"] = release
    if language:
        params["language"] = language
    if min_score is not None:
        params["min_score"] = min_score
    if max_score is not None:
        params["max_score"] = max_score

    resp = _api_get(f"{MR_API}/dashboard/cases", params=params)
    resp.raise_for_status()
    return resp.json()


# ---------------------------------------------------------------------------
# 7) HTML 导出 — MR 翻译结果
# ---------------------------------------------------------------------------
def _word_diff_html(before, after):
    """HTML word-level diff: red strikethrough = deleted, green highlight = added."""
    before_words = before.split()
    after_words = after.split()
    sm = difflib.SequenceMatcher(None, before_words, after_words)
    parts = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            parts.append(html_mod.escape(" ".join(before_words[i1:i2])))
        elif tag == "replace":
            old = html_mod.escape(" ".join(before_words[i1:i2]))
            new = html_mod.escape(" ".join(after_words[j1:j2]))
            parts.append(f'<del style="background:#fdd;text-decoration:line-through;color:#c00;">{old}</del>')
            parts.append(f'<ins style="background:#dfd;text-decoration:none;color:#060;">{new}</ins>')
        elif tag == "delete":
            old = html_mod.escape(" ".join(before_words[i1:i2]))
            parts.append(f'<del style="background:#fdd;text-decoration:line-through;color:#c00;">{old}</del>')
        elif tag == "insert":
            new = html_mod.escape(" ".join(after_words[j1:j2]))
            parts.append(f'<ins style="background:#dfd;text-decoration:none;color:#060;">{new}</ins>')
    return " ".join(parts)


def _word_diff_text(before, after):
    """Plain text word-level diff: [-deleted] [+added]."""
    before_words = before.split()
    after_words = after.split()
    sm = difflib.SequenceMatcher(None, before_words, after_words)
    parts = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            parts.append(" ".join(before_words[i1:i2]))
        elif tag == "replace":
            parts.append("[-" + " ".join(before_words[i1:i2]) + "]")
            parts.append("[+" + " ".join(after_words[j1:j2]) + "]")
        elif tag == "delete":
            parts.append("[-" + " ".join(before_words[i1:i2]) + "]")
        elif tag == "insert":
            parts.append("[+" + " ".join(after_words[j1:j2]) + "]")
    return " ".join(parts)


def write_mr_html(results_data, filename, label):
    """生成 MR 翻译结果 HTML 报告（含 Filter + TMX 导出 + 评估列）"""
    translations = results_data.get("translations", [])
    summary = results_data.get("summary", {})
    task_id = results_data.get("task_id", "")

    # Detect if this is a "changes" export (translations have prev_translated_text)
    is_changes = any(t.get("prev_translated_text") is not None for t in translations)

    lang_colors = [
        ("#4472C4", "#e8eef7"), ("#E67E22", "#fdf2e6"),
        ("#27AE60", "#e8f5ee"), ("#8E44AD", "#f3eaf8"),
        ("#E74C3C", "#fceaea"), ("#16A085", "#e6f4f1"),
        ("#D4AC0D", "#faf6e4"), ("#2C3E50", "#eaecee"),
    ]

    # Group by language
    groups = OrderedDict()
    for t in translations:
        lang = t.get("target_language", "(unknown)")
        groups.setdefault(lang, []).append(t)

    # --- Build JSON data for JS ---
    js_rows = []
    all_langs = set()
    for lang_rows_list in groups.values():
        for r in lang_rows_list:
            row_data = {
                "source_text": r.get("source_text", ""),
                "translated_text": r.get("translated_text", ""),
                "language": r.get("target_language", ""),
                "string_key": r.get("opus_id", ""),
                "score": r.get("final_score"),
                "error_category": r.get("error_category") or "",
                "reason": r.get("reason") or "",
            }
            if is_changes:
                row_data["prev_translated_text"] = r.get("prev_translated_text", "")
                row_data["fixed_by"] = r.get("fixed_by", "")
                row_data["mr_link"] = r.get("mr_link", "")
                row_data["release"] = r.get("release", "")
                row_data["jira_ticket_id"] = r.get("jira_ticket_id", "")
            js_rows.append(row_data)
            all_langs.add(r.get("target_language", ""))
    rows_json = json.dumps(js_rows, ensure_ascii=False)
    langs_json = json.dumps(sorted(all_langs), ensure_ascii=False)

    # Build sections with checkboxes
    sections = []
    global_idx = 0
    for lang_i, (lang_name, lang_rows) in enumerate(groups.items()):
        color = lang_colors[lang_i % len(lang_colors)]
        header_bg, section_bg = color

        rows_html = []
        for r in lang_rows:
            score = r.get("final_score")
            score_str = f"{score}" if score is not None else "—"
            score_class = ""
            if score is not None:
                if score < 80:
                    score_class = ' style="color:#E74C3C;font-weight:bold"'
                elif score < 95:
                    score_class = ' style="color:#E67E22;font-weight:bold"'
                else:
                    score_class = ' style="color:#27AE60"'

            err_cat = r.get("error_category") or "—"
            reason = r.get("reason") or ""

            # Build diff columns for changes mode
            diff_cols = ""
            if is_changes:
                prev_text = r.get("prev_translated_text", "")
                curr_text = r.get("translated_text", "")
                diff_html = _word_diff_html(prev_text, curr_text)
                task_time = r.get("task_created", "")[:16].replace("T", " ")
                fixed_by = r.get("fixed_by", "")
                mr_link = r.get("mr_link", "")
                mr_link_html = (f'<a href="{html_mod.escape(mr_link)}" target="_blank">'
                                f'MR#{r.get("mr_iid", "")}</a>') if mr_link else "—"
                release = r.get("release", "") or "—"
                jira_id = r.get("jira_ticket_id", "")
                jira_html = (f'<a href="https://jira.ringcentral.com/browse/{html_mod.escape(jira_id)}"'
                             f' target="_blank">{html_mod.escape(jira_id)}</a>') if jira_id else "—"
                diff_cols = (
                    f'<td class="prev-translated">{html_mod.escape(prev_text)}</td>'
                    f'<td class="diff">{diff_html}</td>'
                    f'<td class="task-time">{html_mod.escape(task_time)}</td>'
                    f'<td class="fixed-by">{html_mod.escape(fixed_by) if fixed_by else "—"}</td>'
                    f'<td class="mr-link">{mr_link_html}</td>'
                    f'<td class="release">{html_mod.escape(release)}</td>'
                    f'<td class="jira">{jira_html}</td>'
                )

            rows_html.append(
                f'<tr>'
                f'<td class="cb-cell"><input type="checkbox" class="row-cb" data-idx="{global_idx}"></td>'
                f'<td class="num">{global_idx + 1}</td>'
                f'<td class="key">{html_mod.escape(r.get("opus_id", ""))}</td>'
                f'<td class="lang">{html_mod.escape(lang_name)}</td>'
                f'<td class="source">{html_mod.escape(r.get("source_text", ""))}</td>'
                f'{diff_cols}'
                f'<td class="translated">{html_mod.escape(r.get("translated_text", ""))}</td>'
                f'<td class="score"{score_class}>{score_str}</td>'
                f'<td class="err-cat">{html_mod.escape(err_cat)}</td>'
                f'<td class="reason">{html_mod.escape(reason)}</td>'
                f'</tr>'
            )
            global_idx += 1

        table_rows = "\n".join(rows_html)

        by_lang = summary.get("by_language", {})
        lang_info = by_lang.get(lang_name, {})
        lang_avg = lang_info.get("average_score", "—")
        lang_count = lang_info.get("count", len(lang_rows))

        sections.append(
            f'<div class="lang-section" style="background:{section_bg};border-left:4px solid {header_bg};border-radius:8px;padding:16px;margin-bottom:24px;">'
            f'<h2 style="color:{header_bg};margin-bottom:12px;font-size:17px;">'
            f'🌐 {html_mod.escape(lang_name)}'
            f'<span class="count" style="background:{header_bg};">{lang_count} entries</span>'
            f'<span class="avg-score" style="margin-left:12px;font-size:13px;color:#666;">Avg Score: {lang_avg}</span>'
            f'</h2>'
            f'<table><thead><tr style="background:{header_bg};">'
            f'<th class="cb-cell"><input type="checkbox" class="section-cb"></th>'
            f'<th>#</th><th>String Key</th><th>Lang</th>'
            f'<th>Source (en-US)</th>'
            + (f'<th>Previous Translation</th><th>Diff</th><th>Changed At</th>'
               f'<th>Fixed By</th><th>MR</th><th>Release</th><th>JIRA</th>' if is_changes else '')
            + f'<th>Translated</th>'
            f'<th>Score</th><th>Error Category</th><th>Reason</th>'
            f'</tr></thead><tbody>{table_rows}</tbody></table></div>'
        )

    sections_html = "\n".join(sections)

    # TOC
    toc_items = ""
    for lang_i, (lang_name, lang_rows) in enumerate(groups.items()):
        color = lang_colors[lang_i % len(lang_colors)]
        toc_items += (
            f'<span style="display:inline-block;background:{color[0]};'
            f'color:#fff;border-radius:16px;padding:4px 14px;margin:4px;'
            f'font-size:13px;">{html_mod.escape(lang_name)} ({len(lang_rows)})</span>'
        )

    total_items = summary.get("total_items", len(translations))

    page = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>MR Pipeline Translations - {html_mod.escape(label)}</title>
<style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family:-apple-system,"Segoe UI",Roboto,Arial,sans-serif; background:#f5f6fa; padding:24px; padding-top:72px; color:#333; }}
    h1 {{ font-size:20px; margin-bottom:4px; }}
    .meta {{ color:#666; font-size:14px; margin-bottom:8px; }}
    .summary-bar {{ background:#1e293b; color:#cbd5e1; border-radius:10px; padding:16px 20px; margin-bottom:20px; display:flex; gap:32px; align-items:center; }}
    .summary-item {{ text-align:center; }}
    .summary-value {{ font-size:24px; font-weight:bold; color:#fff; }}
    .summary-label {{ font-size:11px; color:#94a3b8; text-transform:uppercase; letter-spacing:.5px; }}
    .toc {{ margin-bottom:20px; }}
    .count {{ display:inline-block; color:#fff; border-radius:12px; padding:2px 10px; font-size:13px; margin-left:8px; }}
    table {{ border-collapse:collapse; width:100%; background:#fff; border-radius:6px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.08); font-size:13px; }}
    th {{ color:#fff; padding:10px 8px; text-align:left; white-space:nowrap; }}
    td {{ padding:8px; border-bottom:1px solid #e8e8e8; vertical-align:top; line-height:1.5; }}
    tr:hover {{ background:rgba(255,255,255,.6); }}
    .num {{ text-align:center; color:#999; min-width:30px; }}
    .lang {{ text-align:center; }}
    .key {{ font-family:monospace; font-size:12px; max-width:260px; word-break:break-all; color:#555; }}
    .source {{ background:#fff8e6; min-width:200px; }}
    .translated {{ background:#edf7ed; min-width:200px; }}
    .score {{ text-align:center; font-weight:600; min-width:60px; }}
    .err-cat {{ white-space:nowrap; font-size:12px; }}
    .reason {{ max-width:200px; font-size:12px; color:#666; }}

    /* --- Toolbar --- */
    .tmx-toolbar {{
        position:fixed; top:0; left:0; right:0; z-index:1000;
        background:#fff; border-bottom:1px solid #dde; padding:10px 24px;
        display:flex; align-items:center; gap:12px;
        box-shadow:0 2px 8px rgba(0,0,0,.08);
    }}
    .tmx-toolbar .btn {{
        border:none; border-radius:6px; padding:7px 16px; cursor:pointer;
        font-size:13px; font-weight:600; transition:background .15s, transform .1s;
    }}
    .tmx-toolbar .btn:active {{ transform:scale(.97); }}
    .btn-select {{ background:#eef1f5; color:#333; }}
    .btn-select:hover {{ background:#dde3ea; }}
    .btn-filter {{ background:#eef1f5; color:#333; }}
    .btn-filter:hover {{ background:#dde3ea; }}
    .btn-filter.active {{ background:#4472C4; color:#fff; }}
    .btn-export {{ background:#4472C4; color:#fff; }}
    .btn-export:hover {{ background:#3461b0; }}
    .btn-export:disabled {{ background:#a8b8d0; cursor:not-allowed; opacity:.7; }}
    .badge {{ background:#4472C4; color:#fff; border-radius:12px; padding:2px 10px; font-size:12px; font-weight:700; min-width:28px; text-align:center; }}
    .badge.zero {{ background:#bbb; }}
    .toolbar-sep {{ width:1px; height:24px; background:#dde; }}

    /* checkbox */
    .cb-cell {{ text-align:center; width:32px; min-width:32px; }}
    .cb-cell input[type="checkbox"] {{ width:15px; height:15px; cursor:pointer; accent-color:#4472C4; }}
    tr.row-selected {{ background:#e8eef7 !important; }}
    tr.row-hidden {{ display:none !important; }}

    /* --- Filter Panel --- */
    .filter-panel {{
        background:#1e293b; color:#cbd5e1; border-radius:10px;
        margin-bottom:20px; overflow:hidden;
        max-height:0; opacity:0; transition:max-height .35s ease, opacity .25s ease, margin .25s ease, padding .25s ease;
        padding:0 20px;
    }}
    .filter-panel.open {{ max-height:800px; opacity:1; padding:20px; }}
    .fp-row {{ display:flex; gap:16px; flex-wrap:wrap; margin-bottom:16px; }}
    .fp-row:last-child {{ margin-bottom:0; }}
    .fp-simple {{ display:flex; align-items:center; gap:8px; }}
    .fp-simple label {{ font-size:12px; font-weight:600; color:#94a3b8; text-transform:uppercase; letter-spacing:.5px; white-space:nowrap; }}
    .fp-simple input[type="text"] {{
        background:#0f172a; border:1px solid #334155; border-radius:6px;
        color:#e2e8f0; padding:6px 10px; font-size:13px; width:170px; outline:none;
    }}
    .fp-simple input[type="text"]:focus {{ border-color:#4472C4; }}
    .fp-simple select {{
        background:#0f172a; border:1px solid #334155; border-radius:6px;
        color:#e2e8f0; padding:6px 10px; font-size:13px; width:130px; outline:none; cursor:pointer;
    }}
    .tf-card {{
        background:#0f172a; border:1px solid #334155; border-radius:8px;
        padding:12px 14px; min-width:280px; flex:1;
    }}
    .tf-header {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:10px; }}
    .tf-title {{ font-size:12px; font-weight:700; color:#38bdf8; text-transform:uppercase; letter-spacing:.5px; }}
    .tf-logic {{ display:flex; gap:0; }}
    .tf-logic button {{
        padding:3px 12px; font-size:11px; font-weight:700; border:1px solid #334155;
        cursor:pointer; background:transparent; color:#64748b; transition:all .15s;
    }}
    .tf-logic button:first-child {{ border-radius:4px 0 0 4px; }}
    .tf-logic button:last-child {{ border-radius:0 4px 4px 0; border-left:0; }}
    .tf-logic button.active {{ background:#0ea5e9; color:#fff; border-color:#0ea5e9; }}
    .tf-input-row {{ display:flex; align-items:center; gap:8px; margin-bottom:8px; }}
    .tf-label {{ font-size:11px; font-weight:700; padding:3px 8px; border-radius:4px; text-align:center; min-width:36px; }}
    .tf-label.pos {{ background:#fff; color:#1e293b; }}
    .tf-label.neg {{ background:transparent; color:#f87171; border:1px solid #7f1d1d; }}
    .tf-input {{ flex:1; background:#1e293b; border:1px solid #334155; border-radius:6px; color:#e2e8f0; padding:6px 10px; font-size:13px; outline:none; }}
    .tf-input:focus {{ border-color:#4472C4; }}
    .tf-input.neg-input {{ border-color:#7f1d1d; }}
    .tf-input.neg-input:focus {{ border-color:#ef4444; }}
    .tf-opts {{ display:flex; flex-wrap:wrap; gap:6px 14px; margin-top:6px; }}
    .tf-opt {{ display:flex; align-items:center; gap:4px; font-size:11px; color:#94a3b8; cursor:pointer; user-select:none; }}
    .tf-opt input {{ accent-color:#4472C4; cursor:pointer; }}
    .fp-actions {{ display:flex; gap:10px; align-items:center; margin-left:auto; }}
    .fp-actions .btn {{ border:none; border-radius:6px; padding:7px 18px; cursor:pointer; font-size:13px; font-weight:600; }}
    .btn-apply {{ background:#0ea5e9; color:#fff; }}
    .btn-apply:hover {{ background:#0284c7; }}
    .btn-clear {{ background:#334155; color:#cbd5e1; }}
    .btn-clear:hover {{ background:#475569; }}
    .filter-info {{ font-size:12px; color:#94a3b8; }}
</style>
</head>
<body>

<!-- Toolbar -->
<div class="tmx-toolbar">
    <button class="btn btn-select" id="btnSelectAll" onclick="toggleSelectAll()">☑ Select All</button>
    <div class="toolbar-sep"></div>
    <span style="font-size:13px;color:#555;">Selected:</span>
    <span class="badge zero" id="selCount">0</span>
    <div class="toolbar-sep"></div>
    <button class="btn btn-export" id="btnExport" onclick="exportTMX()" disabled>📦 Export TMX</button>
    <span id="exportStatus" style="font-size:12px;color:#888;"></span>
    <div class="toolbar-sep"></div>
    <button class="btn btn-filter" id="btnFilterToggle" onclick="toggleFilterPanel()">🔍 Filters</button>
    <span class="filter-info" id="filterInfo"></span>
</div>

    <h1>MR Pipeline Translations <span class="count" style="background:#4472C4;">{total_items} entries</span></h1>
    <p class="meta">{html_mod.escape(label)}</p>

<!-- Filter Panel -->
<div class="filter-panel" id="filterPanel">
    <div class="fp-row">
        <div class="fp-simple">
            <label>Lang</label>
            <select id="fLang"><option value="">All</option></select>
        </div>
        <div class="fp-simple">
            <label>Score</label>
            <select id="fScore">
                <option value="">All</option>
                <option value="lt80">&lt; 80</option>
                <option value="lt95">&lt; 95</option>
                <option value="gte95">≥ 95</option>
                <option value="eq100">= 100</option>
            </select>
        </div>
        <div class="fp-actions">
            <button class="btn btn-apply" onclick="applyFilters()">▶ Apply</button>
            <button class="btn btn-clear" onclick="clearFilters()">✕ Clear</button>
        </div>
    </div>
    <div class="fp-row" id="tfRow">
        <!-- TextFilter cards rendered by JS -->
    </div>
</div>

    <div class="summary-bar">
        <div class="summary-item">
            <div class="summary-value">{total_items}</div>
            <div class="summary-label">Total Items</div>
        </div>
        <div class="summary-item">
            <div class="summary-value">{len(groups)}</div>
            <div class="summary-label">Languages</div>
        </div>
    </div>

    <div class="toc">{toc_items}</div>
    {sections_html}

<script src="https://cdn.jsdelivr.net/npm/jszip@3/dist/jszip.min.js"></script>
<script>
// ============================================================
// Row data & selection state
// ============================================================
const ROWS = {rows_json};
const ALL_LANGS = {langs_json};
let allSelected = false;

// Populate lang dropdown
(function() {{
    const sel = document.getElementById('fLang');
    ALL_LANGS.forEach(l => {{
        const opt = document.createElement('option');
        opt.value = l; opt.textContent = l;
        sel.appendChild(opt);
    }});
}})();

// ============================================================
// TextFilter card rendering
// ============================================================
const TF_FIELDS = [
    {{ id: 'stringKey',  label: 'STRING KEY',      dataKey: 'string_key' }},
    {{ id: 'source',     label: 'SOURCE (EN-US)',   dataKey: 'source_text' }},
    {{ id: 'translated', label: 'TRANSLATED TEXT',  dataKey: 'translated_text' }},
];

(function renderTFCards() {{
    const container = document.getElementById('tfRow');
    TF_FIELDS.forEach(f => {{
        container.innerHTML += `
        <div class="tf-card" data-field="${{f.id}}">
            <div class="tf-header">
                <span class="tf-title">${{f.label}}</span>
                <div class="tf-logic">
                    <button class="active" data-val="AND" onclick="toggleLogic(this)">AND</button>
                    <button data-val="OR" onclick="toggleLogic(this)">OR</button>
                </div>
            </div>
            <div class="tf-input-row">
                <span class="tf-label pos">Pos</span>
                <input class="tf-input" data-role="pos" placeholder="Positive keyword…">
            </div>
            <div class="tf-input-row">
                <span class="tf-label neg">Neg</span>
                <input class="tf-input neg-input" data-role="neg" placeholder="Negative keyword (Exclude)…">
            </div>
            <div class="tf-opts">
                <label class="tf-opt"><input type="checkbox" data-role="matchWhole"> Match whole</label>
                <label class="tf-opt"><input type="checkbox" data-role="posCaseSensitive"> Match case (Pos)</label>
                <label class="tf-opt"><input type="checkbox" data-role="posRegex"> Regex (Pos)</label>
                <label class="tf-opt"><input type="checkbox" data-role="negCaseSensitive"> Match case (Neg)</label>
                <label class="tf-opt"><input type="checkbox" data-role="negRegex"> Regex (Neg)</label>
            </div>
        </div>`;
    }});
}})();

function toggleLogic(btn) {{
    const group = btn.parentElement;
    group.querySelectorAll('button').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
}}

function toggleFilterPanel() {{
    const panel = document.getElementById('filterPanel');
    const btn = document.getElementById('btnFilterToggle');
    panel.classList.toggle('open');
    btn.classList.toggle('active');
}}

// ============================================================
// Filter engine
// ============================================================
function getTextFilterState(cardEl) {{
    const logicBtn = cardEl.querySelector('.tf-logic button.active');
    return {{
        logic: logicBtn ? logicBtn.dataset.val : 'AND',
        pos: cardEl.querySelector('[data-role="pos"]').value,
        neg: cardEl.querySelector('[data-role="neg"]').value,
        posCaseSensitive: cardEl.querySelector('[data-role="posCaseSensitive"]').checked,
        posRegex: cardEl.querySelector('[data-role="posRegex"]').checked,
        negCaseSensitive: cardEl.querySelector('[data-role="negCaseSensitive"]').checked,
        negRegex: cardEl.querySelector('[data-role="negRegex"]').checked,
        matchWhole: cardEl.querySelector('[data-role="matchWhole"]').checked,
    }};
}}

function testMatch(text, keyword, caseSensitive, isRegex, matchWhole) {{
    if (!keyword) return null;
    try {{
        let pattern;
        if (isRegex) {{
            pattern = new RegExp(keyword, caseSensitive ? '' : 'i');
        }} else {{
            const escaped = keyword.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&');
            const p = matchWhole ? ('\\\\b' + escaped + '\\\\b') : escaped;
            pattern = new RegExp(p, caseSensitive ? '' : 'i');
        }}
        return pattern.test(text);
    }} catch(e) {{
        return false;
    }}
}}

function evaluateTextFilter(text, tf) {{
    const posResult = testMatch(text, tf.pos, tf.posCaseSensitive, tf.posRegex, tf.matchWhole);
    const negResult = testMatch(text, tf.neg, tf.negCaseSensitive, tf.negRegex, tf.matchWhole);
    if (tf.logic === 'AND') {{
        const posPass = (posResult === null) ? true : posResult;
        const negPass = (negResult === null) ? true : !negResult;
        return posPass && negPass;
    }} else {{
        if (posResult === null && negResult === null) return true;
        const posPass = posResult === true;
        const negPass = (negResult !== null) && !negResult;
        return posPass || negPass;
    }}
}}

function applyFilters() {{
    const fLang = document.getElementById('fLang').value;
    const fScore = document.getElementById('fScore').value;

    const tfCards = document.querySelectorAll('.tf-card');
    const tfStates = {{}};
    tfCards.forEach(card => {{
        const field = card.dataset.field;
        tfStates[field] = getTextFilterState(card);
    }});

    const allRows = document.querySelectorAll('input.row-cb');
    let visibleCount = 0;
    let totalCount = allRows.length;

    allRows.forEach(cb => {{
        const idx = parseInt(cb.dataset.idx);
        const row = ROWS[idx];
        const tr = cb.closest('tr');
        let pass = true;

        if (fLang && pass) {{
            if (row.language !== fLang) pass = false;
        }}
        if (fScore && pass) {{
            const s = row.score;
            if (fScore === 'lt80')  {{ if (s === null || s >= 80)  pass = false; }}
            if (fScore === 'lt95')  {{ if (s === null || s >= 95)  pass = false; }}
            if (fScore === 'gte95') {{ if (s === null || s < 95)   pass = false; }}
            if (fScore === 'eq100') {{ if (s !== 100)              pass = false; }}
        }}

        if (pass) {{
            TF_FIELDS.forEach(f => {{
                if (!pass) return;
                const tf = tfStates[f.id];
                if (!tf || (!tf.pos && !tf.neg)) return;
                const text = row[f.dataKey] || '';
                if (!evaluateTextFilter(text, tf)) pass = false;
            }});
        }}

        if (pass) {{
            tr.classList.remove('row-hidden');
            visibleCount++;
        }} else {{
            tr.classList.add('row-hidden');
        }}
    }});

    const info = document.getElementById('filterInfo');
    if (visibleCount < totalCount) {{
        info.textContent = 'Showing ' + visibleCount + ' / ' + totalCount;
        info.style.color = '#e67e22';
    }} else {{
        info.textContent = '';
    }}

    document.querySelectorAll('.lang-section table').forEach(table => {{
        const visibleCbs = table.querySelectorAll('tbody tr:not(.row-hidden) input.row-cb');
        const checkedCbs = table.querySelectorAll('tbody tr:not(.row-hidden) input.row-cb:checked');
        const sectionCb = table.querySelector('input.section-cb');
        if (sectionCb) {{
            sectionCb.checked = visibleCbs.length > 0 && visibleCbs.length === checkedCbs.length;
        }}
    }});
    updateBadge();
}}

function clearFilters() {{
    document.getElementById('fLang').value = '';
    document.getElementById('fScore').value = '';

    document.querySelectorAll('.tf-card').forEach(card => {{
        card.querySelector('[data-role="pos"]').value = '';
        card.querySelector('[data-role="neg"]').value = '';
        card.querySelectorAll('.tf-opts input[type="checkbox"]').forEach(cb => cb.checked = false);
        const btns = card.querySelectorAll('.tf-logic button');
        btns.forEach(b => b.classList.remove('active'));
        btns[0].classList.add('active');
    }});

    document.querySelectorAll('tr.row-hidden').forEach(tr => tr.classList.remove('row-hidden'));
    document.getElementById('filterInfo').textContent = '';
    updateBadge();
}}

// ============================================================
// Selection helpers
// ============================================================
function getVisibleRowCheckboxes() {{
    return document.querySelectorAll('tr:not(.row-hidden) input.row-cb');
}}

function updateBadge() {{
    const n = document.querySelectorAll('input.row-cb:checked').length;
    const badge = document.getElementById('selCount');
    badge.textContent = n;
    badge.className = n ? 'badge' : 'badge zero';
    document.getElementById('btnExport').disabled = (n === 0);
}}

function toggleSelectAll() {{
    allSelected = !allSelected;
    getVisibleRowCheckboxes().forEach(cb => {{ cb.checked = allSelected; highlightRow(cb); }});
    document.querySelectorAll('input.section-cb').forEach(cb => cb.checked = allSelected);
    document.getElementById('btnSelectAll').textContent = allSelected ? '☐ Deselect All' : '☑ Select All';
    updateBadge();
}}

document.addEventListener('change', function(e) {{
    if (e.target.classList.contains('section-cb')) {{
        const tbody = e.target.closest('table').querySelector('tbody');
        tbody.querySelectorAll('tr:not(.row-hidden) input.row-cb').forEach(cb => {{
            cb.checked = e.target.checked;
            highlightRow(cb);
        }});
        updateBadge();
    }}
    if (e.target.classList.contains('row-cb')) {{
        highlightRow(e.target);
        const tbody = e.target.closest('tbody');
        const visibleInSection = tbody.querySelectorAll('tr:not(.row-hidden) input.row-cb');
        const checkedInSection = tbody.querySelectorAll('tr:not(.row-hidden) input.row-cb:checked');
        const sectionCb = e.target.closest('table').querySelector('input.section-cb');
        if (sectionCb) sectionCb.checked = (visibleInSection.length === checkedInSection.length);
        updateBadge();
    }}
}});

function highlightRow(cb) {{
    const tr = cb.closest('tr');
    if (cb.checked) tr.classList.add('row-selected');
    else tr.classList.remove('row-selected');
}}

// ============================================================
// TMX generation
// ============================================================
function escapeXml(s) {{
    return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&apos;');
}}

function langUnderscore(lang) {{
    return lang.replace('-', '_');
}}

function buildTmx(srcLang, tgtLang, entries) {{
    let xml = '<?xml version="1.0" encoding="UTF-8"?>\\n';
    xml += '<!DOCTYPE tmx SYSTEM "tmx14.dtd">\\n';
    xml += '<tmx version="1.4">';
    xml += '<header adminlang="en-US" creationtool="Tranzor" creationtoolversion="1.0"'
         + ' datatype="xml" o-tmf="Tranzor" srclang="' + srcLang + '" segtype="sentence">';
    xml += '</header>';
    xml += '<body>\\n';

    entries.forEach(function(e, i) {{
        const tuid = 'tranzor-mr-' + i;
        xml += '<tu tuid="' + tuid + '">';
        xml += '<prop type="x-segment-id">' + escapeXml(e.string_key) + '</prop>';
        xml += '\\n<tuv xml:lang="' + srcLang + '"><seg>' + escapeXml(e.source_text) + '</seg></tuv>\\n';
        xml += '<tuv xml:lang="' + tgtLang + '"><seg>' + escapeXml(e.translated_text) + '</seg></tuv>\\n';
        xml += '</tu>\\n';
    }});

    xml += '</body></tmx>';
    return xml;
}}

// ============================================================
// Download helpers
// ============================================================
function downloadBlob(blob, name) {{
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = name;
    document.body.appendChild(a);
    a.click();
    setTimeout(() => {{ URL.revokeObjectURL(a.href); a.remove(); }}, 100);
}}

async function exportTMX() {{
    const checked = document.querySelectorAll('tr:not(.row-hidden) input.row-cb:checked');
    if (!checked.length) return;

    const status = document.getElementById('exportStatus');
    status.textContent = 'Generating…';

    const byLang = {{}};
    checked.forEach(cb => {{
        const idx = parseInt(cb.dataset.idx);
        const row = ROWS[idx];
        const lang = row.language;
        if (!byLang[lang]) byLang[lang] = [];
        byLang[lang].push(row);
    }});

    const langs = Object.keys(byLang);

    if (langs.length === 1) {{
        const lang = langs[0];
        const tmx = buildTmx('en-US', lang, byLang[lang]);
        const blob = new Blob([tmx], {{ type: 'application/xml' }});
        const fn = 'tranzor_mr_translations_en_US-' + langUnderscore(lang) + '.tmx';
        downloadBlob(blob, fn);
        status.textContent = '✓ Downloaded ' + fn + ' (' + byLang[lang].length + ' entries)';
    }} else {{
        if (typeof JSZip === 'undefined') {{
            status.textContent = '⚠ JSZip not loaded. Check your internet connection and reload.';
            return;
        }}
        const zip = new JSZip();
        let totalEntries = 0;
        langs.forEach(lang => {{
            const tmx = buildTmx('en-US', lang, byLang[lang]);
            const fn = 'en_US-' + langUnderscore(lang) + '.tmx';
            zip.file(fn, tmx);
            totalEntries += byLang[lang].length;
        }});
        const content = await zip.generateAsync({{ type: 'blob' }});
        downloadBlob(content, 'tranzor_mr_translations_export.zip');
        status.textContent = '✓ Downloaded ZIP with ' + langs.length + ' languages, ' + totalEntries + ' entries';
    }}
}}
</script>
</body>
</html>"""

    with open(filename, "w", encoding="utf-8") as f:
        f.write(page)


# ---------------------------------------------------------------------------
# 8) Excel 导出 — MR 翻译结果
# ---------------------------------------------------------------------------
def write_mr_excel(results_data, filename):
    """生成 MR 翻译结果 Excel 报告"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("错误: 缺少 openpyxl 包，请先运行: pip install openpyxl")
        return

    translations = results_data.get("translations", [])
    is_changes = any(t.get("prev_translated_text") is not None for t in translations)

    wb = Workbook()
    ws = wb.active
    ws.title = "MR Translations"

    # Header — extra columns for changes mode
    if is_changes:
        headers = ["#", "String Key", "Language", "Source Text",
                   "Previous Translation", "Current Translation", "Diff",
                   "Changed At", "Fixed By", "MR", "Release", "JIRA",
                   "Score", "Error Category", "Reason"]
        widths = [6, 30, 10, 40, 40, 40, 50, 18, 25, 12, 18, 16, 8, 18, 30]
    else:
        headers = ["#", "String Key", "Language", "Source Text",
                   "Translated Text", "Score", "Error Category", "Reason", "Iteration"]
        widths = [6, 30, 10, 40, 40, 8, 18, 30, 8]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    diff_del_font = Font(color="CC0000", strikethrough=True)
    diff_ins_font = Font(color="006600", bold=True)

    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_i, t in enumerate(translations, 2):
        ws.cell(row=row_i, column=1, value=row_i - 1)
        ws.cell(row=row_i, column=2, value=t.get("opus_id", ""))
        ws.cell(row=row_i, column=3, value=t.get("target_language", ""))
        ws.cell(row=row_i, column=4, value=t.get("source_text", ""))

        if is_changes:
            ws.cell(row=row_i, column=5, value=t.get("prev_translated_text", ""))
            ws.cell(row=row_i, column=6, value=t.get("translated_text", ""))
            diff_text = _word_diff_text(
                t.get("prev_translated_text", ""),
                t.get("translated_text", ""))
            ws.cell(row=row_i, column=7, value=diff_text)
            task_time = t.get("task_created", "")[:16].replace("T", " ")
            ws.cell(row=row_i, column=8, value=task_time)
            ws.cell(row=row_i, column=9, value=t.get("fixed_by") or "")
            mr_link = t.get("mr_link", "")
            ws.cell(row=row_i, column=10, value=f'MR#{t.get("mr_iid", "")}')
            if mr_link:
                ws.cell(row=row_i, column=10).hyperlink = mr_link
                ws.cell(row=row_i, column=10).font = Font(color="0563C1", underline="single")
            ws.cell(row=row_i, column=11, value=t.get("release") or "")
            jira_id = t.get("jira_ticket_id") or ""
            ws.cell(row=row_i, column=12, value=jira_id)
            if jira_id:
                ws.cell(row=row_i, column=12).hyperlink = f"https://jira.ringcentral.com/browse/{jira_id}"
                ws.cell(row=row_i, column=12).font = Font(color="0563C1", underline="single")
            score = t.get("final_score")
            ws.cell(row=row_i, column=13, value=score if score is not None else "")
            ws.cell(row=row_i, column=14, value=t.get("error_category") or "")
            ws.cell(row=row_i, column=15, value=t.get("reason") or "")

            if score is not None:
                score_cell = ws.cell(row=row_i, column=13)
                if score < 80:
                    score_cell.font = Font(color="E74C3C", bold=True)
                elif score < 95:
                    score_cell.font = Font(color="E67E22", bold=True)
                else:
                    score_cell.font = Font(color="27AE60")
        else:
            ws.cell(row=row_i, column=5, value=t.get("translated_text", ""))
            score = t.get("final_score")
            ws.cell(row=row_i, column=6, value=score if score is not None else "")
            ws.cell(row=row_i, column=7, value=t.get("error_category") or "")
            ws.cell(row=row_i, column=8, value=t.get("reason") or "")
            ws.cell(row=row_i, column=9, value=t.get("iteration", 1))

            if score is not None:
                score_cell = ws.cell(row=row_i, column=6)
                if score < 80:
                    score_cell.font = Font(color="E74C3C", bold=True)
                elif score < 95:
                    score_cell.font = Font(color="E67E22", bold=True)
                else:
                    score_cell.font = Font(color="27AE60")

    # Column widths
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)
        ws.column_dimensions[col_letter].width = w

    wb.save(filename)


# ---------------------------------------------------------------------------
# 9) 统一保存入口
# ---------------------------------------------------------------------------
def save_mr_file(results_data, filename, label, fmt):
    """保存 MR 翻译结果，文件被占用时自动加序号"""
    base, ext = os.path.splitext(filename)
    save_path = filename
    for attempt in range(100):
        try:
            if fmt == "html":
                write_mr_html(results_data, save_path, label)
            else:
                write_mr_excel(results_data, save_path)
            print(f"已导出: {save_path}")
            if fmt == "html":
                from export_gui import open_in_browser
                open_in_browser(save_path)
            return save_path
        except PermissionError:
            attempt_num = attempt + 1
            save_path = f"{base}_{attempt_num}{ext}"
            print(f"  文件被占用，尝试保存为: {save_path}")
    return None
