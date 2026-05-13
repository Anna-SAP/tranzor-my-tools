"""
Unit tests for terminology_watchtower.py (pure logic).

Run:  python -m unittest test_terminology_watchtower
or:   python test_terminology_watchtower.py
"""
from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import terminology_watchtower as tw


# Sample CSV (matches PRD §8 example, plus a UTF-8 BOM variant)
SAMPLE_CSV = (
    "source_term,locale,approved_translation,source_aliases,"
    "forbidden_translations,product_scope,severity,case_sensitive,enabled,notes\n"
    "AI receptionist,fr_FR,Réceptionniste IA,AI receptionist,Standard IA,"
    "Voice|AI,Critical,false,true,Approved\n"
    "AI receptionist,de_DE,KI-Telefonzentrale,AI receptionist,KI-Rezeptionistin,"
    ",High,false,true,Approved\n"
    "AI receptionist,zh_CN,AI 接待员,AI receptionist,AI 前台,"
    "Voice|AI,High,false,true,Approved\n"
)


def make_candidate(**kw):
    base = dict(
        candidate_id="c1", product="Voice", locale="fr_fr",
        key="opus.welcome", source_text="Configure your AI receptionist.",
        target_text="", source_kind=tw.SOURCE_KIND_MR,
    )
    base.update(kw)
    base["locale"] = tw.normalize_locale(base["locale"])
    return tw.TranslationCandidate(**base)


class NormalizationTests(unittest.TestCase):

    def test_normalize_text_collapses_whitespace(self):
        self.assertEqual(tw.normalize_text("  hello   world  "), "hello world")
        self.assertEqual(tw.normalize_text(None), "")
        self.assertEqual(tw.normalize_text(""), "")

    def test_normalize_locale(self):
        self.assertEqual(tw.normalize_locale("zh-CN"), "zh_cn")
        self.assertEqual(tw.normalize_locale("ZH_CN"), "zh_cn")
        self.assertEqual(tw.normalize_locale(None), "")

    def test_pipe_list(self):
        self.assertEqual(tw.parse_pipe_list("a|b||c"), ["a", "b", "c"])
        self.assertEqual(tw.parse_pipe_list(""), [])

    def test_severity_max(self):
        self.assertEqual(tw.severity_max("Low", "High"), "High")
        self.assertEqual(tw.severity_max("Medium", "Critical", "High"), "Critical")
        self.assertEqual(tw.severity_max("xxx"), "Low")

    def test_normalize_severity(self):
        self.assertEqual(tw.normalize_severity("critical"), "Critical")
        self.assertEqual(tw.normalize_severity(""), "High")
        self.assertIsNone(tw.normalize_severity("urgent"))


class GlossaryImportTests(unittest.TestCase):

    def test_import_basic(self):
        res = tw.import_glossary_csv_text(SAMPLE_CSV)
        self.assertEqual(res.imported, 3)
        self.assertEqual(res.skipped, 0)
        self.assertEqual(len(res.rules), 3)
        self.assertEqual(res.rules[0].locale, "fr_fr")
        self.assertEqual(res.rules[0].severity, "Critical")

    def test_import_with_bom(self):
        text = "﻿" + SAMPLE_CSV
        res = tw.import_glossary_csv_text(text)
        self.assertEqual(res.imported, 3)

    def test_missing_required_column(self):
        bad = "source_term,locale\nfoo,en_US\n"
        res = tw.import_glossary_csv_text(bad)
        self.assertEqual(res.imported, 0)
        self.assertTrue(res.errors)
        self.assertIn("approved_translation", res.errors[0][1])

    def test_invalid_severity(self):
        bad = ("source_term,locale,approved_translation,severity\n"
               "AI,en_US,AI,urgent\n")
        res = tw.import_glossary_csv_text(bad)
        self.assertEqual(res.imported, 0)
        self.assertEqual(res.skipped, 1)
        self.assertTrue(res.errors)

    def test_default_severity_when_empty(self):
        body = ("source_term,locale,approved_translation\n"
                "AI receptionist,en_US,AI Receptionist\n")
        res = tw.import_glossary_csv_text(body)
        self.assertEqual(res.imported, 1)
        self.assertEqual(res.rules[0].severity, "High")
        self.assertTrue(res.rules[0].enabled)


class ScannerTests(unittest.TestCase):

    def setUp(self):
        self.rules = tw.import_glossary_csv_text(SAMPLE_CSV).rules

    def test_locale_normalization_match(self):
        cand = make_candidate(
            locale="fr-FR",
            target_text="Standard IA installé.",
        )
        issues, _ = tw.scan_candidates(self.rules, [cand])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].locale, "fr_fr")
        self.assertEqual(issues[0].issue_type, tw.ISSUE_BOTH)
        self.assertEqual(issues[0].severity, "Critical")

    def test_required_term_missing(self):
        cand = make_candidate(
            locale="zh_CN",
            source_text="The AI receptionist greets users.",
            target_text="该助手会打招呼。",
        )
        issues, _ = tw.scan_candidates(self.rules, [cand])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].issue_type, tw.ISSUE_REQUIRED_MISSING)

    def test_forbidden_used_only(self):
        cand = make_candidate(
            locale="zh_CN",
            source_text="The AI receptionist greets users.",
            target_text="AI 接待员 也叫 AI 前台。",
        )
        issues, _ = tw.scan_candidates(self.rules, [cand])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].issue_type, tw.ISSUE_FORBIDDEN_USED)
        self.assertEqual(issues[0].severity, tw.severity_max("High", "Medium"))
        self.assertIn("AI 前台", issues[0].forbidden_found)

    def test_missing_target(self):
        # de_DE rule has severity High → MissingTarget severity = max(High, High) = High
        cand = make_candidate(
            locale="de_DE",
            source_text="Configure your AI receptionist.",
            target_text="",
        )
        issues, _ = tw.scan_candidates(self.rules, [cand])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].issue_type, tw.ISSUE_MISSING_TARGET)
        self.assertEqual(issues[0].severity, "High")

    def test_missing_target_critical_rule_stays_critical(self):
        # fr_FR rule has Critical severity → MissingTarget severity stays Critical
        cand = make_candidate(
            locale="fr_FR",
            source_text="Configure your AI receptionist.",
            target_text="",
        )
        issues, _ = tw.scan_candidates(self.rules, [cand])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].severity, "Critical")

    def test_compliant_translation_no_issue(self):
        cand = make_candidate(
            locale="fr_FR",
            source_text="Configure your AI receptionist.",
            target_text="Configurez votre Réceptionniste IA.",
        )
        issues, _ = tw.scan_candidates(self.rules, [cand])
        self.assertEqual(len(issues), 0)

    def test_accent_preservation(self):
        # "Receptionniste" without accent must NOT count as containing
        # the approved "Réceptionniste IA".
        cand = make_candidate(
            locale="fr_FR",
            source_text="Configure your AI receptionist.",
            target_text="Receptionniste IA installé.",
        )
        issues, _ = tw.scan_candidates(self.rules, [cand])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].issue_type, tw.ISSUE_REQUIRED_MISSING)

    def test_disabled_rule_ignored(self):
        rules = [tw.TermRule(
            rule_id="r1", source_term="AI receptionist",
            locale="fr_fr", approved_translation="Réceptionniste IA",
            enabled=False, severity="Critical",
        )]
        cand = make_candidate(target_text="anything")
        issues, _ = tw.scan_candidates(rules, [cand])
        self.assertEqual(len(issues), 0)

    def test_product_scope_filter(self):
        rule = tw.TermRule(
            rule_id="r1", source_term="AI receptionist",
            locale="fr_fr", approved_translation="Réceptionniste IA",
            product_scope=["Voice", "AI"], severity="High",
        )
        out_of_scope = make_candidate(product="Meetings", target_text="bonjour")
        in_scope = make_candidate(product="Voice", target_text="bonjour")
        issues, _ = tw.scan_candidates([rule], [out_of_scope, in_scope])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].product, "Voice")

    def test_case_sensitive_rule(self):
        rule = tw.TermRule(
            rule_id="r1", source_term="AI receptionist",
            locale="fr_fr", approved_translation="Réceptionniste IA",
            case_sensitive=True, severity="High",
        )
        cand_no_match = make_candidate(
            source_text="ai receptionist is here", target_text="x")
        issues, _ = tw.scan_candidates([rule], [cand_no_match])
        self.assertEqual(len(issues), 0)

    def test_stable_issue_id(self):
        cand = make_candidate(target_text="x")
        issues1, _ = tw.scan_candidates(self.rules, [cand])
        issues2, _ = tw.scan_candidates(self.rules, [cand])
        self.assertEqual(issues1[0].issue_id, issues2[0].issue_id)
        # changing actual text should not change the id
        cand2 = make_candidate(target_text="y")
        issues3, _ = tw.scan_candidates(self.rules, [cand2])
        self.assertEqual(issues1[0].issue_id, issues3[0].issue_id)


class StatusStoreTests(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "issue_statuses.json")

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_status_persists_across_rebuild(self):
        s1 = tw.StatusStore(self.path)
        s1.set("ISS-x", "Reported")
        self.assertEqual(s1.get("ISS-x"), "Reported")

        s2 = tw.StatusStore(self.path)
        self.assertEqual(s2.get("ISS-x"), "Reported")
        self.assertEqual(s2.get("ISS-unknown"), tw.DEFAULT_STATUS)

    def test_invalid_status_rejected(self):
        s = tw.StatusStore(self.path)
        with self.assertRaises(ValueError):
            s.set("ISS-x", "Bogus")


class FilterTests(unittest.TestCase):

    def test_filter_by_severity_and_search(self):
        rules = tw.import_glossary_csv_text(SAMPLE_CSV).rules
        cands = [
            make_candidate(candidate_id="a", locale="fr_FR",
                           target_text="x"),
            make_candidate(candidate_id="b", locale="zh_CN",
                           source_text="The AI receptionist greets users.",
                           target_text="AI 接待员 也叫 AI 前台。"),
        ]
        issues, _ = tw.scan_candidates(rules, cands)
        crit = tw.filter_issues(issues, severity="Critical")
        self.assertTrue(all(i.severity == "Critical" for i in crit))
        zh = tw.filter_issues(issues, locale="zh_CN")
        self.assertTrue(all(i.locale == "zh_cn" for i in zh))


class ExportTests(unittest.TestCase):

    def setUp(self):
        rules = tw.import_glossary_csv_text(SAMPLE_CSV).rules
        cand = make_candidate(target_text="Standard IA installé.")
        self.issues, self.summary = tw.scan_candidates(rules, [cand])
        self.tmp = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_html_export(self):
        path = os.path.join(self.tmp, "evidence.html")
        tw.write_evidence_html(self.issues, self.summary, path)
        with open(path, "r", encoding="utf-8") as f:
            html = f.read()
        self.assertIn("<meta charset=\"UTF-8\">", html)
        self.assertIn("AI receptionist", html)
        self.assertIn("Réceptionniste IA", html)
        self.assertIn("Standard IA", html)

    def test_markdown_export(self):
        path = os.path.join(self.tmp, "evidence.md")
        tw.write_evidence_markdown(self.issues, self.summary, path)
        with open(path, "r", encoding="utf-8") as f:
            md = f.read()
        self.assertIn("AI receptionist", md)
        self.assertIn("Critical", md)

    def test_xlsx_export(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not available")
        path = os.path.join(self.tmp, "evidence.xlsx")
        tw.export_evidence_xlsx(self.issues, self.summary, path)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        self.assertIn("Issues", wb.sheetnames)
        self.assertIn("Summary", wb.sheetnames)
        ws = wb["Issues"]
        headers = [c.value for c in ws[1]]
        self.assertIn("Severity", headers)
        self.assertIn("Source Term", headers)


class GlossaryXlsxImportTests(unittest.TestCase):

    def setUp(self):
        try:
            from openpyxl import Workbook  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not available")
        self.tmp = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _write_xlsx(self, name, rows):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Terms"
        for r in rows:
            ws.append(r)
        path = os.path.join(self.tmp, name)
        wb.save(path)
        return path

    def test_standard_ringcentral_format(self):
        path = self._write_xlsx("Glossary_DE-DE_test.xlsx", [
            ("Term ID", "Source Term (EN)", "Target Term (DE-DE)",
             "Part of Speech", "Definition", "Context", "Status", "Notes",
             "Last Mod Date"),
            ("DE-DE_00001", "decision-making", "Entscheiden",
             "noun", None, None, "VALID", None, "2025-05-14"),
            ("DE-DE_00002", "AI receptionist", "KI-Telefonzentrale",
             "noun", "Approved product term", None, "VALID",
             "DNT: No", "2026-03-06"),
        ])
        res = tw.import_glossary(path)
        self.assertEqual(res.imported, 2, msg=str(res.errors))
        self.assertEqual(res.rules[0].rule_id, "DE-DE_00001")
        self.assertEqual(res.rules[0].locale, "de_de")
        self.assertEqual(res.rules[0].locale_display, "DE-DE")
        self.assertEqual(res.rules[1].source_term, "AI receptionist")
        self.assertEqual(res.rules[1].approved_translation, "KI-Telefonzentrale")
        self.assertTrue(res.rules[1].enabled)
        # Notes aggregated from Notes + Definition + Context + POS
        self.assertIn("Approved product term", res.rules[1].notes)
        self.assertIn("DNT: No", res.rules[1].notes)

    def test_status_other_than_valid_disables(self):
        path = self._write_xlsx("Glossary_FR-FR_test.xlsx", [
            ("Term ID", "Source Term (EN)", "Target Term (FR-FR)", "Status"),
            ("FR-FR_001", "AI receptionist", "Réceptionniste IA", "VALID"),
            ("FR-FR_002", "obsolete term", "vieux terme", "DEPRECATED"),
            ("FR-FR_003", "draft term", "brouillon", "PENDING"),
        ])
        res = tw.import_glossary(path)
        self.assertEqual(res.imported, 3)
        self.assertTrue(res.rules[0].enabled)
        self.assertFalse(res.rules[1].enabled)
        self.assertFalse(res.rules[2].enabled)
        self.assertIn("status=DEPRECATED", res.rules[1].notes)

    def test_locale_from_filename_fallback(self):
        path = self._write_xlsx("Glossary_zh-CN_skill.xlsx", [
            ("Term ID", "Source Term (EN)", "Target Term", "Status"),
            ("X1", "AI receptionist", "AI 接待员", "VALID"),
        ])
        res = tw.import_glossary(path)
        self.assertEqual(res.imported, 1, msg=str(res.errors))
        self.assertEqual(res.rules[0].locale, "zh_cn")

    def test_skip_empty_target(self):
        path = self._write_xlsx("Glossary_DE-DE_x.xlsx", [
            ("Term ID", "Source Term (EN)", "Target Term (DE-DE)", "Status"),
            ("X1", "with target", "Übersetzung", "VALID"),
            ("X2", "no target", None, "VALID"),
        ])
        res = tw.import_glossary(path)
        self.assertEqual(res.imported, 1)

    def test_multi_locale_one_sheet(self):
        path = self._write_xlsx("Glossary_multi.xlsx", [
            ("Term ID", "Source Term (EN)",
             "Target Term (DE-DE)", "Target Term (FR-FR)", "Status"),
            ("M1", "AI receptionist", "KI-Telefonzentrale",
             "Réceptionniste IA", "VALID"),
        ])
        res = tw.import_glossary(path)
        self.assertEqual(res.imported, 2)
        locales = sorted(r.locale for r in res.rules)
        self.assertEqual(locales, ["de_de", "fr_fr"])

    def test_real_fixture_if_present(self):
        fixture = "D:/Downloads_D/Terminology_\U0001f495SKILL-SKILL/Glossary_DE-DE_from_skill.xlsx"
        if not os.path.exists(fixture):
            self.skipTest("real fixture not present")
        res = tw.import_glossary(fixture)
        self.assertGreater(res.imported, 100,
                           msg=f"only {res.imported} imported, errors={res.errors[:5]}")
        # Sanity: locale normalized to de_de
        self.assertTrue(all(r.locale == "de_de" for r in res.rules))
        # And every rule has a source + approved
        for r in res.rules[:50]:
            self.assertTrue(r.source_term)
            self.assertTrue(r.approved_translation)


class CandidateAdapterTests(unittest.TestCase):

    def test_candidates_from_inventory(self):
        class FakeInv:
            data = {
                "Voice": {
                    "en-US": {"opus.welcome": "Configure your AI receptionist."},
                    "fr-FR": {"opus.welcome": "Standard IA."},
                }
            }
            sources = {
                "Voice": {
                    "fr-FR": {"opus.welcome": {"source": "MR",
                                               "task_id": "t1"}},
                }
            }

        cands = tw.candidates_from_full_inventory(FakeInv())
        # Only fr-FR is a target (en-US is reference)
        self.assertEqual(len(cands), 1)
        c = cands[0]
        self.assertEqual(c.locale, "fr_fr")
        self.assertEqual(c.source_kind, tw.SOURCE_KIND_MR)
        self.assertEqual(c.task_id, "t1")
        self.assertIn("AI receptionist", c.source_text)


class MergeAndQuickCheckTests(unittest.TestCase):

    def test_merge_by_rule_id_incoming_wins(self):
        existing = [
            tw.TermRule(rule_id="r1", source_term="A", locale="fr_fr",
                        approved_translation="X", severity="High"),
            tw.TermRule(rule_id="r2", source_term="B", locale="fr_fr",
                        approved_translation="Y", severity="High"),
        ]
        incoming = [
            tw.TermRule(rule_id="r1", source_term="A", locale="fr_fr",
                        approved_translation="X-NEW", severity="Critical"),
            tw.TermRule(rule_id="r3", source_term="C", locale="de_de",
                        approved_translation="Z", severity="High"),
        ]
        merged = tw.merge_rules(existing, incoming)
        self.assertEqual(len(merged), 3)
        ids = [r.rule_id for r in merged]
        self.assertEqual(ids, ["r1", "r2", "r3"])
        r1 = next(r for r in merged if r.rule_id == "r1")
        self.assertEqual(r1.approved_translation, "X-NEW")
        self.assertEqual(r1.severity, "Critical")

    def test_parse_clipboard_pairs_tab(self):
        pairs = tw.parse_clipboard_pairs("fr_FR\tRéceptionniste IA\n"
                                         "de_DE\tKI-Telefonzentrale\n")
        self.assertEqual(pairs, [("fr_FR", "Réceptionniste IA"),
                                 ("de_DE", "KI-Telefonzentrale")])

    def test_parse_clipboard_pairs_arrow_and_comma(self):
        pairs = tw.parse_clipboard_pairs(
            "fr_FR -> Agent IA\n"
            "de_DE → KI-Agent\n"
            "zh_CN, AI 智能体\n"
            "# comment\n"
            "\n"
        )
        self.assertEqual(pairs, [("fr_FR", "Agent IA"),
                                 ("de_DE", "KI-Agent"),
                                 ("zh_CN", "AI 智能体")])

    def test_parse_clipboard_pairs_whitespace(self):
        pairs = tw.parse_clipboard_pairs("fr_FR   Agent IA\n")
        self.assertEqual(pairs, [("fr_FR", "Agent IA")])

    def test_build_quick_check_rules(self):
        rules = tw.build_quick_check_rules(
            "AI Agent",
            [("fr_FR", "Agent IA"), ("de_DE", "KI-Agent"),
             ("zh_CN", "AI 智能体")],
            severity="Critical",
        )
        self.assertEqual(len(rules), 3)
        self.assertTrue(all(r.rule_id.startswith("QC-") for r in rules))
        self.assertTrue(all(r.severity == "Critical" for r in rules))
        self.assertEqual({r.locale for r in rules},
                         {"fr_fr", "de_de", "zh_cn"})

    def test_quick_check_end_to_end(self):
        rules = tw.build_quick_check_rules(
            "AI Agent",
            [("zh_CN", "AI 智能体")],
        )
        cand_bad = make_candidate(
            candidate_id="c1", locale="zh_CN",
            source_text="The AI Agent answers customer questions.",
            target_text="该 AI 代理人会回答客户问题。",
        )
        cand_ok = make_candidate(
            candidate_id="c2", locale="zh_CN",
            source_text="Configure your AI Agent today.",
            target_text="今天就配置你的 AI 智能体。",
        )
        issues, _ = tw.scan_candidates(rules, [cand_bad, cand_ok])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].candidate_id, "c1")
        self.assertEqual(issues[0].issue_type, tw.ISSUE_REQUIRED_MISSING)


class ReorgPasteTests(unittest.TestCase):

    def test_alternating_with_header(self):
        text = (
            "Translations\n"
            "de_DE\nKI-Repräsentant\n"
            "en_GB\nAI Representative\n"
            "es_419\nRepresentante de IA\n"
            "fi_FI\nAI-edustaja\n"
            "fr_CA\nReprésentant IA\n"
        )
        out = tw.reorganize_paste(text)
        pairs = tw.parse_clipboard_pairs(out)
        self.assertEqual(len(pairs), 5)
        self.assertEqual(pairs[0], ("de_DE", "KI-Repräsentant"))
        self.assertEqual(pairs[2], ("es_419", "Representante de IA"))
        self.assertEqual(pairs[4], ("fr_CA", "Représentant IA"))

    def test_already_paired_passthrough(self):
        text = "fr_FR\tAgent IA\nde_DE\tKI-Agent\n"
        out = tw.reorganize_paste(text)
        # Already paired; reorg should not duplicate or drop rows
        pairs = tw.parse_clipboard_pairs(out)
        self.assertEqual(pairs, [("fr_FR", "Agent IA"),
                                 ("de_DE", "KI-Agent")])

    def test_arrow_separator_passthrough(self):
        text = "fr_FR → Agent IA\nde_DE -> KI-Agent\n"
        out = tw.reorganize_paste(text)
        pairs = tw.parse_clipboard_pairs(out)
        self.assertEqual(pairs, [("fr_FR", "Agent IA"),
                                 ("de_DE", "KI-Agent")])

    def test_multiline_translation_value(self):
        text = (
            "fr_FR\n"
            "Agent IA\n"
            "(Réceptionniste IA)\n"
            "de_DE\n"
            "KI-Agent\n"
        )
        out = tw.reorganize_paste(text)
        pairs = tw.parse_clipboard_pairs(out)
        self.assertEqual(len(pairs), 2)
        self.assertEqual(pairs[0][0], "fr_FR")
        self.assertIn("Agent IA", pairs[0][1])
        self.assertIn("Réceptionniste IA", pairs[0][1])
        self.assertEqual(pairs[1], ("de_DE", "KI-Agent"))

    def test_drops_header_words(self):
        text = (
            "Locale\nApproved Translation\n"
            "fr_FR\nAgent IA\n"
        )
        out = tw.reorganize_paste(text)
        self.assertEqual(out, "fr_FR\tAgent IA")

    def test_no_locale_tokens_returns_input(self):
        text = "just\nsome\nrandom lines\n"
        out = tw.reorganize_paste(text)
        # No locale tokens detected; should fall back rather than fabricate
        self.assertIn("just", out)
        self.assertNotIn("\t", out)

    def test_empty_input(self):
        self.assertEqual(tw.reorganize_paste(""), "")
        self.assertEqual(tw.reorganize_paste("\n\n\n"), "")


class LocaleFallbackTests(unittest.TestCase):

    def test_zh_hk_inherits_zh_tw(self):
        # A single zh_tw source rule generates fallbacks for ALL fallback
        # locales whose source is zh_tw — currently zh_hk and zh_mo.
        rule = tw.TermRule(
            rule_id="r1", source_term="AI Agent",
            locale="zh_tw", approved_translation="AI 智能體",
            severity="High", locale_display="zh-TW",
        )
        out = tw.expand_locale_fallbacks([rule])
        locales = {r.locale for r in out}
        self.assertEqual(locales, {"zh_tw", "zh_hk", "zh_mo"})
        fb = next(r for r in out if r.locale == "zh_hk")
        self.assertEqual(fb.approved_translation, "AI 智能體")
        self.assertEqual(fb.locale_display, "zh-HK")
        self.assertEqual(fb.source_term, "AI Agent")
        self.assertEqual(fb.rule_id, "r1::fallback::zh_hk")
        self.assertIn("fallback from zh-TW", fb.notes)

    def test_explicit_zh_hk_not_overridden(self):
        rules = [
            tw.TermRule(rule_id="r1", source_term="AI Agent",
                        locale="zh_tw", approved_translation="AI 智能體",
                        severity="High"),
            tw.TermRule(rule_id="r2", source_term="AI Agent",
                        locale="zh_hk", approved_translation="AI 智能代理",
                        severity="High"),
        ]
        out = tw.expand_locale_fallbacks(rules)
        # zh_hk explicit, zh_mo synthesized from zh_tw (still missing) → 3
        locales_to_count = {}
        for r in out:
            locales_to_count[r.locale] = locales_to_count.get(r.locale, 0) + 1
        self.assertEqual(locales_to_count.get("zh_hk"), 1)  # not duplicated
        zh_hk = next(r for r in out if r.locale == "zh_hk")
        self.assertEqual(zh_hk.approved_translation, "AI 智能代理")
        self.assertNotIn("fallback", zh_hk.notes)  # still the explicit rule

    def test_disabled_source_does_not_propagate(self):
        rule = tw.TermRule(
            rule_id="r1", source_term="X", locale="zh_tw",
            approved_translation="x-tw",
            severity="High", enabled=False,
        )
        out = tw.expand_locale_fallbacks([rule])
        self.assertEqual(len(out), 1)  # original only, no fallback

    def test_per_term_isolation(self):
        rules = [
            tw.TermRule(rule_id="t1", source_term="Term1",
                        locale="zh_tw", approved_translation="t1-tw",
                        severity="High"),
            tw.TermRule(rule_id="t2", source_term="Term2",
                        locale="zh_hk", approved_translation="t2-hk-explicit",
                        severity="High"),
        ]
        out = tw.expand_locale_fallbacks(rules)
        # Term1 gets a zh_hk fallback; Term2 does NOT get a zh_tw fallback
        # (we only synthesize fallback→source direction, not the reverse).
        zh_hks = [r for r in out if r.locale == "zh_hk"]
        self.assertEqual(len(zh_hks), 2)  # Term1 synthetic + Term2 explicit
        # Term2 synthetic should NOT exist — fallback is one-way zh_tw → zh_hk
        zh_tws = [r for r in out if r.locale == "zh_tw"]
        self.assertEqual(len(zh_tws), 1)

    def test_quick_check_rules_apply_fallback(self):
        rules = tw.build_quick_check_rules(
            "AI Agent",
            [("zh-TW", "AI 智能體")],
        )
        locales = {r.locale for r in rules}
        self.assertIn("zh_tw", locales)
        self.assertIn("zh_hk", locales)
        zh_hk = next(r for r in rules if r.locale == "zh_hk")
        self.assertEqual(zh_hk.approved_translation, "AI 智能體")

    def test_quick_check_rules_can_opt_out(self):
        rules = tw.build_quick_check_rules(
            "AI Agent",
            [("zh-TW", "AI 智能體")],
            apply_locale_fallbacks=False,
        )
        locales = {r.locale for r in rules}
        self.assertEqual(locales, {"zh_tw"})

    def test_zh_hk_candidate_now_caught(self):
        """End-to-end: zh-TW glossary rule + zh-HK candidate with a
        non-conforming translation should now produce an issue."""
        rules = tw.build_quick_check_rules(
            "AI agent",
            [("zh-TW", "AI 智能體")],
        )
        cand = make_candidate(
            candidate_id="hk1",
            locale="zh-HK",
            source_text="Chat with your AI agent to test the skill.",
            target_text="與您的 AI 智能代理對話",
        )
        issues, summary = tw.scan_candidates(rules, [cand])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].locale, "zh_hk")
        self.assertEqual(issues[0].issue_type, tw.ISSUE_REQUIRED_MISSING)
        self.assertIn("fallback from zh-TW", issues[0].rule_notes)


class TranzorFallbackTests(unittest.TestCase):
    """Specifically the Tranzor Terminology API path: API has zh-TW but
    no zh-HK. After expansion, zh-HK rules must exist."""

    def test_details_to_rules_synthesizes_zh_hk(self):
        import tranzor_terminology as tz
        detail = {
            "id": 8858, "code": "abc", "name": "AI Agent",
            "scope": "Global", "dnt": False,
            "translations": [
                {"language_code": "zh-TW", "translated_name": "AI 智能體"},
                {"language_code": "zh-CN", "translated_name": "AI 智能体"},
            ],
        }
        rules = tz.details_to_rules([detail])
        locales = {r.locale for r in rules}
        # zh_tw spawns fallbacks for zh_hk and zh_mo
        self.assertEqual(locales, {"zh_tw", "zh_cn", "zh_hk", "zh_mo"})
        zh_hk = next(r for r in rules if r.locale == "zh_hk")
        self.assertEqual(zh_hk.approved_translation, "AI 智能體")
        self.assertIn("fallback", zh_hk.notes)

    def test_details_to_rules_can_opt_out(self):
        import tranzor_terminology as tz
        detail = {
            "id": 1, "code": "x", "name": "Term", "dnt": False,
            "translations": [
                {"language_code": "zh-TW", "translated_name": "TW"},
            ],
        }
        rules = tz.details_to_rules([detail], apply_locale_fallbacks=False)
        self.assertEqual({r.locale for r in rules}, {"zh_tw"})


class ScannerPerfSmokeTests(unittest.TestCase):
    """Not a strict benchmark — just a sanity ceiling on a 2350-rule glossary."""

    def test_2350_rules_30k_candidates_under_10s(self):
        import time
        # Build 2350 de_DE rules
        rules = []
        for n in range(2350):
            rules.append(tw.TermRule(
                rule_id=f"r{n:04d}",
                source_term=f"term{n:04d}",
                locale="de_de",
                approved_translation=f"deutsche{n:04d}",
                severity="High",
            ))
        # Build 30k candidates: 1k de_DE (some matching), 29k other locales
        cands = []
        for i in range(1000):
            cands.append(tw.TranslationCandidate(
                candidate_id=f"de-{i}",
                product="Voice", locale="de_de", key=f"k{i}",
                source_text=f"prefix term{i % 2350:04d} suffix",
                target_text=f"falsch{i:04d}",
                source_kind=tw.SOURCE_KIND_MR,
            ))
        for i in range(29000):
            cands.append(tw.TranslationCandidate(
                candidate_id=f"fr-{i}",
                product="Voice", locale="fr_fr", key=f"k{i}",
                source_text="random source text",
                target_text="random target",
                source_kind=tw.SOURCE_KIND_MR,
            ))

        t0 = time.perf_counter()
        issues, summary = tw.scan_candidates(rules, cands)
        elapsed = time.perf_counter() - t0

        # Should produce roughly 1 issue per de_DE candidate that matches
        self.assertGreater(len(issues), 500)
        self.assertEqual(summary.scanned_candidates, 30000)
        self.assertLess(elapsed, 10.0,
                        msg=f"perf regression: {elapsed:.2f}s on 30k×2350")
        # Print for visibility (only when running in verbose)
        if "-v" in sys.argv or "--verbose" in sys.argv:
            print(f"\n[perf] 30k×2350 scan: {elapsed:.2f}s, {len(issues)} issues")


class TranzorTerminologyClientTests(unittest.TestCase):
    """Pure tests — no live network."""

    def test_term_detail_to_rules(self):
        import tranzor_terminology as tz
        detail = {
            "id": 8858,
            "code": "019d8a594b19f2779ce7456fc92cd84c",
            "name": "AI Representative",
            "scope": "Global",
            "dnt": False,
            "definition": "An evolution of the AI receptionist",
            "part_of_speech": "Noun",
            "translations": [
                {"language_code": "de-DE", "translated_name": "KI-Repräsentant"},
                {"language_code": "fr-FR", "translated_name": "Représentant IA"},
                {"language_code": "zh-CN", "translated_name": "AI 代表"},
                {"language_code": "ja-JP", "translated_name": ""},  # skipped
            ],
        }
        rules = tz.term_detail_to_rules(detail, severity="Critical")
        self.assertEqual(len(rules), 3)
        # All rules share term + severity, differ by locale
        self.assertTrue(all(r.source_term == "AI Representative" for r in rules))
        self.assertTrue(all(r.severity == "Critical" for r in rules))
        self.assertEqual({r.locale for r in rules},
                         {"de_de", "fr_fr", "zh_cn"})
        self.assertTrue(all(r.rule_id.startswith("TZ-") for r in rules))
        # Notes carry definition/POS for traceability
        self.assertIn("AI receptionist", rules[0].notes)
        self.assertIn("Noun", rules[0].notes)

    def test_dnt_term_skipped_by_default(self):
        import tranzor_terminology as tz
        detail = {
            "id": 1, "code": "x", "name": "Beetexting", "dnt": True,
            "translations": [],
        }
        self.assertEqual(tz.term_detail_to_rules(detail), [])
        # Passthrough mode emits no rules either when there are no
        # locale codes to anchor against.
        self.assertEqual(
            tz.term_detail_to_rules(detail, include_dnt_as_passthrough=True),
            [],
        )

    def test_dnt_term_passthrough_with_locales(self):
        import tranzor_terminology as tz
        detail = {
            "id": 2, "code": "y", "name": "Avaya", "dnt": True,
            "translations": [
                {"language_code": "de-DE", "translated_name": ""},
                {"language_code": "fr-FR", "translated_name": ""},
            ],
        }
        rules = tz.term_detail_to_rules(detail,
                                        include_dnt_as_passthrough=True)
        self.assertEqual(len(rules), 2)
        # Approved == source term — DNT semantics
        self.assertTrue(all(r.approved_translation == "Avaya" for r in rules))

    def test_details_to_rules_bulk(self):
        import tranzor_terminology as tz
        details = [
            {"id": 1, "code": "a", "name": "A", "dnt": False,
             "translations": [{"language_code": "de-DE",
                               "translated_name": "A-de"}]},
            {"id": 2, "code": "b", "name": "B", "dnt": False,
             "translations": [{"language_code": "fr-FR",
                               "translated_name": "B-fr"}]},
        ]
        rules = tz.details_to_rules(details)
        self.assertEqual(len(rules), 2)
        self.assertEqual({r.source_term for r in rules}, {"A", "B"})

    def test_terminology_app_url(self):
        import tranzor_terminology as tz
        url = tz.terminology_app_url()
        self.assertTrue(url.startswith(tz.TRANZOR_URL))
        # SPA entry — the static asset path, NOT the API. Past mistake
        # was returning /context/terminology/{id} which 404s because
        # that's the backend API namespace.
        self.assertTrue(url.endswith("/static/terminology"))
        self.assertNotIn("/api/", url)
        self.assertFalse(hasattr(tz, "terminology_detail_url"),
                         "stale per-id URL builder must be removed — "
                         "SPA has no deep links per term")


if __name__ == "__main__":
    unittest.main()
