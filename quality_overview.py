"""
质量概览模块
=============
质量数据聚合与可视化，支持 tkinter Canvas 图表 + HTML / Excel 导出。
支持 MR 翻译链路和 Legacy（文件翻译）链路的双路质量分析。
"""

import html as html_mod
import math
import os
import platform
import webbrowser
from collections import OrderedDict
from datetime import date, datetime

# ---------------------------------------------------------------------------
# 跨平台字体适配
# ---------------------------------------------------------------------------
if platform.system() == "Darwin":  # macOS
    FONT_FAMILY = "Helvetica Neue"
else:  # Windows / Linux
    FONT_FAMILY = "Segoe UI"


# ---------------------------------------------------------------------------
# 1) 常量与配置
# ---------------------------------------------------------------------------
DEFAULT_THRESHOLD = 98

SCORE_BINS = [
    ("0-79", 0, 80),
    ("80-89", 80, 90),
    ("90-94", 90, 95),
    ("95-97", 95, 98),
    ("98-99", 98, 100),
    ("100", 100, 101),
]

ERROR_CATEGORIES = [
    "Accuracy", "Fluency", "Terminology", "Consistency",
    "Locale Convention", "Variable/Number Mismatch",
]


# ---------------------------------------------------------------------------
# 2) 数据聚合 — 通用内部函数
# ---------------------------------------------------------------------------
def _as_score(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value))
    except (TypeError, ValueError):
        return None


def _normalize_error_category(category):
    if category in (None, "", "None"):
        return None
    return str(category)


def _has_refinement(item):
    iteration = item.get("iteration")
    if iteration is not None:
        try:
            if int(iteration) >= 2:
                return True
        except (TypeError, ValueError):
            pass
    return bool(item.get("iteration_history"))


def _has_human_touch(item):
    for field in ("reviewer_comment", "reviewer_notes"):
        value = item.get(field)
        if value and str(value).strip():
            return True
    if item.get("fixed_by_lead"):
        return True
    try:
        return int(item.get("edit_log_count") or 0) > 0
    except (TypeError, ValueError):
        return False


def _warning_key(source_id, language):
    return str(source_id or ""), str(language or "")


def _parse_event_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"

    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        try:
            return datetime.fromisoformat(text.split("T")[0]).date()
        except ValueError:
            return None


def get_iteration_snapshot(item, iteration_key="iteration_1"):
    """Return a normalized iteration snapshot from iteration_history."""
    history = item.get("iteration_history")
    if not isinstance(history, dict):
        return {}

    raw = history.get(iteration_key) or {}
    evaluation = raw.get("evaluation") or {}
    return {
        "translation": raw.get("translation"),
        "final_score": _as_score(evaluation.get("final_score", raw.get("final_score"))),
        "error_category": evaluation.get("error_category", raw.get("error_category")),
        "reason": evaluation.get("reason", raw.get("reason")),
        "penalty_points": evaluation.get("penalty_points", raw.get("penalty_points")),
        "critical_count": evaluation.get("critical_count", raw.get("critical_count")),
        "major_count": evaluation.get("major_count", raw.get("major_count")),
        "minor_count": evaluation.get("minor_count", raw.get("minor_count")),
    }


def _build_trend_points(items, threshold):
    buckets = {}
    for item in items:
        event_date = _parse_event_date(item.get("_quality_event_at"))
        score = _as_score(item.get("final_score"))
        if event_date is None or score is None:
            continue
        bucket = buckets.setdefault(event_date, {"scores": [], "below": 0})
        bucket["scores"].append(score)
        if score < threshold:
            bucket["below"] += 1

    trend_points = []
    for event_date, bucket in sorted(buckets.items()):
        count = len(bucket["scores"])
        if count == 0:
            continue
        trend_points.append({
            "date": event_date.isoformat(),
            "label": event_date.strftime("%m-%d"),
            "avg_score": round(sum(bucket["scores"]) / count, 2),
            "below_rate": round(bucket["below"] / count * 100, 1),
            "count": count,
        })
    return trend_points


def _build_error_by_language(items):
    grouped = {}
    for item in items:
        category = _normalize_error_category(item.get("error_category"))
        if not category:
            continue
        language = item.get("target_language") or "(unknown)"
        lang_bucket = grouped.setdefault(language, {})
        lang_bucket[category] = lang_bucket.get(category, 0) + 1

    ranked = sorted(
        grouped.items(),
        key=lambda pair: (-sum(pair[1].values()), pair[0])
    )
    return OrderedDict(ranked[:10])


def _build_legacy_warning_index(warnings):
    index = {}
    for warning in warnings.get("inconsistent", []) or []:
        language = warning.get("target_language")
        for source_id in warning.get("source_ids") or []:
            key = _warning_key(source_id, language)
            index.setdefault(key, set()).add("inconsistent")

    for warning in warnings.get("untranslated", []) or []:
        key = _warning_key(warning.get("source_id"), warning.get("target_language"))
        index.setdefault(key, set()).add("untranslated")

    return index


def _compute_metrics(items, threshold=DEFAULT_THRESHOLD):
    """从归一化后的条目列表计算质量概览指标。"""
    total_items = len(items)
    scored = []
    for item in items:
        score = _as_score(item.get("final_score"))
        if score is None:
            continue
        normalized = dict(item)
        normalized["final_score"] = score
        scored.append(normalized)

    scores = [item["final_score"] for item in scored]
    avg_score = round(sum(scores) / len(scores), 2) if scores else None

    low_items = sorted(
        [item for item in scored if item["final_score"] < threshold],
        key=lambda item: (
            item["final_score"],
            item.get("target_language", ""),
            item.get("opus_id", ""),
        ),
    )
    below_count = len(low_items)
    below_rate = round(below_count / len(scored) * 100, 1) if scored else 0

    refined_count = sum(1 for item in items if _has_refinement(item))
    refined_rate = round(refined_count / total_items * 100, 1) if total_items else 0

    human_count = sum(1 for item in items if _has_human_touch(item))
    human_rate = round(human_count / total_items * 100, 1) if total_items else 0

    cached_count = sum(
        1 for item in items
        if item.get("cached") or item.get("ice_match")
    )

    err_dist = OrderedDict((category, 0) for category in ERROR_CATEGORIES)
    for item in items:
        category = _normalize_error_category(item.get("error_category"))
        if not category:
            continue
        err_dist[category] = err_dist.get(category, 0) + 1

    score_dist = OrderedDict((label, 0) for label, _, _ in SCORE_BINS)
    for item in scored:
        score = item["final_score"]
        for label, lower, upper in SCORE_BINS:
            if lower <= score < upper:
                score_dist[label] += 1
                break

    by_lang = {}
    for item in items:
        language = item.get("target_language") or "(unknown)"
        entry = by_lang.setdefault(language, {
            "count": 0,
            "score_sum": 0,
            "scored_count": 0,
            "below": 0,
            "refined": 0,
            "human": 0,
            "warnings": 0,
        })
        entry["count"] += 1
        score = _as_score(item.get("final_score"))
        if score is not None:
            entry["score_sum"] += score
            entry["scored_count"] += 1
            if score < threshold:
                entry["below"] += 1
        if _has_refinement(item):
            entry["refined"] += 1
        if _has_human_touch(item):
            entry["human"] += 1
        if item.get("warning_flags"):
            entry["warnings"] += 1

    lang_details = []
    for language, data in sorted(by_lang.items()):
        avg = data["score_sum"] / data["scored_count"] if data["scored_count"] else None
        below_pct = round(data["below"] / data["scored_count"] * 100, 1) if data["scored_count"] else 0
        refined_pct = round(data["refined"] / data["count"] * 100, 1) if data["count"] else 0
        human_pct = round(data["human"] / data["count"] * 100, 1) if data["count"] else 0
        lang_details.append({
            "language": language,
            "count": data["count"],
            "average_score": round(avg, 1) if avg is not None else None,
            "below_threshold_pct": below_pct,
            "refined_pct": refined_pct,
            "human_touched_pct": human_pct,
            "warnings": data["warnings"],
        })

    return {
        "total_items": total_items,
        "overall_avg_score": avg_score if avg_score is not None else 0,
        "low_score_count": below_count,
        "below_threshold_rate": below_rate,
        "refined_rate": refined_rate,
        "human_touch_rate": human_rate,
        "cached_count": cached_count,
        "threshold": threshold,
        "error_distribution": err_dist,
        "score_distribution": score_dist,
        "by_language": lang_details,
        "low_items": low_items,
        "trend_points": _build_trend_points(scored, threshold),
        "error_by_language": _build_error_by_language(items),
    }


# ---------------------------------------------------------------------------
# 3) 数据聚合 — MR 链路
# ---------------------------------------------------------------------------
def aggregate_mr_quality(overview_data, cases_data, threshold=DEFAULT_THRESHOLD):
    """从 dashboard overview + cases 聚合 MR 链路质量数据。"""
    all_items = []
    for mr in cases_data.get("mrs", []) or []:
        mr_iid = mr.get("mr_iid", "")
        project_id = mr.get("project_id", "")
        merged_at = mr.get("merged_at")
        for item in mr.get("cases", []) or []:
            normalized = dict(item)
            normalized["final_score"] = _as_score(item.get("final_score"))
            normalized["error_category"] = _normalize_error_category(item.get("error_category"))
            normalized["_source_type"] = "MR"
            normalized["_scope_name"] = f"MR !{mr_iid}" if mr_iid else ""
            normalized["_project_id"] = project_id
            normalized["_quality_event_at"] = merged_at
            all_items.append(normalized)

    metrics = _compute_metrics(all_items, threshold)
    metrics["total_tasks"] = overview_data.get("total_tasks", 0)
    metrics["completed"] = overview_data.get("completed", 0)
    metrics["failed"] = overview_data.get("failed", 0)
    return metrics


def aggregate_quality_data(overview_data, cases_data, threshold=DEFAULT_THRESHOLD):
    """向后兼容入口。"""
    return aggregate_mr_quality(overview_data, cases_data, threshold)


# ---------------------------------------------------------------------------
# 4) 数据聚合 — Legacy (文件翻译) 链路
# ---------------------------------------------------------------------------
def aggregate_legacy_quality(tasks, translations_map, warnings_map=None,
                             threshold=DEFAULT_THRESHOLD):
    """聚合 Legacy 文件翻译链路质量数据。"""
    if warnings_map is None:
        warnings_map = {}

    all_items = []
    for task in tasks:
        task_id = task.get("task_id") or task.get("id")
        task_name = task.get("task_name") or task.get("name", "")
        project_name = task.get("project_name", "")
        created_at = task.get("created_at")
        trans_list = translations_map.get(str(task_id), [])
        warning_index = _build_legacy_warning_index(
            warnings_map.get(str(task_id), {}) or {}
        )

        for item in trans_list:
            normalized = dict(item)
            normalized["final_score"] = _as_score(item.get("final_score"))
            normalized["error_category"] = _normalize_error_category(item.get("error_category"))
            normalized["_source_type"] = "File"
            normalized["_scope_name"] = task_name
            normalized["_task_id"] = str(task_id)
            normalized["_project_name"] = project_name
            normalized["_quality_event_at"] = created_at
            if not normalized.get("opus_id") and normalized.get("unit_id"):
                normalized["opus_id"] = normalized["unit_id"]

            warning_key = _warning_key(
                normalized.get("source_id"),
                normalized.get("target_language"),
            )
            warning_types = sorted(warning_index.get(warning_key, set()))
            if warning_types:
                normalized["warning_flags"] = True
                normalized["warning_types"] = warning_types
            all_items.append(normalized)

    metrics = _compute_metrics(all_items, threshold)
    metrics["total_tasks"] = len(tasks)
    metrics["completed"] = sum(
        1 for task in tasks if (task.get("status") or "").lower() == "completed"
    )
    metrics["failed"] = sum(
        1 for task in tasks if (task.get("status") or "").lower() == "failed"
    )
    return metrics


# ---------------------------------------------------------------------------
# 5) tkinter Canvas 图表
# ---------------------------------------------------------------------------
PIE_COLORS = [
    "#4472C4", "#E67E22", "#27AE60", "#8E44AD",
    "#E74C3C", "#16A085", "#D4AC0D", "#2C3E50",
]

BAR_COLORS = [
    "#27AE60", "#4472C4", "#D4AC0D",
    "#E67E22", "#E74C3C", "#8E44AD",
]


def draw_pie_chart(canvas, data, width, height, title=""):
    """在 tkinter Canvas 上绘制饼图
    data: dict { label: count, ... }
    """
    canvas.delete("all")

    # Title
    if title:
        canvas.create_text(width // 2, 16, text=title,
                           fill="#ccc", font=(FONT_FAMILY, 11, "bold"))

    total = sum(data.values())
    if total == 0:
        canvas.create_text(width // 2, height // 2, text="No Data",
                           fill="#666", font=(FONT_FAMILY, 10))
        return

    cx, cy = width // 2 - 40, height // 2 + 10
    r = min(width, height) // 2 - 40
    if r < 30:
        r = 30

    start = 0
    legend_items = []
    for i, (label, count) in enumerate(data.items()):
        if count == 0:
            continue
        extent = (count / total) * 360
        color = PIE_COLORS[i % len(PIE_COLORS)]
        canvas.create_arc(
            cx - r, cy - r, cx + r, cy + r,
            start=start, extent=extent,
            fill=color, outline="#1a1a2e", width=1,
            style="pieslice"
        )
        legend_items.append((label, count, color))
        start += extent

    # Legend
    lx = cx + r + 24
    ly = cy - r + 10
    for i, (label, count, color) in enumerate(legend_items):
        y = ly + i * 20
        canvas.create_rectangle(lx, y, lx + 12, y + 12, fill=color, outline="")
        pct = (count / total * 100) if total else 0
        canvas.create_text(
            lx + 18, y + 6, anchor="w",
            text=f"{label}: {count} ({pct:.0f}%)",
            fill="#bbb", font=(FONT_FAMILY, 9)
        )


def draw_bar_chart(canvas, data, width, height, title=""):
    """在 tkinter Canvas 上绘制柱状图
    data: OrderedDict { label: count, ... }
    """
    canvas.delete("all")

    # Title
    if title:
        canvas.create_text(width // 2, 16, text=title,
                           fill="#ccc", font=(FONT_FAMILY, 11, "bold"))

    labels = list(data.keys())
    values = list(data.values())
    max_val = max(values) if values else 1
    if max_val == 0:
        max_val = 1

    n = len(labels)
    if n == 0:
        canvas.create_text(width // 2, height // 2, text="No Data",
                           fill="#666", font=(FONT_FAMILY, 10))
        return

    pad_left = 50
    pad_right = 20
    pad_top = 40
    pad_bottom = 40
    chart_w = width - pad_left - pad_right
    chart_h = height - pad_top - pad_bottom
    bar_w = chart_w // n - 8
    if bar_w < 10:
        bar_w = 10

    # Draw bars
    for i, (label, val) in enumerate(zip(labels, values)):
        x = pad_left + i * (chart_w // n) + (chart_w // n - bar_w) // 2
        bar_h = int((val / max_val) * chart_h)
        y_top = pad_top + chart_h - bar_h
        y_bot = pad_top + chart_h
        color = BAR_COLORS[i % len(BAR_COLORS)]

        canvas.create_rectangle(x, y_top, x + bar_w, y_bot,
                                fill=color, outline="")

        # Value label on top
        canvas.create_text(x + bar_w // 2, y_top - 8, text=str(val),
                           fill="#ccc", font=(FONT_FAMILY, 9, "bold"))

        # X-axis label
        canvas.create_text(x + bar_w // 2, y_bot + 14, text=label,
                           fill="#888", font=(FONT_FAMILY, 8))

    # Y-axis line
    canvas.create_line(pad_left - 2, pad_top, pad_left - 2,
                       pad_top + chart_h, fill="#444", width=1)
    # X-axis line
    canvas.create_line(pad_left - 2, pad_top + chart_h,
                       width - pad_right, pad_top + chart_h, fill="#444", width=1)


def draw_trend_chart(canvas, data_points, width, height,
                     threshold=DEFAULT_THRESHOLD, title=""):
    """在 tkinter Canvas 上绘制质量趋势折线图。

    data_points: list of {"label": str, "avg_score": float, "below_rate": float}
    """
    canvas.delete("all")

    if title:
        canvas.create_text(width // 2, 16, text=title,
                           fill="#ccc", font=(FONT_FAMILY, 11, "bold"))

    if not data_points or len(data_points) < 2:
        canvas.create_text(width // 2, height // 2,
                           text="Insufficient data for trend",
                           fill="#666", font=(FONT_FAMILY, 10))
        return

    pad_left = 55
    pad_right = 20
    pad_top = 40
    pad_bottom = 45
    chart_w = width - pad_left - pad_right
    chart_h = height - pad_top - pad_bottom

    scores = [dp["avg_score"] for dp in data_points if dp.get("avg_score") is not None]
    if not scores:
        return
    min_score = max(0, min(min(scores), threshold) - 5)
    max_score = min(100, max(max(scores), threshold) + 2)
    score_range = max_score - min_score or 1

    n = len(data_points)

    # Axes
    canvas.create_line(pad_left, pad_top, pad_left, pad_top + chart_h, fill="#444", width=1)
    canvas.create_line(pad_left, pad_top + chart_h,
                       pad_left + chart_w, pad_top + chart_h, fill="#444", width=1)

    # Threshold line
    th_y = pad_top + chart_h - int((threshold - min_score) / score_range * chart_h)
    canvas.create_line(pad_left, th_y, pad_left + chart_w, th_y,
                       fill="#E74C3C", width=1, dash=(4, 4))
    canvas.create_text(pad_left - 5, th_y, anchor="e",
                       text=str(threshold), fill="#E74C3C",
                       font=(FONT_FAMILY, 8))

    # Plot points and lines
    points = []
    for i, dp in enumerate(data_points):
        s = dp.get("avg_score")
        if s is None:
            continue
        x = pad_left + int(i / (n - 1) * chart_w) if n > 1 else pad_left + chart_w // 2
        y = pad_top + chart_h - int((s - min_score) / score_range * chart_h)
        points.append((x, y, s, dp.get("label", "")))

    # Draw lines
    for i in range(len(points) - 1):
        canvas.create_line(points[i][0], points[i][1],
                           points[i + 1][0], points[i + 1][1],
                           fill="#4472C4", width=2)

    # Draw dots and labels
    for x, y, s, label in points:
        canvas.create_oval(x - 3, y - 3, x + 3, y + 3,
                           fill="#4472C4", outline="#fff", width=1)
        canvas.create_text(x, y - 12, text=f"{s:.1f}",
                           fill="#ccc", font=(FONT_FAMILY, 8))

    # X-axis labels (show first, last, and middle)
    step = max(1, len(points) // 5)
    for i, (x, y, s, label) in enumerate(points):
        if i % step == 0 or i == len(points) - 1:
            canvas.create_text(x, pad_top + chart_h + 14, text=label[:10],
                               fill="#888", font=(FONT_FAMILY, 7))

    # Y-axis labels
    for val in range(int(min_score), int(max_score) + 1, max(1, int(score_range // 4))):
        y = pad_top + chart_h - int((val - min_score) / score_range * chart_h)
        canvas.create_text(pad_left - 5, y, anchor="e",
                           text=str(val), fill="#888", font=(FONT_FAMILY, 8))

    # Legend
    canvas.create_line(width - 120, pad_top + 5, width - 100, pad_top + 5,
                       fill="#4472C4", width=2)
    canvas.create_text(width - 95, pad_top + 5, anchor="w",
                       text="Avg Score", fill="#bbb", font=(FONT_FAMILY, 8))
    canvas.create_line(width - 120, pad_top + 20, width - 100, pad_top + 20,
                       fill="#E74C3C", width=1, dash=(4, 4))
    canvas.create_text(width - 95, pad_top + 20, anchor="w",
                       text=f"Threshold ({threshold})", fill="#bbb",
                       font=(FONT_FAMILY, 8))


def draw_stacked_bar_chart(canvas, data, width, height, title=""):
    """绘制按语言分组的错误类别堆叠水平条形图。

    data: dict { "ja-JP": {"Accuracy": 5, "Fluency": 3, ...}, ... }
    """
    canvas.delete("all")

    if title:
        canvas.create_text(width // 2, 16, text=title,
                           fill="#ccc", font=(FONT_FAMILY, 11, "bold"))

    if not data:
        canvas.create_text(width // 2, height // 2, text="No Data",
                           fill="#666", font=(FONT_FAMILY, 10))
        return

    pad_left = 70
    pad_right = 20
    pad_top = 36
    pad_bottom = 20
    chart_w = width - pad_left - pad_right
    chart_h = height - pad_top - pad_bottom

    langs = list(data.keys())[:10]  # max 10 languages
    n = len(langs)
    if n == 0:
        return
    bar_h = max(12, chart_h // n - 4)

    # Global max for scale
    max_total = max(sum(data[lang].values()) for lang in langs) or 1

    # Category order for consistent coloring
    categories = list(ERROR_CATEGORIES)

    for i, lang in enumerate(langs):
        y = pad_top + i * (chart_h // n)
        # Language label
        canvas.create_text(pad_left - 4, y + bar_h // 2, anchor="e",
                           text=lang[:8], fill="#bbb", font=(FONT_FAMILY, 8))

        x_offset = pad_left
        for ci, cat in enumerate(categories):
            count = data[lang].get(cat, 0)
            if count == 0:
                continue
            seg_w = max(1, int(count / max_total * chart_w))
            color = PIE_COLORS[ci % len(PIE_COLORS)]
            canvas.create_rectangle(x_offset, y, x_offset + seg_w, y + bar_h,
                                    fill=color, outline="")
            if seg_w > 20:
                canvas.create_text(x_offset + seg_w // 2, y + bar_h // 2,
                                   text=str(count), fill="#fff",
                                   font=(FONT_FAMILY, 7))
            x_offset += seg_w

    # Legend
    lx = pad_left
    ly = pad_top + chart_h + 2
    for ci, cat in enumerate(categories):
        if lx + 80 > width:
            break
        color = PIE_COLORS[ci % len(PIE_COLORS)]
        canvas.create_rectangle(lx, ly, lx + 8, ly + 8, fill=color, outline="")
        canvas.create_text(lx + 11, ly + 4, anchor="w", text=cat[:12],
                           fill="#888", font=(FONT_FAMILY, 7))
        lx += 90


# ---------------------------------------------------------------------------
# 6) HTML 导出 — 质量报告
# ---------------------------------------------------------------------------
def write_quality_html(aggregated, filename, label):
    """生成质量报告 HTML"""
    a = aggregated
    threshold = a.get("threshold", DEFAULT_THRESHOLD)

    # Error distribution as chart bars (inline CSS)
    err_bars = ""
    err_total = sum(a["error_distribution"].values())
    for i, (cat, count) in enumerate(a["error_distribution"].items()):
        pct = (count / err_total * 100) if err_total else 0
        color = PIE_COLORS[i % len(PIE_COLORS)]
        err_bars += (
            f'<div style="display:flex;align-items:center;gap:8px;margin:4px 0;">'
            f'<span style="min-width:180px;font-size:13px;">{html_mod.escape(cat)}</span>'
            f'<div style="background:#eee;border-radius:6px;flex:1;height:20px;">'
            f'<div style="background:{color};border-radius:6px;height:100%;width:{pct:.1f}%;"></div>'
            f'</div>'
            f'<span style="min-width:60px;font-size:12px;color:#888;">{count} ({pct:.0f}%)</span>'
            f'</div>'
        )

    # Score distribution as chart
    score_bars = ""
    score_max = max(a["score_distribution"].values()) if a["score_distribution"] else 1
    if score_max == 0:
        score_max = 1
    for i, (label_text, count) in enumerate(a["score_distribution"].items()):
        pct = (count / score_max * 100)
        color = BAR_COLORS[i % len(BAR_COLORS)]
        score_bars += (
            f'<div style="display:flex;align-items:center;gap:8px;margin:4px 0;">'
            f'<span style="min-width:60px;font-size:13px;">{label_text}</span>'
            f'<div style="background:#eee;border-radius:6px;flex:1;height:20px;">'
            f'<div style="background:{color};border-radius:6px;height:100%;width:{pct:.1f}%;"></div>'
            f'</div>'
            f'<span style="min-width:40px;font-size:12px;color:#888;">{count}</span>'
            f'</div>'
        )

    # Language detail table
    lang_rows = ""
    for ld in a["by_language"]:
        avg_str = f'{ld["average_score"]}' if ld["average_score"] is not None else "—"
        lang_rows += (
            f'<tr>'
            f'<td>{html_mod.escape(ld["language"])}</td>'
            f'<td style="text-align:center">{ld["count"]}</td>'
            f'<td style="text-align:center">{avg_str}</td>'
            f'<td style="text-align:center;color:#E74C3C">{ld["below_threshold_pct"]}%</td>'
            f'<td style="text-align:center">{ld["refined_pct"]}%</td>'
            f'<td style="text-align:center">{ld["human_touched_pct"]}%</td>'
            f'<td style="text-align:center">{ld["warnings"]}</td>'
            f'</tr>'
        )

    # Low-score items table
    low_rows = ""
    for i, it in enumerate(a["low_items"]):
        score = it.get("final_score", "—")
        score_style = ' style="color:#E74C3C;font-weight:bold"' if isinstance(score, (int, float)) and score < 80 else ""
        source_type = it.get("_source_type", "")
        scope_name = it.get("_scope_name", "")
        low_rows += (
            f'<tr>'
            f'<td class="num">{i + 1}</td>'
            f'<td>{html_mod.escape(source_type)}</td>'
            f'<td>{html_mod.escape(scope_name)}</td>'
            f'<td class="key">{html_mod.escape(it.get("opus_id", ""))}</td>'
            f'<td>{html_mod.escape(it.get("target_language", ""))}</td>'
            f'<td>{html_mod.escape(it.get("source_text", "")[:120])}</td>'
            f'<td>{html_mod.escape(it.get("translated_text", "")[:120])}</td>'
            f'<td{score_style}>{score}</td>'
            f'<td>{html_mod.escape(it.get("error_category") or "—")}</td>'
            f'<td>{html_mod.escape(it.get("reason") or "")}</td>'
            f'</tr>'
        )

    page = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Quality Overview - {html_mod.escape(label)}</title>
<style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family:-apple-system,"Segoe UI",Roboto,Arial,sans-serif; background:#f5f6fa; padding:24px; color:#333; }}
    h1 {{ font-size:20px; margin-bottom:4px; }}
    h2 {{ font-size:16px; margin:20px 0 12px; color:#2c3e50; }}
    .meta {{ color:#666; font-size:14px; margin-bottom:16px; }}

    .cards {{ display:flex; gap:16px; margin-bottom:24px; flex-wrap:wrap; }}
    .card {{
        background:#fff; border-radius:10px; padding:20px; min-width:140px;
        box-shadow:0 2px 8px rgba(0,0,0,.06); text-align:center; flex:1;
    }}
    .card-value {{ font-size:24px; font-weight:bold; }}
    .card-label {{ font-size:11px; color:#888; text-transform:uppercase; letter-spacing:.5px; margin-top:4px; }}
    .card-accent {{ color:#4472C4; }}
    .card-warn {{ color:#E74C3C; }}
    .card-ok {{ color:#27AE60; }}
    .card-info {{ color:#16A085; }}

    .chart-section {{ background:#fff; border-radius:10px; padding:20px; margin-bottom:20px; box-shadow:0 2px 8px rgba(0,0,0,.06); }}

    table {{ border-collapse:collapse; width:100%; background:#fff; border-radius:6px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.08); font-size:13px; margin-bottom:20px; }}
    th {{ background:#4472C4; color:#fff; padding:10px 8px; text-align:left; white-space:nowrap; }}
    td {{ padding:8px; border-bottom:1px solid #e8e8e8; vertical-align:top; line-height:1.5; }}
    tr:hover {{ background:#f8f9fa; }}
    .num {{ text-align:center; color:#999; min-width:30px; }}
    .key {{ font-family:monospace; font-size:12px; max-width:200px; word-break:break-all; color:#555; }}
</style>
</head>
<body>
    <h1>Quality Overview</h1>
    <p class="meta">{html_mod.escape(label)}</p>

    <div class="cards">
        <div class="card">
            <div class="card-value card-accent">{a['total_tasks']}</div>
            <div class="card-label">Work Items</div>
        </div>
        <div class="card">
            <div class="card-value card-accent">{a['total_items']}</div>
            <div class="card-label">Segments</div>
        </div>
        <div class="card">
            <div class="card-value card-ok">{a['overall_avg_score']}</div>
            <div class="card-label">Avg Score</div>
        </div>
        <div class="card">
            <div class="card-value card-warn">{a['below_threshold_rate']}%</div>
            <div class="card-label">Below {threshold}</div>
        </div>
        <div class="card">
            <div class="card-value card-info">{a['refined_rate']}%</div>
            <div class="card-label">Refined</div>
        </div>
        <div class="card">
            <div class="card-value card-info">{a['human_touch_rate']}%</div>
            <div class="card-label">Human Touch</div>
        </div>
    </div>

    <div class="chart-section">
        <h2>Error Category Distribution</h2>
        {err_bars}
    </div>

    <div class="chart-section">
        <h2>Score Distribution</h2>
        {score_bars}
    </div>

    <h2>By Language Breakdown</h2>
    <table>
        <thead><tr>
            <th>Language</th><th>Segments</th><th>Avg Score</th>
            <th>Below {threshold}%</th><th>Refined %</th><th>Human Touch %</th><th>Warnings</th>
        </tr></thead>
        <tbody>{lang_rows}</tbody>
    </table>

    <h2>Low-Score Items (Score &lt; {threshold})</h2>
    <table>
        <thead><tr>
            <th>#</th><th>Type</th><th>Task/MR</th><th>String Key</th><th>Language</th>
            <th>Source</th><th>Translated</th><th>Score</th>
            <th>Error Category</th><th>Reason</th>
        </tr></thead>
        <tbody>{low_rows if low_rows else '<tr><td colspan="10" style="text-align:center;color:#888;">No low-score items</td></tr>'}</tbody>
    </table>
</body>
</html>"""

    with open(filename, "w", encoding="utf-8") as f:
        f.write(page)


# ---------------------------------------------------------------------------
# 7) Excel 导出 — 质量报告
# ---------------------------------------------------------------------------
def write_quality_excel(aggregated, filename):
    """生成质量报告 Excel（多 Sheet）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("错误: 缺少 openpyxl 包，请先运行: pip install openpyxl")
        return

    a = aggregated
    threshold = a.get("threshold", DEFAULT_THRESHOLD)
    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    summary_data = [
        ("Metric", "Value"),
        ("Total Work Items", a.get("total_tasks", 0)),
        ("Total Segments", a.get("total_items", 0)),
        ("Average Score", a.get("overall_avg_score", 0)),
        (f"Below Threshold (< {threshold})", a.get("low_score_count", 0)),
        ("Below Threshold Rate", f'{a.get("below_threshold_rate", 0)}%'),
        ("Refined Rate", f'{a.get("refined_rate", 0)}%'),
        ("Human Touch Rate", f'{a.get("human_touch_rate", 0)}%'),
    ]
    for row_i, (metric, val) in enumerate(summary_data, 1):
        c1 = ws.cell(row=row_i, column=1, value=metric)
        c2 = ws.cell(row=row_i, column=2, value=val)
        if row_i == 1:
            c1.fill = header_fill
            c1.font = header_font
            c2.fill = header_fill
            c2.font = header_font
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

    # --- Sheet 2: By Language ---
    ws2 = wb.create_sheet("By Language")
    lang_headers = ["Language", "Segments", "Avg Score", f"Below {threshold} %",
                    "Refined %", "Human Touch %", "Warnings"]
    for col_i, h in enumerate(lang_headers, 1):
        cell = ws2.cell(row=1, column=col_i, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_i, ld in enumerate(a.get("by_language", []), 2):
        ws2.cell(row=row_i, column=1, value=ld["language"])
        ws2.cell(row=row_i, column=2, value=ld["count"])
        ws2.cell(row=row_i, column=3, value=ld["average_score"] if ld["average_score"] else "")
        ws2.cell(row=row_i, column=4, value=ld["below_threshold_pct"])
        ws2.cell(row=row_i, column=5, value=ld["refined_pct"])
        ws2.cell(row=row_i, column=6, value=ld["human_touched_pct"])
        ws2.cell(row=row_i, column=7, value=ld["warnings"])

    for i, w in enumerate([15, 10, 12, 14, 12, 14, 10], 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    # --- Sheet 3: Low-Score Items ---
    ws3 = wb.create_sheet("Low Score Items")
    low_headers = ["#", "Source Type", "Task/MR", "String Key", "Language",
                   "Source", "Translated", "Score", "Error Category", "Reason"]
    for col_i, h in enumerate(low_headers, 1):
        cell = ws3.cell(row=1, column=col_i, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_i, it in enumerate(a.get("low_items", []), 2):
        ws3.cell(row=row_i, column=1, value=row_i - 1)
        ws3.cell(row=row_i, column=2, value=it.get("_source_type", ""))
        ws3.cell(row=row_i, column=3, value=it.get("_scope_name", ""))
        ws3.cell(row=row_i, column=4, value=it.get("opus_id", ""))
        ws3.cell(row=row_i, column=5, value=it.get("target_language", ""))
        ws3.cell(row=row_i, column=6, value=it.get("source_text", ""))
        ws3.cell(row=row_i, column=7, value=it.get("translated_text", ""))
        score = it.get("final_score")
        score_cell = ws3.cell(row=row_i, column=8, value=score if score is not None else "")
        if score is not None and score < 80:
            score_cell.font = Font(color="E74C3C", bold=True)
        ws3.cell(row=row_i, column=9, value=it.get("error_category") or "")
        ws3.cell(row=row_i, column=10, value=it.get("reason") or "")

    for i, w in enumerate([6, 10, 20, 30, 10, 40, 40, 8, 18, 30], 1):
        col_letter = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
        ws3.column_dimensions[col_letter].width = w

    wb.save(filename)


# ---------------------------------------------------------------------------
# 8) 统一保存入口
# ---------------------------------------------------------------------------
def save_quality_file(aggregated, filename, label, fmt):
    """保存质量报告，文件被占用时自动加序号"""
    base, ext = os.path.splitext(filename)
    save_path = filename
    for attempt in range(100):
        try:
            if fmt == "html":
                write_quality_html(aggregated, save_path, label)
            else:
                write_quality_excel(aggregated, save_path)
            print(f"已导出: {save_path}")
            if fmt == "html":
                from export_gui import open_in_browser
                open_in_browser(save_path)
            return save_path
        except PermissionError:
            attempt_num = attempt + 1
            save_path = f"{base}_{attempt_num}{ext}"
            print(f"  文件被占用，尝试保存为: {save_path}")
    return None
