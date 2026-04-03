"""
Tranzor 变更导出工具
====================
通过 Tranzor HTTP API 导出所有手动翻译编辑记录（全量）。
默认输出 HTML（带可视化 Diff），加 --xlsx 输出 Excel。

用法:
    python export_changes.py                 # 导出所有变更 (HTML)
    python export_changes.py --xlsx          # 输出 Excel 格式
    python export_changes.py --task 123      # 只导出某个 task 的变更

首次使用前:
    pip install openpyxl requests
"""

import argparse
import difflib
import html
import os
import sys
import webbrowser
from datetime import date

try:
    import requests
except ImportError:
    print("错误: 缺少 requests 包，请先运行:")
    print("  pip install openpyxl requests")
    sys.exit(1)

# ---------------------------------------------------------------------------
# 配置 — 你的 Tranzor 平台地址（就是你浏览器里打开的那个地址）
# ---------------------------------------------------------------------------
TRANZOR_URL = "http://tranzor-platform.int.rclabenv.com"

API = f"{TRANZOR_URL}/api/v1/legacy"


# ---------------------------------------------------------------------------
# 1) 获取 task 列表
# ---------------------------------------------------------------------------
def fetch_tasks(created_after=None, created_before=None):
    """获取所有 task（分页遍历）"""
    all_tasks = []
    offset = 0
    limit = 200

    while True:
        params = {"limit": limit, "offset": offset, "status": "Completed"}
        if created_after:
            params["created_after"] = created_after
        if created_before:
            params["created_before"] = created_before

        resp = _api_get(f"{API}/tasks", params=params)
        resp.raise_for_status()
        data = resp.json()

        tasks = data.get("tasks", [])
        all_tasks.extend(tasks)

        if len(tasks) < limit:
            break
        offset += limit

    return all_tasks


# ---------------------------------------------------------------------------
# 2) 获取某个 task 下所有 "Manual Edit" 类型的翻译
# ---------------------------------------------------------------------------
def fetch_manual_edits(task_id):
    """获取某个 task 中所有手动编辑过的翻译"""
    manual_translations = []
    offset = 0
    limit = 200

    while True:
        params = {"limit": limit, "offset": offset}
        resp = _api_get(
            f"{API}/tasks/{task_id}/translations", params=params
        )
        resp.raise_for_status()
        data = resp.json()

        for entry in data.get("entries", []):
            t_type = entry.get("translation_type", "")
            if t_type in ("Manual Edit", "LLM Retranslate"):
                manual_translations.append(entry)

        total = data.get("total", 0)
        offset += limit
        if offset >= total:
            break

    return manual_translations


# ---------------------------------------------------------------------------
# 3) 获取单条翻译的 edit log
# ---------------------------------------------------------------------------
def fetch_edit_logs(task_id, translation_id):
    """获取单条翻译的编辑历史"""
    resp = _api_get(
        f"{API}/tasks/{task_id}/translations/{translation_id}/edit-logs",
    )
    resp.raise_for_status()
    return resp.json()


# ---------------------------------------------------------------------------
# 4) 主流程：遍历 tasks → 找 Manual Edit → 拉 edit logs（并发优化版）
# ---------------------------------------------------------------------------
# 共享 HTTP session（连接池复用）
_session = requests.Session()

# 并发工作线程数（不宜过高，否则服务端可能超时）
MAX_WORKERS = 4

# 最大重试次数
MAX_RETRIES = 3


def _api_get(url, **kwargs):
    """带重试的 GET 请求，遇到超时/连接错误自动重试"""
    import time
    kwargs.setdefault("timeout", 30)
    for attempt in range(MAX_RETRIES):
        try:
            return _session.get(url, **kwargs)
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError) as e:
            if attempt < MAX_RETRIES - 1:
                wait = 2 ** attempt  # 1s, 2s, 4s
                print(f"    ⚠ 请求超时，{wait}s 后重试 ({attempt+1}/{MAX_RETRIES})...")
                time.sleep(wait)
            else:
                raise


def _process_single_task(task_info):
    """处理单个 task：拉取手动编辑 → 并发拉取 edit logs → 返回结果列表"""
    tid, tname, idx, total = task_info

    try:
        manual = fetch_manual_edits(tid)
    except Exception as e:
        print(f"  ⚠ [{idx}/{total}] Task '{tname}' - 获取翻译列表失败: {e}")
        return []

    if not manual:
        return []

    print(f"  [{idx}/{total}] Task '{tname}' - 发现 {len(manual)} 条手动编辑")

    # 并发拉取该 task 下所有 edit logs
    from concurrent.futures import ThreadPoolExecutor, as_completed

    entries_with_logs = []

    def _fetch_one_log(entry):
        tr_id = entry.get("translation_id")
        if not tr_id:
            return []
        try:
            logs = fetch_edit_logs(tid, tr_id)
        except Exception as e:
            print(f"    ⚠ 获取 edit log 失败 (task={tid}, tr={tr_id}): {e}")
            return []
        row_results = []
        for log in logs:
            row_results.append({
                "edit_id": log.get("id", ""),
                "edit_time": log.get("created_at", ""),
                "editor": log.get("user_name", ""),
                "task_id": tid,
                "task_name": tname,
                "language": entry.get("target_language", ""),
                "string_key": entry.get("opus_id", ""),
                "source_text": entry.get("source_text", ""),
                "before": log.get("original_text", ""),
                "after": log.get("edited_text", ""),
                "notes": log.get("notes", ""),
            })
        return row_results

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = [pool.submit(_fetch_one_log, entry) for entry in manual]
        for f in as_completed(futures):
            entries_with_logs.extend(f.result())

    return entries_with_logs


def collect_changes(task_id=None):
    """收集所有变更记录（全量，并发优化）"""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # 获取 task 列表
    if task_id:
        tasks = [{"id": task_id, "task_name": f"Task {task_id}"}]
        try:
            resp = requests.get(f"{API}/tasks/{task_id}", timeout=30)
            resp.raise_for_status()
            task_data = resp.json()
            tasks[0]["task_name"] = task_data.get("task_name", f"Task {task_id}")
        except Exception:
            pass
    else:
        print("  正在获取 task 列表...")
        tasks = fetch_tasks()
        print(f"  找到 {len(tasks)} 个已完成的 task")

    total = len(tasks)
    task_infos = [
        (task["id"], task.get("task_name", ""), i + 1, total)
        for i, task in enumerate(tasks)
    ]

    # 并发处理所有 task
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_process_single_task, info): info for info in task_infos}
        for f in as_completed(futures):
            results.extend(f.result())

    return results


# ---------------------------------------------------------------------------
# 5) 单词级别 Diff（纯文本版，用于 Excel）
# ---------------------------------------------------------------------------
def word_diff_text(before, after):
    """纯文本 diff: [-删除] [+新增]"""
    before_words = before.split()
    after_words = after.split()
    sm = difflib.SequenceMatcher(None, before_words, after_words)
    parts = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            parts.append(" ".join(before_words[i1:i2]))
        elif tag == "replace":
            parts.append("[-" + " ".join(before_words[i1:i2]) + "]")
            parts.append("[+" + " ".join(after_words[j1:j2]) + "]")
        elif tag == "delete":
            parts.append("[-" + " ".join(before_words[i1:i2]) + "]")
        elif tag == "insert":
            parts.append("[+" + " ".join(after_words[j1:j2]) + "]")
    return " ".join(parts)


# ---------------------------------------------------------------------------
# 6) 单词级别 Diff（HTML 版，红色删除线 + 绿色高亮）
# ---------------------------------------------------------------------------
def word_diff_html(before, after):
    """HTML diff: 红色删除线 = 删除, 绿色高亮 = 新增"""
    before_words = before.split()
    after_words = after.split()
    sm = difflib.SequenceMatcher(None, before_words, after_words)
    parts = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            parts.append(html.escape(" ".join(before_words[i1:i2])))
        elif tag == "replace":
            old = html.escape(" ".join(before_words[i1:i2]))
            new = html.escape(" ".join(after_words[j1:j2]))
            parts.append(f'<del>{old}</del>')
            parts.append(f'<ins>{new}</ins>')
        elif tag == "delete":
            old = html.escape(" ".join(before_words[i1:i2]))
            parts.append(f'<del>{old}</del>')
        elif tag == "insert":
            new = html.escape(" ".join(after_words[j1:j2]))
            parts.append(f'<ins>{new}</ins>')
    return " ".join(parts)


def write_html(rows, filename, label):
    """生成带可视化 Diff 的 HTML 报告，按 Editor 分组，支持筛选和交互式 TMX 导出"""
    import json
    from collections import OrderedDict

    # 按 Editor 分组，保持出现顺序
    groups = OrderedDict()
    for r in rows:
        editor = r.get("editor") or "(unknown)"
        groups.setdefault(editor, []).append(r)

    # 为每个 Editor 分配一个颜色
    editor_colors = [
        ("#4472C4", "#e8eef7"),  # 蓝
        ("#E67E22", "#fdf2e6"),  # 橙
        ("#27AE60", "#e8f5ee"),  # 绿
        ("#8E44AD", "#f3eaf8"),  # 紫
        ("#E74C3C", "#fceaea"),  # 红
        ("#16A085", "#e6f4f1"),  # 青
        ("#D4AC0D", "#faf6e4"),  # 金
        ("#2C3E50", "#eaecee"),  # 灰
    ]

    # --- 构建 JSON 数据供 JS 使用（必须按分组后的顺序，与 HTML 表格一致） ---
    js_rows = []
    all_langs = set()
    for editor_rows_list in groups.values():
        for r in editor_rows_list:
            js_rows.append({
                "source_text": r["source_text"],
                "before": r["before"],
                "after": r["after"],
                "language": r["language"],
                "string_key": r["string_key"],
                "edit_time": r["edit_time"],
                "editor": r.get("editor") or "(unknown)",
                "task_id": r.get("task_id", ""),
                "task_name": r["task_name"],
            })
            all_langs.add(r["language"])
    rows_json = json.dumps(js_rows, ensure_ascii=False)
    langs_json = json.dumps(sorted(all_langs), ensure_ascii=False)

    sections_html = ""
    global_idx = 0
    for editor_i, (editor_name, editor_rows) in enumerate(groups.items()):
        color_pair = editor_colors[editor_i % len(editor_colors)]
        header_bg = color_pair[0]
        section_bg = color_pair[1]

        table_rows = ""
        for r in editor_rows:
            diff = word_diff_html(r["before"], r["after"])
            table_rows += f"""
            <tr>
                <td class="cb-cell"><input type="checkbox" class="row-cb" data-idx="{global_idx}"></td>
                <td class="num">{global_idx + 1}</td>
                <td class="time">{html.escape(r['edit_time'][:19].replace('T', ' '))}</td>
                <td class="task-id">{r.get('task_id', '')}</td>
                <td class="task">{html.escape(r['task_name'])}</td>
                <td class="lang">{html.escape(r['language'])}</td>
                <td class="key">{html.escape(r['string_key'])}</td>
                <td>{html.escape(r['source_text'])}</td>
                <td class="before">{html.escape(r['before'])}</td>
                <td class="after">{html.escape(r['after'])}</td>
                <td class="diff">{diff}</td>
                <td>{html.escape(r['notes'] or '')}</td>
            </tr>"""
            global_idx += 1

        sections_html += f"""
        <div class="editor-section" style="background: {section_bg}; border-left: 4px solid {header_bg}; border-radius: 8px; padding: 16px; margin-bottom: 24px;">
            <h2 style="color: {header_bg}; margin-bottom: 12px; font-size: 17px;">
                👤 {html.escape(editor_name)}
                <span class="count" style="background: {header_bg};">{len(editor_rows)} edits</span>
            </h2>
            <table>
                <thead>
                    <tr style="background: {header_bg};">
                        <th class="cb-cell"><input type="checkbox" class="section-cb"></th>
                        <th>#</th><th>Time</th><th>Task ID</th><th>Task</th><th>Lang</th>
                        <th>String Key</th><th>Source (en-US)</th><th>Before</th><th>After</th>
                        <th>Diff</th><th>Notes</th>
                    </tr>
                </thead>
                <tbody>{table_rows}
                </tbody>
            </table>
        </div>"""

    # 顶部 Editor 索引
    toc_items = ""
    for editor_i, (editor_name, editor_rows) in enumerate(groups.items()):
        color_pair = editor_colors[editor_i % len(editor_colors)]
        toc_items += f'<span style="display:inline-block; background:{color_pair[0]}; color:#fff; border-radius:16px; padding:4px 14px; margin:4px; font-size:13px;">{html.escape(editor_name)} ({len(editor_rows)})</span>'

    page = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Tranzor Changes - {html.escape(label)}</title>
<style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: -apple-system, "Segoe UI", Roboto, Arial, sans-serif; background: #f5f6fa; padding: 24px; padding-top: 72px; color: #333; }}
    h1 {{ font-size: 20px; margin-bottom: 4px; }}
    .meta {{ color: #666; font-size: 14px; margin-bottom: 8px; }}
    .toc {{ margin-bottom: 20px; }}
    .count {{ display: inline-block; color: #fff; border-radius: 12px; padding: 2px 10px; font-size: 13px; margin-left: 8px; }}
    table {{ border-collapse: collapse; width: 100%; background: #fff; border-radius: 6px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,.08); font-size: 13px; }}
    th {{ color: #fff; padding: 10px 8px; text-align: left; white-space: nowrap; }}
    td {{ padding: 8px; border-bottom: 1px solid #e8e8e8; vertical-align: top; line-height: 1.5; }}
    tr:hover {{ background: rgba(255,255,255,.6); }}
    .num {{ text-align: center; color: #999; min-width: 30px; }}
    .time {{ white-space: nowrap; font-size: 12px; color: #666; }}
    .task-id {{ text-align: center; font-family: monospace; font-size: 12px; color: #888; white-space: nowrap; }}
    .task {{ max-width: 200px; }}
    .lang {{ text-align: center; }}
    .key {{ font-family: monospace; font-size: 12px; max-width: 260px; word-break: break-all; color: #555; }}
    .before {{ background: #fff8e6; }}
    .after {{ background: #edf7ed; }}
    .diff {{ background: #f0f5ff; min-width: 250px; }}
    del {{ background: #fdd; color: #c00; text-decoration: line-through; padding: 1px 3px; border-radius: 3px; }}
    ins {{ background: #dfd; color: #060; text-decoration: none; padding: 1px 3px; border-radius: 3px; }}

    /* --- Toolbar --- */
    .tmx-toolbar {{
        position: fixed; top: 0; left: 0; right: 0; z-index: 1000;
        background: #fff; border-bottom: 1px solid #dde; padding: 10px 24px;
        display: flex; align-items: center; gap: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,.08);
    }}
    .tmx-toolbar .btn {{
        border: none; border-radius: 6px; padding: 7px 16px; cursor: pointer;
        font-size: 13px; font-weight: 600; transition: background .15s, transform .1s;
    }}
    .tmx-toolbar .btn:active {{ transform: scale(.97); }}
    .btn-select {{ background: #eef1f5; color: #333; }}
    .btn-select:hover {{ background: #dde3ea; }}
    .btn-filter {{ background: #eef1f5; color: #333; }}
    .btn-filter:hover {{ background: #dde3ea; }}
    .btn-filter.active {{ background: #4472C4; color: #fff; }}
    .btn-export {{ background: #4472C4; color: #fff; }}
    .btn-export:hover {{ background: #3461b0; }}
    .btn-export:disabled {{ background: #a8b8d0; cursor: not-allowed; opacity: .7; }}
    .badge {{ background: #4472C4; color: #fff; border-radius: 12px; padding: 2px 10px; font-size: 12px; font-weight: 700; min-width: 28px; text-align: center; }}
    .badge.zero {{ background: #bbb; }}
    .toolbar-sep {{ width: 1px; height: 24px; background: #dde; }}

    /* checkbox column */
    .cb-cell {{ text-align: center; width: 32px; min-width: 32px; }}
    .cb-cell input[type="checkbox"] {{ width: 15px; height: 15px; cursor: pointer; accent-color: #4472C4; }}
    tr.row-selected {{ background: #e8eef7 !important; }}
    tr.row-hidden {{ display: none !important; }}

    /* --- Filter Panel --- */
    .filter-panel {{
        background: #1e293b; color: #cbd5e1; border-radius: 10px;
        margin-bottom: 20px; overflow: hidden;
        max-height: 0; opacity: 0; transition: max-height .35s ease, opacity .25s ease, margin .25s ease, padding .25s ease;
        padding: 0 20px;
    }}
    .filter-panel.open {{
        max-height: 800px; opacity: 1; padding: 20px;
    }}
    .fp-row {{
        display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 16px;
    }}
    .fp-row:last-child {{ margin-bottom: 0; }}

    /* Simple filters (Time, Task, Lang) */
    .fp-simple {{ display: flex; align-items: center; gap: 8px; }}
    .fp-simple label {{ font-size: 12px; font-weight: 600; color: #94a3b8; text-transform: uppercase; letter-spacing: .5px; white-space: nowrap; }}
    .fp-simple input[type="text"] {{
        background: #0f172a; border: 1px solid #334155; border-radius: 6px;
        color: #e2e8f0; padding: 6px 10px; font-size: 13px; width: 170px;
        outline: none; transition: border-color .15s;
    }}
    .fp-simple input[type="text"]:focus {{ border-color: #4472C4; }}
    .fp-simple select {{
        background: #0f172a; border: 1px solid #334155; border-radius: 6px;
        color: #e2e8f0; padding: 6px 10px; font-size: 13px; width: 130px;
        outline: none; cursor: pointer;
    }}

    /* TextFilter card */
    .tf-card {{
        background: #0f172a; border: 1px solid #334155; border-radius: 8px;
        padding: 12px 14px; min-width: 280px; flex: 1;
    }}
    .tf-header {{
        display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;
    }}
    .tf-title {{ font-size: 12px; font-weight: 700; color: #38bdf8; text-transform: uppercase; letter-spacing: .5px; }}
    .tf-logic {{ display: flex; gap: 0; }}
    .tf-logic button {{
        padding: 3px 12px; font-size: 11px; font-weight: 700; border: 1px solid #334155;
        cursor: pointer; background: transparent; color: #64748b; transition: all .15s;
    }}
    .tf-logic button:first-child {{ border-radius: 4px 0 0 4px; }}
    .tf-logic button:last-child {{ border-radius: 0 4px 4px 0; border-left: 0; }}
    .tf-logic button.active {{ background: #0ea5e9; color: #fff; border-color: #0ea5e9; }}

    .tf-input-row {{
        display: flex; align-items: center; gap: 8px; margin-bottom: 8px;
    }}
    .tf-label {{
        font-size: 11px; font-weight: 700; padding: 3px 8px; border-radius: 4px;
        text-align: center; min-width: 36px;
    }}
    .tf-label.pos {{ background: #fff; color: #1e293b; }}
    .tf-label.neg {{ background: transparent; color: #f87171; border: 1px solid #7f1d1d; }}
    .tf-input {{
        flex: 1; background: #1e293b; border: 1px solid #334155; border-radius: 6px;
        color: #e2e8f0; padding: 6px 10px; font-size: 13px; outline: none;
    }}
    .tf-input:focus {{ border-color: #4472C4; }}
    .tf-input.neg-input {{ border-color: #7f1d1d; }}
    .tf-input.neg-input:focus {{ border-color: #ef4444; }}

    .tf-opts {{
        display: flex; flex-wrap: wrap; gap: 6px 14px; margin-top: 6px;
    }}
    .tf-opt {{
        display: flex; align-items: center; gap: 4px; font-size: 11px; color: #94a3b8; cursor: pointer; user-select: none;
    }}
    .tf-opt input {{ accent-color: #4472C4; cursor: pointer; }}
    .tf-opt .regex-label {{ color: #eab308; }}
    .tf-opt .neg-label {{ color: #94a3b8; }}

    /* Filter action buttons */
    .fp-actions {{
        display: flex; gap: 10px; align-items: center; margin-left: auto;
    }}
    .fp-actions .btn {{
        border: none; border-radius: 6px; padding: 7px 18px; cursor: pointer;
        font-size: 13px; font-weight: 600;
    }}
    .btn-apply {{ background: #0ea5e9; color: #fff; }}
    .btn-apply:hover {{ background: #0284c7; }}
    .btn-clear {{ background: #334155; color: #cbd5e1; }}
    .btn-clear:hover {{ background: #475569; }}
    .filter-info {{ font-size: 12px; color: #94a3b8; }}
</style>
</head>
<body>

<!-- Toolbar -->
<div class="tmx-toolbar">
    <button class="btn btn-select" id="btnSelectAll" onclick="toggleSelectAll()">☑ Select All</button>
    <div class="toolbar-sep"></div>
    <span style="font-size:13px;color:#555;">Selected:</span>
    <span class="badge zero" id="selCount">0</span>
    <div class="toolbar-sep"></div>
    <button class="btn btn-export" id="btnExport" onclick="exportTMX()" disabled>📦 Export TMX</button>
    <span id="exportStatus" style="font-size:12px;color:#888;"></span>
    <div class="toolbar-sep"></div>
    <button class="btn btn-filter" id="btnFilterToggle" onclick="toggleFilterPanel()">🔍 Filters</button>
    <span class="filter-info" id="filterInfo"></span>
</div>

    <h1>Tranzor Translation Changes <span class="count" style="background:#4472C4;">{len(rows)} edits</span></h1>
    <p class="meta">{html.escape(label)}</p>

<!-- Filter Panel -->
<div class="filter-panel" id="filterPanel">
    <div class="fp-row">
        <div class="fp-simple">
            <label>Time</label>
            <input type="text" id="fTime" placeholder="e.g. 2026-03 or 07:37">
        </div>
        <div class="fp-simple">
            <label>Task ID</label>
            <input type="text" id="fTaskId" placeholder="e.g. 12345">
        </div>
        <div class="fp-simple">
            <label>Task</label>
            <input type="text" id="fTask" placeholder="Keyword…">
        </div>
        <div class="fp-simple">
            <label>Lang</label>
            <select id="fLang"><option value="">All</option></select>
        </div>
        <div class="fp-actions">
            <button class="btn btn-apply" onclick="applyFilters()">▶ Apply</button>
            <button class="btn btn-clear" onclick="clearFilters()">✕ Clear</button>
        </div>
    </div>
    <div class="fp-row" id="tfRow">
        <!-- 4 TextFilter cards rendered by JS -->
    </div>
</div>

    <div class="toc">{toc_items}</div>
    {sections_html}

<script src="https://cdn.jsdelivr.net/npm/jszip@3/dist/jszip.min.js"></script>
<script>
// ============================================================
// Row data & selection state
// ============================================================
const ROWS = {rows_json};
const ALL_LANGS = {langs_json};
let allSelected = false;

// Populate lang dropdown
(function() {{
    const sel = document.getElementById('fLang');
    ALL_LANGS.forEach(l => {{
        const opt = document.createElement('option');
        opt.value = l; opt.textContent = l;
        sel.appendChild(opt);
    }});
}})();

// ============================================================
// TextFilter card rendering
// ============================================================
const TF_FIELDS = [
    {{ id: 'stringKey', label: 'STRING KEY', dataKey: 'string_key' }},
    {{ id: 'source',    label: 'SOURCE (EN-US)', dataKey: 'source_text' }},
    {{ id: 'before',    label: 'BEFORE', dataKey: 'before' }},
    {{ id: 'after',     label: 'AFTER',  dataKey: 'after' }},
];

(function renderTFCards() {{
    const container = document.getElementById('tfRow');
    TF_FIELDS.forEach(f => {{
        container.innerHTML += `
        <div class="tf-card" data-field="${{f.id}}">
            <div class="tf-header">
                <span class="tf-title">${{f.label}}</span>
                <div class="tf-logic">
                    <button class="active" data-val="AND" onclick="toggleLogic(this)">AND</button>
                    <button data-val="OR" onclick="toggleLogic(this)">OR</button>
                </div>
            </div>
            <div class="tf-input-row">
                <span class="tf-label pos">Pos</span>
                <input class="tf-input" data-role="pos" placeholder="Positive keyword…">
            </div>
            <div class="tf-input-row">
                <span class="tf-label neg">Neg</span>
                <input class="tf-input neg-input" data-role="neg" placeholder="Negative keyword (Exclude)…">
            </div>
            <div class="tf-opts">
                <label class="tf-opt"><input type="checkbox" data-role="matchWhole"> Match whole</label>
                <label class="tf-opt"><input type="checkbox" data-role="posCaseSensitive"> Match case <span class="neg-label">(Pos)</span></label>
                <label class="tf-opt"><input type="checkbox" data-role="posRegex"> <span class="regex-label">Regex (Pos)</span></label>
                <label class="tf-opt"><input type="checkbox" data-role="negCaseSensitive"> Match case <span class="neg-label">(Neg)</span></label>
                <label class="tf-opt"><input type="checkbox" data-role="negRegex"> <span class="regex-label">Regex (Neg)</span></label>
            </div>
        </div>`;
    }});
}})();

function toggleLogic(btn) {{
    const group = btn.parentElement;
    group.querySelectorAll('button').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
}}

function toggleFilterPanel() {{
    const panel = document.getElementById('filterPanel');
    const btn = document.getElementById('btnFilterToggle');
    panel.classList.toggle('open');
    btn.classList.toggle('active');
}}

// ============================================================
// Filter engine
// ============================================================
function getTextFilterState(cardEl) {{
    const logicBtn = cardEl.querySelector('.tf-logic button.active');
    return {{
        logic: logicBtn ? logicBtn.dataset.val : 'AND',
        pos: cardEl.querySelector('[data-role="pos"]').value,
        neg: cardEl.querySelector('[data-role="neg"]').value,
        posCaseSensitive: cardEl.querySelector('[data-role="posCaseSensitive"]').checked,
        posRegex: cardEl.querySelector('[data-role="posRegex"]').checked,
        negCaseSensitive: cardEl.querySelector('[data-role="negCaseSensitive"]').checked,
        negRegex: cardEl.querySelector('[data-role="negRegex"]').checked,
        matchWhole: cardEl.querySelector('[data-role="matchWhole"]').checked,
    }};
}}

function testMatch(text, keyword, caseSensitive, isRegex, matchWhole) {{
    if (!keyword) return null; // null = no condition
    try {{
        let pattern;
        if (isRegex) {{
            pattern = new RegExp(keyword, caseSensitive ? '' : 'i');
        }} else {{
            const escaped = keyword.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&');
            const p = matchWhole ? ('\\\\b' + escaped + '\\\\b') : escaped;
            pattern = new RegExp(p, caseSensitive ? '' : 'i');
        }}
        return pattern.test(text);
    }} catch(e) {{
        return false;
    }}
}}

function evaluateTextFilter(text, tf) {{
    const posResult = testMatch(text, tf.pos, tf.posCaseSensitive, tf.posRegex, tf.matchWhole);
    const negResult = testMatch(text, tf.neg, tf.negCaseSensitive, tf.negRegex, tf.matchWhole);

    if (tf.logic === 'AND') {{
        const posPass = (posResult === null) ? true : posResult;
        const negPass = (negResult === null) ? true : !negResult;
        return posPass && negPass;
    }} else {{
        // OR: pass if either positive matches, or negative excludes correctly
        if (posResult === null && negResult === null) return true;
        const posPass = posResult === true;
        const negPass = (negResult !== null) && !negResult;
        return posPass || negPass;
    }}
}}

function applyFilters() {{
    const fTime = document.getElementById('fTime').value.trim();
    const fTaskId = document.getElementById('fTaskId').value.trim();
    const fTask = document.getElementById('fTask').value.trim();
    const fLang = document.getElementById('fLang').value;

    // Collect TextFilter states
    const tfCards = document.querySelectorAll('.tf-card');
    const tfStates = {{}};
    tfCards.forEach(card => {{
        const field = card.dataset.field;
        tfStates[field] = getTextFilterState(card);
    }});

    const allRows = document.querySelectorAll('input.row-cb');
    let visibleCount = 0;
    let totalCount = allRows.length;

    allRows.forEach(cb => {{
        const idx = parseInt(cb.dataset.idx);
        const row = ROWS[idx];
        const tr = cb.closest('tr');
        let pass = true;

        // Time filter
        if (fTime && pass) {{
            const timeStr = (row.edit_time || '').replace('T', ' ');
            if (timeStr.indexOf(fTime) === -1) pass = false;
        }}

        // Task ID filter
        if (fTaskId && pass) {{
            if (String(row.task_id).indexOf(fTaskId) === -1) pass = false;
        }}

        // Task filter
        if (fTask && pass) {{
            if (row.task_name.toLowerCase().indexOf(fTask.toLowerCase()) === -1) pass = false;
        }}

        // Lang filter
        if (fLang && pass) {{
            if (row.language !== fLang) pass = false;
        }}

        // TextFilter fields
        if (pass) {{
            TF_FIELDS.forEach(f => {{
                if (!pass) return;
                const tf = tfStates[f.id];
                if (!tf || (!tf.pos && !tf.neg)) return; // skip empty filters
                const text = row[f.dataKey] || '';
                if (!evaluateTextFilter(text, tf)) pass = false;
            }});
        }}

        if (pass) {{
            tr.classList.remove('row-hidden');
            visibleCount++;
        }} else {{
            tr.classList.add('row-hidden');
        }}
    }});

    // Update filter info
    const info = document.getElementById('filterInfo');
    if (visibleCount < totalCount) {{
        info.textContent = 'Showing ' + visibleCount + ' / ' + totalCount;
        info.style.color = '#e67e22';
    }} else {{
        info.textContent = '';
    }}

    // Sync section header checkboxes for visible rows
    document.querySelectorAll('.editor-section table').forEach(table => {{
        const visibleCbs = table.querySelectorAll('tbody tr:not(.row-hidden) input.row-cb');
        const checkedCbs = table.querySelectorAll('tbody tr:not(.row-hidden) input.row-cb:checked');
        const sectionCb = table.querySelector('input.section-cb');
        if (sectionCb) {{
            sectionCb.checked = visibleCbs.length > 0 && visibleCbs.length === checkedCbs.length;
        }}
    }});
    updateBadge();
}}

function clearFilters() {{
    document.getElementById('fTime').value = '';
    document.getElementById('fTaskId').value = '';
    document.getElementById('fTask').value = '';
    document.getElementById('fLang').value = '';

    document.querySelectorAll('.tf-card').forEach(card => {{
        card.querySelector('[data-role="pos"]').value = '';
        card.querySelector('[data-role="neg"]').value = '';
        card.querySelectorAll('.tf-opts input[type="checkbox"]').forEach(cb => cb.checked = false);
        const btns = card.querySelectorAll('.tf-logic button');
        btns.forEach(b => b.classList.remove('active'));
        btns[0].classList.add('active'); // reset to AND
    }});

    // Show all rows
    document.querySelectorAll('tr.row-hidden').forEach(tr => tr.classList.remove('row-hidden'));
    document.getElementById('filterInfo').textContent = '';
    updateBadge();
}}

// ============================================================
// Selection helpers
// ============================================================
function getRowCheckboxes() {{
    return document.querySelectorAll('input.row-cb');
}}

function getVisibleRowCheckboxes() {{
    return document.querySelectorAll('tr:not(.row-hidden) input.row-cb');
}}

function updateBadge() {{
    const n = document.querySelectorAll('input.row-cb:checked').length;
    const badge = document.getElementById('selCount');
    badge.textContent = n;
    badge.className = n ? 'badge' : 'badge zero';
    document.getElementById('btnExport').disabled = (n === 0);
}}

function toggleSelectAll() {{
    allSelected = !allSelected;
    // Only operate on visible rows
    getVisibleRowCheckboxes().forEach(cb => {{ cb.checked = allSelected; highlightRow(cb); }});
    document.querySelectorAll('input.section-cb').forEach(cb => cb.checked = allSelected);
    document.getElementById('btnSelectAll').textContent = allSelected ? '☐ Deselect All' : '☑ Select All';
    updateBadge();
}}

// Per-section select all
document.addEventListener('change', function(e) {{
    if (e.target.classList.contains('section-cb')) {{
        const tbody = e.target.closest('table').querySelector('tbody');
        tbody.querySelectorAll('tr:not(.row-hidden) input.row-cb').forEach(cb => {{
            cb.checked = e.target.checked;
            highlightRow(cb);
        }});
        updateBadge();
    }}
    if (e.target.classList.contains('row-cb')) {{
        highlightRow(e.target);
        const tbody = e.target.closest('tbody');
        const visibleInSection = tbody.querySelectorAll('tr:not(.row-hidden) input.row-cb');
        const checkedInSection = tbody.querySelectorAll('tr:not(.row-hidden) input.row-cb:checked');
        const sectionCb = e.target.closest('table').querySelector('input.section-cb');
        if (sectionCb) sectionCb.checked = (visibleInSection.length === checkedInSection.length);
        updateBadge();
    }}
}});

function highlightRow(cb) {{
    const tr = cb.closest('tr');
    if (cb.checked) tr.classList.add('row-selected');
    else tr.classList.remove('row-selected');
}}

// ============================================================
// TMX generation
// ============================================================
function escapeXml(s) {{
    return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&apos;');
}}

function fmtDate(isoStr) {{
    if (!isoStr) return '20260101T000000Z';
    return isoStr.replace(/[-:]/g, '').replace(/\\.\\d+/, '');
}}

function langUnderscore(lang) {{
    return lang.replace('-', '_');
}}

function buildTmx(srcLang, tgtLang, entries) {{
    let xml = '<?xml version="1.0" encoding="UTF-8"?>\\n';
    xml += '<!DOCTYPE tmx SYSTEM "tmx14.dtd">\\n';
    xml += '<tmx version="1.4">';
    xml += '<header adminlang="en-US" creationtool="Tranzor" creationtoolversion="1.0"'
         + ' datatype="xml" o-tmf="Tranzor" srclang="' + srcLang + '" segtype="sentence">';
    xml += '</header>';
    xml += '<body>\\n';

    entries.forEach(function(e, i) {{
        const d = fmtDate(e.edit_time);
        const editor = escapeXml(e.editor);
        const tuid = 'tranzor-' + i;
        xml += '<tu creationdate="' + d + '" creationid="' + editor
             + '" changedate="' + d + '" changeid="' + editor
             + '" tuid="' + tuid + '">';
        xml += '<prop type="x-segment-id">' + escapeXml(e.string_key) + '</prop>';
        xml += '<prop type="x-Project">' + escapeXml(e.task_name) + '</prop>';
        xml += '\\n<tuv xml:lang="' + srcLang + '"><seg>' + escapeXml(e.source_text) + '</seg></tuv>\\n';
        xml += '<tuv xml:lang="' + tgtLang + '"><seg>' + escapeXml(e.after) + '</seg></tuv>\\n';
        xml += '</tu>\\n';
    }});

    xml += '</body></tmx>';
    return xml;
}}

// ============================================================
// Download helpers
// ============================================================
function downloadBlob(blob, name) {{
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = name;
    document.body.appendChild(a);
    a.click();
    setTimeout(() => {{ URL.revokeObjectURL(a.href); a.remove(); }}, 100);
}}

async function exportTMX() {{
    const checked = document.querySelectorAll('tr:not(.row-hidden) input.row-cb:checked');
    if (!checked.length) return;

    const status = document.getElementById('exportStatus');
    status.textContent = 'Generating…';

    const byLang = {{}};
    checked.forEach(cb => {{
        const idx = parseInt(cb.dataset.idx);
        const row = ROWS[idx];
        const lang = row.language;
        if (!byLang[lang]) byLang[lang] = [];
        byLang[lang].push(row);
    }});

    const langs = Object.keys(byLang);

    if (langs.length === 1) {{
        const lang = langs[0];
        const tmx = buildTmx('en-US', lang, byLang[lang]);
        const blob = new Blob([tmx], {{ type: 'application/xml' }});
        const fn = 'tranzor_tm_en_US-' + langUnderscore(lang) + '.tmx';
        downloadBlob(blob, fn);
        status.textContent = '✓ Downloaded ' + fn + ' (' + byLang[lang].length + ' entries)';
    }} else {{
        if (typeof JSZip === 'undefined') {{
            status.textContent = '⚠ JSZip not loaded. Check your internet connection and reload.';
            return;
        }}
        const zip = new JSZip();
        let totalEntries = 0;
        langs.forEach(lang => {{
            const tmx = buildTmx('en-US', lang, byLang[lang]);
            const fn = 'en_US-' + langUnderscore(lang) + '.tmx';
            zip.file(fn, tmx);
            totalEntries += byLang[lang].length;
        }});
        const content = await zip.generateAsync({{ type: 'blob' }});
        downloadBlob(content, 'tranzor_tm_export.zip');
        status.textContent = '✓ Downloaded ZIP with ' + langs.length + ' languages, ' + totalEntries + ' entries';
    }}
}}
</script>
</body>
</html>"""

    with open(filename, "w", encoding="utf-8") as f:
        f.write(page)


# ---------------------------------------------------------------------------
# 8) 写入 Excel
# ---------------------------------------------------------------------------
EXCEL_COLUMNS = [
    "Edit ID", "Edit Time", "Editor",
    "Task Name", "Language", "String Key",
    "Source (en-US)", "Before Edit", "After Edit", "Diff", "Notes",
]


def write_excel(rows, filename):
    """将结果写入 Excel 文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("错误: 缺少 openpyxl，请运行: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Translation Changes"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    before_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    after_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    diff_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for row_idx, r in enumerate(rows, 2):
        diff = word_diff_text(r["before"], r["after"])
        values = [
            r["edit_id"], r["edit_time"], r["editor"],
            r["task_name"], r["language"], r["string_key"],
            r["source_text"], r["before"], r["after"], diff, r["notes"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 8:
                cell.fill = before_fill
            elif col_idx == 9:
                cell.fill = after_fill
            elif col_idx == 10:
                cell.fill = diff_fill

    col_widths = [8, 20, 15, 25, 10, 35, 40, 40, 40, 50, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(filename)


# ---------------------------------------------------------------------------
# 9) 保存文件（含文件被占用时自动重命名）
# ---------------------------------------------------------------------------
def save_file(rows, filename, label, fmt):
    """保存文件，文件被占用时自动加序号"""
    base, ext = os.path.splitext(filename)
    save_path = filename
    for attempt in range(100):
        try:
            if fmt == "html":
                write_html(rows, save_path, label)
            else:
                write_excel(rows, save_path)
            print(f"已导出: {save_path}")
            # HTML 自动在浏览器中打开
            if fmt == "html":
                from export_gui import open_in_browser
                open_in_browser(save_path)
            return
        except PermissionError:
            attempt_num = attempt + 1
            save_path = f"{base}_{attempt_num}{ext}"
            print(f"  文件被占用，尝试保存为: {save_path}")


# ---------------------------------------------------------------------------
# 10) 命令行入口
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="导出 Tranzor 翻译手动变更记录",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python export_changes.py                 导出所有变更 (HTML，自动打开浏览器)
  python export_changes.py --xlsx          导出所有变更 (Excel)
  python export_changes.py --task 5678     只导出某个 task 的变更
        """,
    )
    parser.add_argument("--task", type=int, help="只导出指定 task ID 的变更")
    parser.add_argument("--xlsx", action="store_true", help="输出 Excel 格式（默认 HTML）")
    parser.add_argument("--output", "-o", help="输出文件名 (默认自动生成)")

    args = parser.parse_args()

    if args.task:
        print(f"指定 Task: {args.task}")

    rows = collect_changes(task_id=args.task)
    print(f"\n共找到 {len(rows)} 条变更记录")

    if not rows:
        print("没有变更记录，无需导出。")
        return

    # 确定输出格式和文件名
    fmt = "xlsx" if args.xlsx else "html"
    ext = ".xlsx" if args.xlsx else ".html"
    today_str = date.today().isoformat()
    label = f"All changes (exported {today_str})"

    if args.output:
        filename = args.output
    else:
        filename = f"tranzor_all_changes_{today_str}{ext}"

    save_file(rows, filename, label, fmt)


if __name__ == "__main__":
    main()
