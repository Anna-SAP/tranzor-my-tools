"""
JSON 导出器 — 统一格式，供翻译质量审计（QA）下游消费
====================================================
把"每行一条 (key, language) 翻译"的平铺数据，透视成
"每个 key 一条记录，包含所有目标语言"的字典结构，
schema 与 `/rc-core-products-trans-checker` 等 LQA Skill
期望的输入完全一致：

    [
      {
        "key": "RingCentral.bui.<hash>.STRING_KEY",
        "en-US": "Cancellation requested",
        "de-DE": "Kündigung beantragt",
        ...
      },
      ...
    ]

三个数据源（File Translation / MR Pipeline / Scan Tasks）的行 schema
略有差异，本模块自动归一化：

  - MR Pipeline / Scan Tasks (results dict):
        { "translations": [
            {"opus_id": ..., "target_language": ...,
             "source_text": ..., "translated_text": ...},
            ...
          ] }
  - File Translation (flat rows):
        [{"string_key": ..., "language": ...,
          "source_text": ..., "translated_text": ...,
          # 或 changes 模式下的 before/after
          "before": ..., "after": ...}, ...]

en-US 列的取值规则：
  1. 若存在显式的 en-US 行（target_language == "en-US"），优先用其 translated_text；
  2. 否则回退为该 key 任意一行的 source_text（en-US 通常就是源文）。
"""
from __future__ import annotations

import json
import os
import re
from collections import OrderedDict


# 列顺序：key 第一，en-US 第二（源语言），其余按字母序
_SOURCE_LANG = "en-US"


def _segment_stable_key(row, key):
    """Keep UNS segmented units from collapsing onto one JSON key.

    Scan (and some MR) ``/results`` rows for Handlebars emails reuse the
    *file-level* ``opus_id`` (e.g. ``common.uns.airLeadCaptured__email_html__1210``)
    across every ``tu_id`` segment, with ``has_seg_units=true``. The QA
    pivot is last-write-wins on ``(key, lang)``, so 13k segment rows
    became 191 file keys whose translations were whichever segment
    happened to arrive last — unusable for LQA.

    When the row is a segment, stamp ``:::seg:::{tu_id}`` (the same
    shape the UNS segmented-email LQA skill already expects). Rows that
    already carry ``:::seg:::`` or have no ``tu_id`` are left alone.
    """
    if not key or ":::seg:::" in key:
        return key
    if not row.get("has_seg_units"):
        return key
    tu = row.get("tu_id")
    if tu in (None, ""):
        return key
    return f"{key}:::seg:::{tu}"


def _normalize_row(row):
    """归一化一条翻译行 → (key, lang, source_text, translated_text)。

    兼容两套字段命名：
      - MR/Scan 用 opus_id / target_language
      - File Translation 用 string_key / language
      - File Translation 的 "changes" 模式额外使用 before/after，
        我们取 after 作为最终译文。
    """
    key = row.get("opus_id") or row.get("string_key") or ""
    key = _segment_stable_key(row, key)
    lang = row.get("target_language") or row.get("language") or ""
    src = row.get("source_text") or ""
    # 优先取 translated_text；其次 after（changes 模式的最终值）
    tgt = row.get("translated_text")
    if tgt is None or tgt == "":
        tgt = row.get("after") or ""
    return key, lang, src, tgt


def build_json_entries(rows, all_languages=None, fill_missing=False):
    """把平铺行透视成 [{key, en-US, de-DE, ...}, ...]。

    - 缺少 key 或 lang 的行被跳过；
    - 同 (key, lang) 多行时，最后一条覆盖先前（通常意味着是更新后的最终译文）；
    - en-US 缺失时回退为 source_text；
    - 输出按 key 升序排序，列顺序固定为 key → en-US → 其余语言（字母序），
      与现有 LQA 工具期望的格式一致。

    全量语言补齐（仅"全量翻译/All Translations"导出启用）
    ----------------------------------------------------
    源数据天然是稀疏的：某个 (key, language) 没有译文时，后端根本不会返回
    对应行（``fetch_all_translations`` 还会主动丢弃 ``translated_text`` 为空
    的行）。直接透视会得到"参差不齐"的结果——同一个 task 里，有的 key 含 16
    种语言、有的只含 13 种。对 QA 全量审计而言，这违反"必须 100% 覆盖所有
    配置的目标语言"的硬要求。

    当 ``fill_missing=True``（或显式传入 ``all_languages``）时，对每个 key 补齐
    其应有的全部目标语言，缺失值填 ``""``（en-US 缺失时优先用 source_text 兜底）。
    空串如实表达"该语言尚无译文"，下游 LQA 会照常将其标记为未翻译，而不是被
    悄悄隐藏。

    语言集合按 **来源 task 维度（观察到的语言）** 计算，绝不跨产品误填——不同
    产品的 key 互不相同，各自只补齐其所属 task 内观察到的语言：
      - ``fill_missing=True``：每个 key 补齐到"其所属 task 内观察到的语言并集"
        （行携带 ``task_id``；无 task 信息时退化为全局观察并集）。这修复了
        "同一 task 里 A key 有 zh-CN、B key 没有"的参差不齐。
      - ``all_languages``：可选的全局权威目标语言列表。整个导出同属一个配置
        集合（例如单个 task 导出）时传入，可额外保证连"零译文"的配置语言也出现。
        多 task 混合导出不传它，避免把别的 task 配置的语言误填进来。

    默认 ``fill_missing=False`` 时行为与历史完全一致（稀疏透视），因此
    Changes/变更导出等共用本函数的调用方不受影响。
    """
    # key -> { lang: translated, "__source__": source_text, "__tasks__": set }
    by_key = OrderedDict()
    # task_id -> set(观察到的语言)
    langs_by_task = {}
    all_observed = set()

    for row in rows or []:
        if not isinstance(row, dict):
            # 非 dict 行直接跳过（与"缺 key/lang 即跳过"的容错语义一致），避免
            # _normalize_row 里 row.get(...) 抛 AttributeError 中断整个导出
            continue
        key, lang, src, tgt = _normalize_row(row)
        if not key or not lang:
            continue
        bucket = by_key.setdefault(key, {})
        # 记住一次非空源文（不同行的 source 应该一致）
        if src and not bucket.get("__source__"):
            bucket["__source__"] = src
        bucket[lang] = tgt
        all_observed.add(lang)
        # 记录该 key 来自哪个 task，并累计每个 task 观察到的语言集合
        tid = row.get("task_id")
        if tid not in (None, ""):
            tid = str(tid)
            bucket.setdefault("__tasks__", set()).add(tid)
            langs_by_task.setdefault(tid, set()).add(lang)

    densify = bool(fill_missing or all_languages)
    global_extra = {str(l) for l in (all_languages or []) if l}

    entries = []
    for key in sorted(by_key.keys()):
        bucket = by_key[key]
        source = bucket.pop("__source__", "")
        tasks = bucket.pop("__tasks__", set())

        # en-US 缺失时用 source_text 兜底
        if _SOURCE_LANG not in bucket and source:
            bucket[_SOURCE_LANG] = source

        if densify:
            # 计算该 key 应覆盖的目标语言集合：按来源 task 维度取"观察到的语言"。
            # 不同产品的 key（opus_id）互不相同，因此天然不会跨产品误填；同一个
            # key 跨 task 命中时取其各 task 观察语言的并集（同产品的真实合并视图）。
            target = set()
            for tid in tasks:
                target |= langs_by_task.get(tid, set())
            if not tasks:
                # 行里没有 task_id 时，退化为全局观察并集（单任务导出即该任务全集）
                target |= all_observed
            target |= global_extra
            for lang in target:
                if lang not in bucket:
                    # en-US 优先用 source 兜底，其余语言填空串
                    bucket[lang] = source if lang == _SOURCE_LANG and source else ""

        ordered = OrderedDict()
        ordered["key"] = key
        if _SOURCE_LANG in bucket:
            ordered[_SOURCE_LANG] = bucket[_SOURCE_LANG]
        for lang in sorted(l for l in bucket if l != _SOURCE_LANG):
            ordered[lang] = bucket[lang]
        entries.append(ordered)

    return entries


def write_translations_json(payload, filename, all_languages=None,
                            fill_missing=False):
    """把翻译数据写成 LQA 工具期望的 JSON 文件。

    payload 可以是：
      - dict（含 ``translations`` 列表）   → MR Pipeline / Scan Tasks
      - list（平铺行）                     → File Translation

    ``all_languages`` / ``fill_missing`` 透传给 :func:`build_json_entries`，
    用于"全量翻译导出"的全量语言补齐（见该函数文档）。默认全部关闭时行为
    与历史一致。

    返回写入的条目列表（便于调用方记日志或测试）。
    """
    if isinstance(payload, dict):
        rows = payload.get("translations") or []
    else:
        rows = payload or []

    entries = build_json_entries(
        rows,
        all_languages=all_languages,
        fill_missing=fill_missing,
    )

    # 始终用 UTF-8 + ensure_ascii=False 让非英文字符直接可读；
    # indent=2 与 BUI 样例对齐，便于人工审查。
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

    lang_count = len({l for e in entries for l in e if l != 'key'})
    print(
        f"已导出: {filename}  "
        f"({len(entries)} 个 key，覆盖 {lang_count} 种语言)"
    )
    # 补齐了空语言节点时给出透明提示（仅全量导出路径会触发）
    if fill_missing or all_languages:
        filled = sum(
            1 for e in entries for k, v in e.items()
            if k != "key" and v == ""
        )
        if filled:
            print(f"  [+] 已补齐 {filled} 个空语言节点，保证每个 key 100% 覆盖目标语言")
    return entries


def save_json_file(payload, filename, all_languages=None, fill_missing=False):
    """带"文件被占用自动加序号"的安全写入入口，对齐
    ``export_mr_pipeline.save_mr_file`` / ``export_translations.save_file``
    的行为，保证 GUI 反复导出时不会因为上一份还开着而失败。
    """
    base, ext = os.path.splitext(filename)
    save_path = filename
    for attempt in range(100):
        try:
            write_translations_json(
                payload, save_path,
                all_languages=all_languages,
                fill_missing=fill_missing,
            )
            return save_path
        except PermissionError:
            attempt_num = attempt + 1
            save_path = f"{base}_{attempt_num}{ext}"
            print(f"  文件被占用，尝试保存为: {save_path}")
    return None


# ---------------------------------------------------------------------------
# Source-only XLSX (Key / en-US Value / task name)
# ---------------------------------------------------------------------------
# One row per unique string key (UNS email segments stay distinct via
# ``:::seg:::{tu_id}`` — see ``_segment_stable_key``). This is the companion
# to the QA JSON export: same keys and en-US values, no target locales.

SOURCE_XLSX_HEADERS = ("Key", "en-US Value", "task name")

_EXCEL_SHEET_INVALID = re.compile(r'[:\\/?*\[\]]')
_EXCEL_SHEET_MAX = 31
_EMPTY_SHEET_MARKERS = frozenset({
    "", "—", "-", "–", "−", "…", "...", "n/a", "N/A", "None", "null",
})


def source_rows_from_payload(payload, task_name=""):
    """Unique ``(Key, en-US Value, task name)`` triples from a JSON payload.

    Reuses :func:`build_json_entries` so UNS segmented units keep the
    ``:::seg:::{tu_id}`` key shape and en-US falls back to ``source_text``.
    ``task_name`` is repeated on every row (typically the companion JSON
    filename, e.g. ``mr_pipeline_MR4100_bd6bba88_all_2026-08-31_08-03-01.json``).
    """
    if isinstance(payload, dict):
        rows = payload.get("translations") or []
    else:
        rows = payload or []
    entries = build_json_entries(rows, fill_missing=False)
    label = str(task_name or "")
    return [
        (entry.get("key") or "", entry.get(_SOURCE_LANG) or "", label)
        for entry in entries
    ]


def sheet_title_for_mr(jira_id, mr_iid):
    """Excel tab title matching the Delta Analysis workbook, e.g. ``BUG-352 MR!4103``.

    Empty / placeholder JIRA (``—``, ``…``) drops the ticket so the sheet is
    just ``MR!{iid}``. Both missing → ``Source``.
    """
    jira = str(jira_id or "").strip()
    if jira in _EMPTY_SHEET_MARKERS:
        jira = ""
    mr = str(mr_iid or "").strip()
    if mr.upper().startswith("MR"):
        mr = mr[2:].lstrip("#!")
    else:
        mr = mr.lstrip("#!")
    if mr in _EMPTY_SHEET_MARKERS:
        mr = ""
    if jira and mr:
        return f"{jira} MR!{mr}"
    if mr:
        return f"MR!{mr}"
    if jira:
        return jira
    return "Source"


def sanitize_excel_sheet_title(name, used=None):
    """Make ``name`` a legal, unique Excel sheet title (max 31 chars).

    Excel rejects ``: \\ / ? * [ ]``, leading/trailing ``'``, and treats
    names as case-insensitive. ``!`` is legal and required for the
    ``BUG-352 MR!4103`` convention. ``used`` is updated in place.
    """
    if used is None:
        used = set()
    text = str(name or "").strip()
    text = _EXCEL_SHEET_INVALID.sub("-", text)
    text = text.strip("'") or "Source"
    text = text[:_EXCEL_SHEET_MAX]
    used_lower = {item.lower() for item in used}
    candidate = text
    n = 1
    while candidate.lower() in used_lower:
        suffix = f" ({n})"
        candidate = text[:_EXCEL_SHEET_MAX - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def write_source_xlsx(sheets, filename):
    """Write a source-only workbook.

    ``sheets`` is an iterable of ``{"title": str, "rows": [(key, en_us, task_name), ...]}``.
    Each sheet gets a header row ``Key | en-US Value | task name``.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("错误: 缺少 openpyxl 包，请先运行: pip install openpyxl")
        return None

    wb = Workbook()
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    used_titles = set()
    first = True

    for sheet in sheets or []:
        title = sanitize_excel_sheet_title(
            (sheet or {}).get("title") or "Source", used_titles)
        rows = (sheet or {}).get("rows") or []
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title)

        for col_i, header in enumerate(SOURCE_XLSX_HEADERS, 1):
            cell = ws.cell(row=1, column=col_i, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_i, triple in enumerate(rows, 2):
            if isinstance(triple, dict):
                key = triple.get("key") or triple.get("Key") or ""
                en_us = (
                    triple.get("en-US Value")
                    or triple.get("en-US")
                    or ""
                )
                task = (
                    triple.get("task name")
                    or triple.get("task_name")
                    or ""
                )
            elif isinstance(triple, (tuple, list)) and len(triple) >= 3:
                key, en_us, task = triple[0], triple[1], triple[2]
            else:
                continue
            ws.cell(row=row_i, column=1, value=key)
            value_cell = ws.cell(row=row_i, column=2, value=en_us)
            value_cell.alignment = wrap
            ws.cell(row=row_i, column=3, value=task)

        ws.column_dimensions[get_column_letter(1)].width = 55
        ws.column_dimensions[get_column_letter(2)].width = 80
        ws.column_dimensions[get_column_letter(3)].width = 55
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:C{max(1, 1 + len(rows))}"

    if first:
        # No sheets were supplied — keep a blank default tab with headers.
        ws = wb.active
        ws.title = sanitize_excel_sheet_title("Source", used_titles)
        for col_i, header in enumerate(SOURCE_XLSX_HEADERS, 1):
            cell = ws.cell(row=1, column=col_i, value=header)
            cell.fill = header_fill
            cell.font = header_font

    wb.save(filename)
    print(f"已导出: {filename}")
    return filename


def save_source_xlsx(sheets, filename):
    """Write a source XLSX, retrying with a numeric suffix if the file is locked."""
    base, ext = os.path.splitext(filename)
    save_path = filename
    for attempt in range(100):
        try:
            written = write_source_xlsx(sheets, save_path)
            return written
        except PermissionError:
            attempt_num = attempt + 1
            save_path = f"{base}_{attempt_num}{ext}"
            print(f"  文件被占用，尝试保存为: {save_path}")
    return None

