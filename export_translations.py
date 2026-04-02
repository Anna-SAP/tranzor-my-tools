"""
Tranzor 全部翻译导出工具
========================
通过 Tranzor HTTP API 导出特定翻译任务的全部翻译结果。
默认输出 HTML（含可视化查看 + TMX 导出），加 --xlsx 输出 Excel。

用法:
    python export_translations.py --task 123          # 导出某个 task 的全部翻译 (HTML)
    python export_translations.py --task 123 --xlsx   # 输出 Excel 格式

首次使用前:
    pip install openpyxl requests
"""

import argparse
import html
import os
import sys
import webbrowser
from datetime import date

try:
    import requests
except ImportError:
    print("错误: 缺少 requests 包，请先运行:")
    print("  pip install openpyxl requests")
    sys.exit(1)

# ---------------------------------------------------------------------------
# 配置 — 你的 Tranzor 平台地址（就是你浏览器里打开的那个地址）
# ---------------------------------------------------------------------------
TRANZOR_URL = "http://tranzor-platform.int.rclabenv.com"

API = f"{TRANZOR_URL}/api/v1/legacy"

# 共享 HTTP session（连接池复用）
_session = requests.Session()

# 最大重试次数
MAX_RETRIES = 3

# 并发工作线程数（task 级别）
MAX_WORKERS = 8

# 单个 task 内的并发 API 请求数
MAX_PAGE_WORKERS = 6


def _api_get(url, **kwargs):
    """带重试的 GET 请求，遇到超时/连接错误自动重试"""
    import time
    kwargs.setdefault("timeout", 30)
    for attempt in range(MAX_RETRIES):
        try:
            return _session.get(url, **kwargs)
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError) as e:
            if attempt < MAX_RETRIES - 1:
                wait = 2 ** attempt  # 1s, 2s, 4s
                print(f"    ⚠ 请求超时，{wait}s 后重试 ({attempt+1}/{MAX_RETRIES})...")
                time.sleep(wait)
            else:
                raise


# ---------------------------------------------------------------------------
# 1) 获取 task 信息
# ---------------------------------------------------------------------------
def fetch_task_info(task_id):
    """获取单个 task 的名称"""
    try:
        resp = _api_get(f"{API}/tasks/{task_id}")
        resp.raise_for_status()
        data = resp.json()
        return data.get("task_name", f"Task {task_id}")
    except Exception:
        return f"Task {task_id}"


# ---------------------------------------------------------------------------
# 2) 获取 task 列表
# ---------------------------------------------------------------------------
def fetch_tasks():
    """获取所有已完成的 task（分页遍历）"""
    all_tasks = []
    offset = 0
    limit = 200

    while True:
        params = {"limit": limit, "offset": offset, "status": "Completed"}
        resp = _api_get(f"{API}/tasks", params=params)
        resp.raise_for_status()
        data = resp.json()

        tasks = data.get("tasks", [])
        all_tasks.extend(tasks)

        if len(tasks) < limit:
            break
        offset += limit

    return all_tasks


# ---------------------------------------------------------------------------
# 3) 获取某个 task 下的全部翻译（不筛选 translation_type）
#    使用并发分页加速大数据量 task
# ---------------------------------------------------------------------------
def fetch_all_translations(task_id):
    """获取某个 task 中所有翻译条目（全量，不筛选类型）。
    对于大数据 task，使用并发分页获取以大幅缩短耗时。
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    limit = 200

    # 第 1 步：探测总条目数（只取 1 条以获取 total）
    probe_resp = _api_get(
        f"{API}/tasks/{task_id}/translations",
        params={"limit": 1, "offset": 0}
    )
    probe_resp.raise_for_status()
    probe_data = probe_resp.json()
    total = probe_data.get("total", 0)

    if total == 0:
        return []

    # 第 2 步：计算所有分页偏移量
    offsets = list(range(0, total, limit))

    # 第 3 步：并发获取所有分页
    def _fetch_page(offset):
        resp = _api_get(
            f"{API}/tasks/{task_id}/translations",
            params={"limit": limit, "offset": offset}
        )
        resp.raise_for_status()
        return offset, resp.json().get("entries", [])

    page_results = {}
    workers = min(MAX_PAGE_WORKERS, len(offsets))

    if workers <= 1:
        # 只有 1 页，直接获取
        _, entries = _fetch_page(0)
        page_results[0] = entries
    else:
        with ThreadPoolExecutor(max_workers=workers) as page_pool:
            futures = {page_pool.submit(_fetch_page, o): o for o in offsets}
            for f in as_completed(futures):
                offset_val, entries = f.result()
                page_results[offset_val] = entries

    # 第 4 步：按 offset 顺序合并结果
    all_translations = []
    for o in offsets:
        for entry in page_results.get(o, []):
            if entry.get("translated_text", ""):
                all_translations.append(entry)

    return all_translations


# ---------------------------------------------------------------------------
# 4) 主流程：收集翻译数据
# ---------------------------------------------------------------------------
def collect_translations(task_id=None):
    """收集全部翻译记录"""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if task_id:
        task_name = fetch_task_info(task_id)
        print(f"  正在获取 Task '{task_name}' (ID: {task_id}) 的全部翻译...")

        translations = fetch_all_translations(task_id)
        print(f"  找到 {len(translations)} 条翻译记录")

        rows = []
        for entry in translations:
            rows.append({
                "task_id": task_id,
                "task_name": task_name,
                "language": entry.get("target_language", ""),
                "string_key": entry.get("opus_id", ""),
                "source_text": entry.get("source_text", ""),
                "translated_text": entry.get("translated_text", ""),
                "translation_type": entry.get("translation_type", ""),
            })
        return rows
    else:
        # 遍历所有已完成的 task
        print("  正在获取 task 列表...")
        tasks = fetch_tasks()
        print(f"  找到 {len(tasks)} 个已完成的 task")

        all_rows = []
        total = len(tasks)

        def _process_task(task_info):
            tid, tname, idx, total_count = task_info
            try:
                translations = fetch_all_translations(tid)
            except Exception as e:
                print(f"  ⚠ [{idx}/{total_count}] Task '{tname}' - 获取翻译失败: {e}")
                return []

            if not translations:
                return []

            print(f"  [{idx}/{total_count}] Task '{tname}' - 找到 {len(translations)} 条翻译")

            rows = []
            for entry in translations:
                rows.append({
                    "task_id": tid,
                    "task_name": tname,
                    "language": entry.get("target_language", ""),
                    "string_key": entry.get("opus_id", ""),
                    "source_text": entry.get("source_text", ""),
                    "translated_text": entry.get("translated_text", ""),
                    "translation_type": entry.get("translation_type", ""),
                })
            return rows

        task_infos = [
            (task["id"], task.get("task_name", ""), i + 1, total)
            for i, task in enumerate(tasks)
        ]

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
            futures = {pool.submit(_process_task, info): info for info in task_infos}
            for f in as_completed(futures):
                all_rows.extend(f.result())

        return all_rows


# ---------------------------------------------------------------------------
# 5) 写入 HTML（含 TMX 导出功能）
# ---------------------------------------------------------------------------
def write_html(rows, filename, label):
    """生成带 TMX 导出功能的 HTML 报告，按语言分组"""
    import json
    from collections import OrderedDict

    # 按语言分组
    groups = OrderedDict()
    for r in rows:
        lang = r.get("language") or "(unknown)"
        groups.setdefault(lang, []).append(r)

    # 语言分组颜色
    lang_colors = [
        ("#4472C4", "#e8eef7"),  # 蓝
        ("#E67E22", "#fdf2e6"),  # 橙
        ("#27AE60", "#e8f5ee"),  # 绿
        ("#8E44AD", "#f3eaf8"),  # 紫
        ("#E74C3C", "#fceaea"),  # 红
        ("#16A085", "#e6f4f1"),  # 青
        ("#D4AC0D", "#faf6e4"),  # 金
        ("#2C3E50", "#eaecee"),  # 灰
    ]

    # --- 构建 JSON 数据供 JS 使用 ---
    js_rows = []
    all_langs = set()
    all_types = set()
    for lang_rows_list in groups.values():
        for r in lang_rows_list:
            js_rows.append({
                "source_text": r["source_text"],
                "translated_text": r["translated_text"],
                "language": r["language"],
                "string_key": r["string_key"],
                "translation_type": r["translation_type"],
                "task_id": r.get("task_id", ""),
                "task_name": r["task_name"],
            })
            all_langs.add(r["language"])
            all_types.add(r["translation_type"])
    rows_json = json.dumps(js_rows, ensure_ascii=False)
    langs_json = json.dumps(sorted(all_langs), ensure_ascii=False)
    types_json = json.dumps(sorted(all_types), ensure_ascii=False)

    sections_parts = []
    global_idx = 0
    for lang_i, (lang_name, lang_rows) in enumerate(groups.items()):
        color_pair = lang_colors[lang_i % len(lang_colors)]
        header_bg = color_pair[0]
        section_bg = color_pair[1]

        row_parts = []
        for r in lang_rows:
            row_parts.append(
                f'<tr>'
                f'<td class="cb-cell"><input type="checkbox" class="row-cb" data-idx="{global_idx}"></td>'
                f'<td class="num">{global_idx + 1}</td>'
                f'<td class="task-id">{r.get("task_id", "")}</td>'
                f'<td class="task">{html.escape(r["task_name"])}</td>'
                f'<td class="lang">{html.escape(r["language"])}</td>'
                f'<td class="key">{html.escape(r["string_key"])}</td>'
                f'<td class="source">{html.escape(r["source_text"])}</td>'
                f'<td class="translated">{html.escape(r["translated_text"])}</td>'
                f'<td class="type">{html.escape(r["translation_type"])}</td>'
                f'</tr>'
            )
            global_idx += 1
        table_rows = '\n'.join(row_parts)

        sections_parts.append(
            f'<div class="lang-section" style="background: {section_bg}; border-left: 4px solid {header_bg}; border-radius: 8px; padding: 16px; margin-bottom: 24px;">'
            f'<h2 style="color: {header_bg}; margin-bottom: 12px; font-size: 17px;">'
            f'🌐 {html.escape(lang_name)}'
            f'<span class="count" style="background: {header_bg};">{len(lang_rows)} entries</span>'
            f'</h2>'
            f'<table><thead><tr style="background: {header_bg};">'
            f'<th class="cb-cell"><input type="checkbox" class="section-cb"></th>'
            f'<th>#</th><th>Task ID</th><th>Task</th><th>Lang</th>'
            f'<th>String Key</th><th>Source (en-US)</th><th>Translated Text</th>'
            f'<th>Type</th>'
            f'</tr></thead><tbody>{table_rows}</tbody></table></div>'
        )
    sections_html = '\n'.join(sections_parts)

    # 顶部语言索引
    toc_items = ""
    for lang_i, (lang_name, lang_rows) in enumerate(groups.items()):
        color_pair = lang_colors[lang_i % len(lang_colors)]
        toc_items += f'<span style="display:inline-block; background:{color_pair[0]}; color:#fff; border-radius:16px; padding:4px 14px; margin:4px; font-size:13px;">{html.escape(lang_name)} ({len(lang_rows)})</span>'

    page = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Tranzor Translations - {html.escape(label)}</title>
<style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: -apple-system, "Segoe UI", Roboto, Arial, sans-serif; background: #f5f6fa; padding: 24px; padding-top: 72px; color: #333; }}
    h1 {{ font-size: 20px; margin-bottom: 4px; }}
    .meta {{ color: #666; font-size: 14px; margin-bottom: 8px; }}
    .toc {{ margin-bottom: 20px; }}
    .count {{ display: inline-block; color: #fff; border-radius: 12px; padding: 2px 10px; font-size: 13px; margin-left: 8px; }}
    table {{ border-collapse: collapse; width: 100%; background: #fff; border-radius: 6px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,.08); font-size: 13px; }}
    th {{ color: #fff; padding: 10px 8px; text-align: left; white-space: nowrap; }}
    td {{ padding: 8px; border-bottom: 1px solid #e8e8e8; vertical-align: top; line-height: 1.5; }}
    tr:hover {{ background: rgba(255,255,255,.6); }}
    .num {{ text-align: center; color: #999; min-width: 30px; }}
    .task-id {{ text-align: center; font-family: monospace; font-size: 12px; color: #888; white-space: nowrap; }}
    .task {{ max-width: 200px; }}
    .lang {{ text-align: center; }}
    .key {{ font-family: monospace; font-size: 12px; max-width: 260px; word-break: break-all; color: #555; }}
    .source {{ background: #fff8e6; min-width: 200px; }}
    .translated {{ background: #edf7ed; min-width: 200px; }}
    .type {{ font-size: 12px; color: #666; white-space: nowrap; }}

    /* --- Toolbar --- */
    .tmx-toolbar {{
        position: fixed; top: 0; left: 0; right: 0; z-index: 1000;
        background: #fff; border-bottom: 1px solid #dde; padding: 10px 24px;
        display: flex; align-items: center; gap: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,.08);
    }}
    .tmx-toolbar .btn {{
        border: none; border-radius: 6px; padding: 7px 16px; cursor: pointer;
        font-size: 13px; font-weight: 600; transition: background .15s, transform .1s;
    }}
    .tmx-toolbar .btn:active {{ transform: scale(.97); }}
    .btn-select {{ background: #eef1f5; color: #333; }}
    .btn-select:hover {{ background: #dde3ea; }}
    .btn-filter {{ background: #eef1f5; color: #333; }}
    .btn-filter:hover {{ background: #dde3ea; }}
    .btn-filter.active {{ background: #4472C4; color: #fff; }}
    .btn-export {{ background: #4472C4; color: #fff; }}
    .btn-export:hover {{ background: #3461b0; }}
    .btn-export:disabled {{ background: #a8b8d0; cursor: not-allowed; opacity: .7; }}
    .badge {{ background: #4472C4; color: #fff; border-radius: 12px; padding: 2px 10px; font-size: 12px; font-weight: 700; min-width: 28px; text-align: center; }}
    .badge.zero {{ background: #bbb; }}
    .toolbar-sep {{ width: 1px; height: 24px; background: #dde; }}

    /* checkbox column */
    .cb-cell {{ text-align: center; width: 32px; min-width: 32px; }}
    .cb-cell input[type="checkbox"] {{ width: 15px; height: 15px; cursor: pointer; accent-color: #4472C4; }}
    tr.row-selected {{ background: #e8eef7 !important; }}
    tr.row-hidden {{ display: none !important; }}

    /* --- Filter Panel --- */
    .filter-panel {{
        background: #1e293b; color: #cbd5e1; border-radius: 10px;
        margin-bottom: 20px; overflow: hidden;
        max-height: 0; opacity: 0; transition: max-height .35s ease, opacity .25s ease, margin .25s ease, padding .25s ease;
        padding: 0 20px;
    }}
    .filter-panel.open {{
        max-height: 800px; opacity: 1; padding: 20px;
    }}
    .fp-row {{
        display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 16px;
    }}
    .fp-row:last-child {{ margin-bottom: 0; }}
    .fp-simple {{ display: flex; align-items: center; gap: 8px; }}
    .fp-simple label {{ font-size: 12px; font-weight: 600; color: #94a3b8; text-transform: uppercase; letter-spacing: .5px; white-space: nowrap; }}
    .fp-simple input[type="text"] {{
        background: #0f172a; border: 1px solid #334155; border-radius: 6px;
        color: #e2e8f0; padding: 6px 10px; font-size: 13px; width: 170px;
        outline: none; transition: border-color .15s;
    }}
    .fp-simple input[type="text"]:focus {{ border-color: #4472C4; }}
    .fp-simple select {{
        background: #0f172a; border: 1px solid #334155; border-radius: 6px;
        color: #e2e8f0; padding: 6px 10px; font-size: 13px; width: 130px;
        outline: none; cursor: pointer;
    }}

    /* TextFilter card */
    .tf-card {{
        background: #0f172a; border: 1px solid #334155; border-radius: 8px;
        padding: 12px 14px; min-width: 280px; flex: 1;
    }}
    .tf-header {{
        display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;
    }}
    .tf-title {{ font-size: 12px; font-weight: 700; color: #38bdf8; text-transform: uppercase; letter-spacing: .5px; }}
    .tf-logic {{ display: flex; gap: 0; }}
    .tf-logic button {{
        padding: 3px 12px; font-size: 11px; font-weight: 700; border: 1px solid #334155;
        cursor: pointer; background: transparent; color: #64748b; transition: all .15s;
    }}
    .tf-logic button:first-child {{ border-radius: 4px 0 0 4px; }}
    .tf-logic button:last-child {{ border-radius: 0 4px 4px 0; border-left: 0; }}
    .tf-logic button.active {{ background: #0ea5e9; color: #fff; border-color: #0ea5e9; }}

    .tf-input-row {{
        display: flex; align-items: center; gap: 8px; margin-bottom: 8px;
    }}
    .tf-label {{
        font-size: 11px; font-weight: 700; padding: 3px 8px; border-radius: 4px;
        text-align: center; min-width: 36px;
    }}
    .tf-label.pos {{ background: #fff; color: #1e293b; }}
    .tf-label.neg {{ background: transparent; color: #f87171; border: 1px solid #7f1d1d; }}
    .tf-input {{
        flex: 1; background: #1e293b; border: 1px solid #334155; border-radius: 6px;
        color: #e2e8f0; padding: 6px 10px; font-size: 13px; outline: none;
    }}
    .tf-input:focus {{ border-color: #4472C4; }}
    .tf-input.neg-input {{ border-color: #7f1d1d; }}
    .tf-input.neg-input:focus {{ border-color: #ef4444; }}

    .tf-opts {{
        display: flex; flex-wrap: wrap; gap: 6px 14px; margin-top: 6px;
    }}
    .tf-opt {{
        display: flex; align-items: center; gap: 4px; font-size: 11px; color: #94a3b8; cursor: pointer; user-select: none;
    }}
    .tf-opt input {{ accent-color: #4472C4; cursor: pointer; }}

    /* Filter action buttons */
    .fp-actions {{
        display: flex; gap: 10px; align-items: center; margin-left: auto;
    }}
    .fp-actions .btn {{
        border: none; border-radius: 6px; padding: 7px 18px; cursor: pointer;
        font-size: 13px; font-weight: 600;
    }}
    .btn-apply {{ background: #0ea5e9; color: #fff; }}
    .btn-apply:hover {{ background: #0284c7; }}
    .btn-clear {{ background: #334155; color: #cbd5e1; }}
    .btn-clear:hover {{ background: #475569; }}
    .filter-info {{ font-size: 12px; color: #94a3b8; }}
</style>
</head>
<body>

<!-- Toolbar -->
<div class="tmx-toolbar">
    <button class="btn btn-select" id="btnSelectAll" onclick="toggleSelectAll()">☑ Select All</button>
    <div class="toolbar-sep"></div>
    <span style="font-size:13px;color:#555;">Selected:</span>
    <span class="badge zero" id="selCount">0</span>
    <div class="toolbar-sep"></div>
    <button class="btn btn-export" id="btnExport" onclick="exportTMX()" disabled>📦 Export TMX</button>
    <span id="exportStatus" style="font-size:12px;color:#888;"></span>
    <div class="toolbar-sep"></div>
    <button class="btn btn-filter" id="btnFilterToggle" onclick="toggleFilterPanel()">🔍 Filters</button>
    <span class="filter-info" id="filterInfo"></span>
</div>

    <h1>Tranzor All Translations <span class="count" style="background:#4472C4;">{len(rows)} entries</span></h1>
    <p class="meta">{html.escape(label)}</p>

<!-- Filter Panel -->
<div class="filter-panel" id="filterPanel">
    <div class="fp-row">
        <div class="fp-simple">
            <label>Task ID</label>
            <input type="text" id="fTaskId" placeholder="e.g. 12345">
        </div>
        <div class="fp-simple">
            <label>Task</label>
            <input type="text" id="fTask" placeholder="Keyword…">
        </div>
        <div class="fp-simple">
            <label>Lang</label>
            <select id="fLang"><option value="">All</option></select>
        </div>
        <div class="fp-simple">
            <label>Type</label>
            <select id="fType"><option value="">All</option></select>
        </div>
        <div class="fp-actions">
            <button class="btn btn-apply" onclick="applyFilters()">▶ Apply</button>
            <button class="btn btn-clear" onclick="clearFilters()">✕ Clear</button>
        </div>
    </div>
    <div class="fp-row" id="tfRow">
        <!-- TextFilter cards rendered by JS -->
    </div>
</div>

    <div class="toc">{toc_items}</div>
    {sections_html}

<script src="https://cdn.jsdelivr.net/npm/jszip@3/dist/jszip.min.js"></script>
<script>
// ============================================================
// Row data & selection state
// ============================================================
const ROWS = {rows_json};
const ALL_LANGS = {langs_json};
const ALL_TYPES = {types_json};
let allSelected = false;

// Populate lang dropdown
(function() {{
    const sel = document.getElementById('fLang');
    ALL_LANGS.forEach(l => {{
        const opt = document.createElement('option');
        opt.value = l; opt.textContent = l;
        sel.appendChild(opt);
    }});
    const typeSel = document.getElementById('fType');
    ALL_TYPES.forEach(t => {{
        const opt = document.createElement('option');
        opt.value = t; opt.textContent = t;
        typeSel.appendChild(opt);
    }});
}})();

// ============================================================
// TextFilter card rendering
// ============================================================
const TF_FIELDS = [
    {{ id: 'stringKey', label: 'STRING KEY', dataKey: 'string_key' }},
    {{ id: 'source',    label: 'SOURCE (EN-US)', dataKey: 'source_text' }},
    {{ id: 'translated', label: 'TRANSLATED TEXT', dataKey: 'translated_text' }},
];

(function renderTFCards() {{
    const container = document.getElementById('tfRow');
    TF_FIELDS.forEach(f => {{
        container.innerHTML += `
        <div class="tf-card" data-field="${{f.id}}">
            <div class="tf-header">
                <span class="tf-title">${{f.label}}</span>
                <div class="tf-logic">
                    <button class="active" data-val="AND" onclick="toggleLogic(this)">AND</button>
                    <button data-val="OR" onclick="toggleLogic(this)">OR</button>
                </div>
            </div>
            <div class="tf-input-row">
                <span class="tf-label pos">Pos</span>
                <input class="tf-input" data-role="pos" placeholder="Positive keyword…">
            </div>
            <div class="tf-input-row">
                <span class="tf-label neg">Neg</span>
                <input class="tf-input neg-input" data-role="neg" placeholder="Negative keyword (Exclude)…">
            </div>
            <div class="tf-opts">
                <label class="tf-opt"><input type="checkbox" data-role="matchWhole"> Match whole</label>
                <label class="tf-opt"><input type="checkbox" data-role="posCaseSensitive"> Match case (Pos)</label>
                <label class="tf-opt"><input type="checkbox" data-role="posRegex"> Regex (Pos)</label>
                <label class="tf-opt"><input type="checkbox" data-role="negCaseSensitive"> Match case (Neg)</label>
                <label class="tf-opt"><input type="checkbox" data-role="negRegex"> Regex (Neg)</label>
            </div>
        </div>`;
    }});
}})();

function toggleLogic(btn) {{
    const group = btn.parentElement;
    group.querySelectorAll('button').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
}}

function toggleFilterPanel() {{
    const panel = document.getElementById('filterPanel');
    const btn = document.getElementById('btnFilterToggle');
    panel.classList.toggle('open');
    btn.classList.toggle('active');
}}

// ============================================================
// Filter engine
// ============================================================
function getTextFilterState(cardEl) {{
    const logicBtn = cardEl.querySelector('.tf-logic button.active');
    return {{
        logic: logicBtn ? logicBtn.dataset.val : 'AND',
        pos: cardEl.querySelector('[data-role="pos"]').value,
        neg: cardEl.querySelector('[data-role="neg"]').value,
        posCaseSensitive: cardEl.querySelector('[data-role="posCaseSensitive"]').checked,
        posRegex: cardEl.querySelector('[data-role="posRegex"]').checked,
        negCaseSensitive: cardEl.querySelector('[data-role="negCaseSensitive"]').checked,
        negRegex: cardEl.querySelector('[data-role="negRegex"]').checked,
        matchWhole: cardEl.querySelector('[data-role="matchWhole"]').checked,
    }};
}}

function testMatch(text, keyword, caseSensitive, isRegex, matchWhole) {{
    if (!keyword) return null;
    try {{
        let pattern;
        if (isRegex) {{
            pattern = new RegExp(keyword, caseSensitive ? '' : 'i');
        }} else {{
            const escaped = keyword.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&');
            const p = matchWhole ? ('\\\\b' + escaped + '\\\\b') : escaped;
            pattern = new RegExp(p, caseSensitive ? '' : 'i');
        }}
        return pattern.test(text);
    }} catch(e) {{
        return false;
    }}
}}

function evaluateTextFilter(text, tf) {{
    const posResult = testMatch(text, tf.pos, tf.posCaseSensitive, tf.posRegex, tf.matchWhole);
    const negResult = testMatch(text, tf.neg, tf.negCaseSensitive, tf.negRegex, tf.matchWhole);

    if (tf.logic === 'AND') {{
        const posPass = (posResult === null) ? true : posResult;
        const negPass = (negResult === null) ? true : !negResult;
        return posPass && negPass;
    }} else {{
        if (posResult === null && negResult === null) return true;
        const posPass = posResult === true;
        const negPass = (negResult !== null) && !negResult;
        return posPass || negPass;
    }}
}}

function applyFilters() {{
    const fTaskId = document.getElementById('fTaskId').value.trim();
    const fTask = document.getElementById('fTask').value.trim();
    const fLang = document.getElementById('fLang').value;
    const fType = document.getElementById('fType').value;

    const tfCards = document.querySelectorAll('.tf-card');
    const tfStates = {{}};
    tfCards.forEach(card => {{
        const field = card.dataset.field;
        tfStates[field] = getTextFilterState(card);
    }});

    const allRows = document.querySelectorAll('input.row-cb');
    let visibleCount = 0;
    let totalCount = allRows.length;

    allRows.forEach(cb => {{
        const idx = parseInt(cb.dataset.idx);
        const row = ROWS[idx];
        const tr = cb.closest('tr');
        let pass = true;

        if (fTaskId && pass) {{
            if (String(row.task_id).indexOf(fTaskId) === -1) pass = false;
        }}
        if (fTask && pass) {{
            if (row.task_name.toLowerCase().indexOf(fTask.toLowerCase()) === -1) pass = false;
        }}
        if (fLang && pass) {{
            if (row.language !== fLang) pass = false;
        }}
        if (fType && pass) {{
            if (row.translation_type !== fType) pass = false;
        }}

        if (pass) {{
            TF_FIELDS.forEach(f => {{
                if (!pass) return;
                const tf = tfStates[f.id];
                if (!tf || (!tf.pos && !tf.neg)) return;
                const text = row[f.dataKey] || '';
                if (!evaluateTextFilter(text, tf)) pass = false;
            }});
        }}

        if (pass) {{
            tr.classList.remove('row-hidden');
            visibleCount++;
        }} else {{
            tr.classList.add('row-hidden');
        }}
    }});

    const info = document.getElementById('filterInfo');
    if (visibleCount < totalCount) {{
        info.textContent = 'Showing ' + visibleCount + ' / ' + totalCount;
        info.style.color = '#e67e22';
    }} else {{
        info.textContent = '';
    }}

    document.querySelectorAll('.lang-section table').forEach(table => {{
        const visibleCbs = table.querySelectorAll('tbody tr:not(.row-hidden) input.row-cb');
        const checkedCbs = table.querySelectorAll('tbody tr:not(.row-hidden) input.row-cb:checked');
        const sectionCb = table.querySelector('input.section-cb');
        if (sectionCb) {{
            sectionCb.checked = visibleCbs.length > 0 && visibleCbs.length === checkedCbs.length;
        }}
    }});
    updateBadge();
}}

function clearFilters() {{
    document.getElementById('fTaskId').value = '';
    document.getElementById('fTask').value = '';
    document.getElementById('fLang').value = '';
    document.getElementById('fType').value = '';

    document.querySelectorAll('.tf-card').forEach(card => {{
        card.querySelector('[data-role="pos"]').value = '';
        card.querySelector('[data-role="neg"]').value = '';
        card.querySelectorAll('.tf-opts input[type="checkbox"]').forEach(cb => cb.checked = false);
        const btns = card.querySelectorAll('.tf-logic button');
        btns.forEach(b => b.classList.remove('active'));
        btns[0].classList.add('active');
    }});

    document.querySelectorAll('tr.row-hidden').forEach(tr => tr.classList.remove('row-hidden'));
    document.getElementById('filterInfo').textContent = '';
    updateBadge();
}}

// ============================================================
// Selection helpers
// ============================================================
function getVisibleRowCheckboxes() {{
    return document.querySelectorAll('tr:not(.row-hidden) input.row-cb');
}}

function updateBadge() {{
    const n = document.querySelectorAll('input.row-cb:checked').length;
    const badge = document.getElementById('selCount');
    badge.textContent = n;
    badge.className = n ? 'badge' : 'badge zero';
    document.getElementById('btnExport').disabled = (n === 0);
}}

function toggleSelectAll() {{
    allSelected = !allSelected;
    getVisibleRowCheckboxes().forEach(cb => {{ cb.checked = allSelected; highlightRow(cb); }});
    document.querySelectorAll('input.section-cb').forEach(cb => cb.checked = allSelected);
    document.getElementById('btnSelectAll').textContent = allSelected ? '☐ Deselect All' : '☑ Select All';
    updateBadge();
}}

document.addEventListener('change', function(e) {{
    if (e.target.classList.contains('section-cb')) {{
        const tbody = e.target.closest('table').querySelector('tbody');
        tbody.querySelectorAll('tr:not(.row-hidden) input.row-cb').forEach(cb => {{
            cb.checked = e.target.checked;
            highlightRow(cb);
        }});
        updateBadge();
    }}
    if (e.target.classList.contains('row-cb')) {{
        highlightRow(e.target);
        const tbody = e.target.closest('tbody');
        const visibleInSection = tbody.querySelectorAll('tr:not(.row-hidden) input.row-cb');
        const checkedInSection = tbody.querySelectorAll('tr:not(.row-hidden) input.row-cb:checked');
        const sectionCb = e.target.closest('table').querySelector('input.section-cb');
        if (sectionCb) sectionCb.checked = (visibleInSection.length === checkedInSection.length);
        updateBadge();
    }}
}});

function highlightRow(cb) {{
    const tr = cb.closest('tr');
    if (cb.checked) tr.classList.add('row-selected');
    else tr.classList.remove('row-selected');
}}

// ============================================================
// TMX generation
// ============================================================
function escapeXml(s) {{
    return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&apos;');
}}

function langUnderscore(lang) {{
    return lang.replace('-', '_');
}}

function buildTmx(srcLang, tgtLang, entries) {{
    let xml = '<?xml version="1.0" encoding="UTF-8"?>\\n';
    xml += '<!DOCTYPE tmx SYSTEM "tmx14.dtd">\\n';
    xml += '<tmx version="1.4">';
    xml += '<header adminlang="en-US" creationtool="Tranzor" creationtoolversion="1.0"'
         + ' datatype="xml" o-tmf="Tranzor" srclang="' + srcLang + '" segtype="sentence">';
    xml += '</header>';
    xml += '<body>\\n';

    entries.forEach(function(e, i) {{
        const tuid = 'tranzor-' + i;
        xml += '<tu tuid="' + tuid + '">';
        xml += '<prop type="x-segment-id">' + escapeXml(e.string_key) + '</prop>';
        xml += '<prop type="x-Project">' + escapeXml(e.task_name) + '</prop>';
        xml += '\\n<tuv xml:lang="' + srcLang + '"><seg>' + escapeXml(e.source_text) + '</seg></tuv>\\n';
        xml += '<tuv xml:lang="' + tgtLang + '"><seg>' + escapeXml(e.translated_text) + '</seg></tuv>\\n';
        xml += '</tu>\\n';
    }});

    xml += '</body></tmx>';
    return xml;
}}

// ============================================================
// Download helpers
// ============================================================
function downloadBlob(blob, name) {{
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = name;
    document.body.appendChild(a);
    a.click();
    setTimeout(() => {{ URL.revokeObjectURL(a.href); a.remove(); }}, 100);
}}

async function exportTMX() {{
    const checked = document.querySelectorAll('tr:not(.row-hidden) input.row-cb:checked');
    if (!checked.length) return;

    const status = document.getElementById('exportStatus');
    status.textContent = 'Generating…';

    const byLang = {{}};
    checked.forEach(cb => {{
        const idx = parseInt(cb.dataset.idx);
        const row = ROWS[idx];
        const lang = row.language;
        if (!byLang[lang]) byLang[lang] = [];
        byLang[lang].push(row);
    }});

    const langs = Object.keys(byLang);

    if (langs.length === 1) {{
        const lang = langs[0];
        const tmx = buildTmx('en-US', lang, byLang[lang]);
        const blob = new Blob([tmx], {{ type: 'application/xml' }});
        const fn = 'tranzor_translations_en_US-' + langUnderscore(lang) + '.tmx';
        downloadBlob(blob, fn);
        status.textContent = '✓ Downloaded ' + fn + ' (' + byLang[lang].length + ' entries)';
    }} else {{
        if (typeof JSZip === 'undefined') {{
            status.textContent = '⚠ JSZip not loaded. Check your internet connection and reload.';
            return;
        }}
        const zip = new JSZip();
        let totalEntries = 0;
        langs.forEach(lang => {{
            const tmx = buildTmx('en-US', lang, byLang[lang]);
            const fn = 'en_US-' + langUnderscore(lang) + '.tmx';
            zip.file(fn, tmx);
            totalEntries += byLang[lang].length;
        }});
        const content = await zip.generateAsync({{ type: 'blob' }});
        downloadBlob(content, 'tranzor_translations_export.zip');
        status.textContent = '✓ Downloaded ZIP with ' + langs.length + ' languages, ' + totalEntries + ' entries';
    }}
}}
</script>
</body>
</html>"""

    with open(filename, "w", encoding="utf-8") as f:
        f.write(page)


# ---------------------------------------------------------------------------
# 6) 写入 Excel
# ---------------------------------------------------------------------------
EXCEL_COLUMNS = [
    "Task ID", "Task Name", "Language", "String Key",
    "Source (en-US)", "Translated Text", "Translation Type",
]


def write_excel(rows, filename):
    """将结果写入 Excel 文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("错误: 缺少 openpyxl，请运行: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "All Translations"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    source_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    translated_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for row_idx, r in enumerate(rows, 2):
        values = [
            r["task_id"], r["task_name"], r["language"], r["string_key"],
            r["source_text"], r["translated_text"], r["translation_type"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 5:
                cell.fill = source_fill
            elif col_idx == 6:
                cell.fill = translated_fill

    col_widths = [10, 25, 10, 35, 40, 40, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(filename)


# ---------------------------------------------------------------------------
# 7) 保存文件（含文件被占用时自动重命名 + 大数据自动分页）
# ---------------------------------------------------------------------------
PAGE_SIZE = 5000  # 超过此阈值按语言分页


def _write_index_html(filepath, label, lang_files, total_count):
    """生成分页索引页面，带语言级别导航链接和统计"""
    import html as _html

    lang_cards = ""
    for lang, fname, count in lang_files:
        color = "#4472C4"
        lang_cards += f"""
        <a href="{_html.escape(fname)}" class="lang-card" style="border-left:4px solid {color};">
            <div class="lang-name">🌐 {_html.escape(lang)}</div>
            <div class="lang-count">{count:,} entries</div>
            <div class="lang-file">{_html.escape(fname)}</div>
        </a>"""

    page = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Tranzor Translations - {_html.escape(label)}</title>
<style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: -apple-system, "Segoe UI", Roboto, Arial, sans-serif; background: #f5f6fa; padding: 40px; color: #333; }}
    .header {{ max-width: 900px; margin: 0 auto 32px; }}
    h1 {{ font-size: 24px; margin-bottom: 8px; }}
    .meta {{ color: #666; font-size: 14px; margin-bottom: 6px; }}
    .stats {{ display: flex; gap: 24px; margin-bottom: 24px; }}
    .stat-card {{
        background: #fff; border-radius: 10px; padding: 20px 28px;
        box-shadow: 0 2px 8px rgba(0,0,0,.06); flex: 1; text-align: center;
    }}
    .stat-num {{ font-size: 32px; font-weight: 700; color: #4472C4; }}
    .stat-label {{ font-size: 13px; color: #888; margin-top: 4px; }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 16px; max-width: 900px; margin: 0 auto; }}
    .lang-card {{
        display: block; background: #fff; border-radius: 10px; padding: 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,.06); text-decoration: none; color: #333;
        transition: transform .15s, box-shadow .15s; cursor: pointer;
    }}
    .lang-card:hover {{ transform: translateY(-3px); box-shadow: 0 6px 20px rgba(0,0,0,.12); }}
    .lang-name {{ font-size: 17px; font-weight: 700; margin-bottom: 6px; }}
    .lang-count {{ font-size: 14px; color: #4472C4; font-weight: 600; margin-bottom: 4px; }}
    .lang-file {{ font-size: 11px; color: #999; font-family: monospace; word-break: break-all; }}
    .note {{ max-width: 900px; margin: 32px auto 0; padding: 16px 20px; background: #fff8e1; border-radius: 8px;
             border-left: 4px solid #ffc107; font-size: 13px; color: #666; }}
</style>
</head>
<body>
<div class="header">
    <h1>📑 Tranzor All Translations</h1>
    <p class="meta">{_html.escape(label)}</p>
    <div class="stats">
        <div class="stat-card">
            <div class="stat-num">{total_count:,}</div>
            <div class="stat-label">Total Entries</div>
        </div>
        <div class="stat-card">
            <div class="stat-num">{len(lang_files)}</div>
            <div class="stat-label">Languages</div>
        </div>
    </div>
</div>
<div class="grid">{lang_cards}</div>
<div class="note">
    ℹ️ Data has been split by language to ensure smooth browsing. Click any language card above to open its report. 
    Each report supports filtering, search, and TMX export.
</div>
</body>
</html>"""

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(page)


def save_file(rows, filename, label, fmt):
    """保存文件，文件被占用时自动加序号；大数据 HTML 自动按语言分页"""
    from collections import OrderedDict

    base, ext = os.path.splitext(filename)

    # HTML 大数据集自动分页
    if fmt == "html" and len(rows) > PAGE_SIZE:
        print(f"\n📊 Data volume ({len(rows):,} rows) exceeds {PAGE_SIZE:,}, splitting by language...")

        # 按语言分组
        groups = OrderedDict()
        for r in rows:
            lang = r.get("language") or "(unknown)"
            groups.setdefault(lang, []).append(r)

        # 生成每语言的子文件
        lang_files = []
        dir_path = os.path.dirname(os.path.abspath(filename))
        base_name = os.path.basename(base)

        for lang, lang_rows in groups.items():
            safe_lang = lang.replace(" ", "_").replace("(", "").replace(")", "")
            sub_filename = f"{base_name}_{safe_lang}{ext}"
            sub_filepath = os.path.join(dir_path, sub_filename)
            sub_label = f"{label} — {lang}"

            # 写入子文件（带重试）
            save_path = sub_filepath
            for attempt in range(100):
                try:
                    write_html(lang_rows, save_path, sub_label)
                    print(f"  ✓ {lang}: {len(lang_rows):,} entries → {os.path.basename(save_path)}")
                    lang_files.append((lang, os.path.basename(save_path), len(lang_rows)))
                    break
                except PermissionError:
                    save_path = f"{os.path.splitext(sub_filepath)[0]}_{attempt+1}{ext}"

        # 生成索引页
        index_path = filename
        for attempt in range(100):
            try:
                _write_index_html(index_path, label, lang_files, len(rows))
                print(f"\n📑 Index page: {os.path.basename(index_path)}")
                webbrowser.open(os.path.abspath(index_path))
                return
            except PermissionError:
                index_path = f"{base}_{attempt+1}{ext}"

    else:
        # 小数据集或 Excel：正常单文件输出
        save_path = filename
        for attempt in range(100):
            try:
                if fmt == "html":
                    write_html(rows, save_path, label)
                else:
                    write_excel(rows, save_path)
                print(f"已导出: {save_path}")
                if fmt == "html":
                    webbrowser.open(os.path.abspath(save_path))
                return
            except PermissionError:
                attempt_num = attempt + 1
                save_path = f"{base}_{attempt_num}{ext}"
                print(f"  文件被占用，尝试保存为: {save_path}")


# ---------------------------------------------------------------------------
# 8) 命令行入口
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="导出 Tranzor 翻译任务的全部翻译记录",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python export_translations.py --task 123         导出指定 task 的全部翻译 (HTML)
  python export_translations.py --task 123 --xlsx  导出指定 task 的全部翻译 (Excel)
  python export_translations.py                    导出所有已完成 task 的全部翻译
        """,
    )
    parser.add_argument("--task", type=int, help="只导出指定 task ID 的翻译")
    parser.add_argument("--xlsx", action="store_true", help="输出 Excel 格式（默认 HTML）")
    parser.add_argument("--output", "-o", help="输出文件名 (默认自动生成)")

    args = parser.parse_args()

    if args.task:
        print(f"指定 Task: {args.task}")

    rows = collect_translations(task_id=args.task)
    print(f"\n共找到 {len(rows)} 条翻译记录")

    if not rows:
        print("没有翻译记录，无需导出。")
        return

    # 确定输出格式和文件名
    fmt = "xlsx" if args.xlsx else "html"
    ext = ".xlsx" if args.xlsx else ".html"
    today_str = date.today().isoformat()
    label = f"All translations (exported {today_str})"

    if args.output:
        filename = args.output
    elif args.task:
        filename = f"tranzor_task_{args.task}_translations_{today_str}{ext}"
    else:
        filename = f"tranzor_all_translations_{today_str}{ext}"

    save_file(rows, filename, label, fmt)


if __name__ == "__main__":
    main()
