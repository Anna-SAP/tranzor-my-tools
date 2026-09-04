"""Tests for the MR Pipeline / Scan Tasks "Export Source XLSX" schema.

The workbook is a source-only companion to the QA JSON export:

  * three columns: Key / en-US Value / task name
  * MR sheet named ``{JIRA} MR!{iid}`` (e.g. ``BUG-352 MR!4103``)
  * Scan sheet named after the task (e.g. ``UNS release_26-3-3 02``)
  * UNS email segments keep the ``:::seg:::{tu_id}`` key shape
  * task name is the companion All-Translations JSON filename

Run:  python -m unittest test_source_xlsx
"""
from __future__ import annotations

import os
import sys
import tempfile
import unittest
from datetime import date
from unittest import mock

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import export_json as ej
from gui_tabs import MRPipelineTab
from gui_tab_scan_tasks import ScanTasksTab
import gui_tab_scan_tasks as scan_mod
import task_post_edit as tpe

_build = MRPipelineTab._build_export_filename
_build_scan = ScanTasksTab._build_export_filename

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


SEG_ROWS = [
    {"opus_id": "common.uns.announcementsOnlyLoginInfo__email_html__3460",
     "has_seg_units": True, "tu_id": 10, "target_language": "de-DE",
     "source_text": "Click on login", "translated_text": "Klicken Sie"},
    {"opus_id": "common.uns.announcementsOnlyLoginInfo__email_html__3460",
     "has_seg_units": True, "tu_id": 4, "target_language": "zh-CN",
     "source_text": "Dear <ph-6-var-ExtensionFullName/>,",
     "translated_text": "亲爱的"},
    {"opus_id": "common.uns.announcementsOnlyLoginInfo__email_html__3460",
     "has_seg_units": True, "tu_id": 10, "target_language": "zh-CN",
     "source_text": "Click on login", "translated_text": "点击登录"},
    {"opus_id": "plain.key", "target_language": "en-US",
     "source_text": "Save", "translated_text": "Save"},
]


class TestSourceRows(unittest.TestCase):

    def test_unique_keys_with_en_us_fallback_and_task_name(self):
        json_name = (
            "mr_pipeline_MR4100_bd6bba88_all_2026-08-31_08-03-01.json")
        rows = ej.source_rows_from_payload(
            {"translations": SEG_ROWS}, task_name=json_name)
        by_key = {key: (en_us, task) for key, en_us, task in rows}
        self.assertEqual(
            by_key["common.uns.announcementsOnlyLoginInfo__email_html__3460:::seg:::10"],
            ("Click on login", json_name))
        self.assertEqual(
            by_key["common.uns.announcementsOnlyLoginInfo__email_html__3460:::seg:::4"],
            ("Dear <ph-6-var-ExtensionFullName/>,", json_name))
        self.assertEqual(by_key["plain.key"], ("Save", json_name))
        # One row per unique key — language fan-out is collapsed.
        self.assertEqual(len(rows), 3)

    def test_segment_keys_sort_lexicographically(self):
        # Screenshot 2 shows :::seg:::10 before :::seg:::4 (string sort).
        json_name = "mr_pipeline_MR4100_bd6bba88_all_2026-08-31_08-03-01.json"
        rows = ej.source_rows_from_payload(
            {"translations": SEG_ROWS}, task_name=json_name)
        keys = [key for key, _, _ in rows]
        self.assertEqual(keys, [
            "common.uns.announcementsOnlyLoginInfo__email_html__3460:::seg:::10",
            "common.uns.announcementsOnlyLoginInfo__email_html__3460:::seg:::4",
            "plain.key",
        ])

    def test_dict_or_list_payload(self):
        from_dict = ej.source_rows_from_payload({"translations": SEG_ROWS})
        from_list = ej.source_rows_from_payload(SEG_ROWS)
        self.assertEqual(from_dict, from_list)

    def test_empty_payload(self):
        self.assertEqual(ej.source_rows_from_payload({}), [])
        self.assertEqual(ej.source_rows_from_payload(None), [])


class TestSheetTitleForMr(unittest.TestCase):

    def test_jira_and_mr_match_screenshot_tabs(self):
        self.assertEqual(
            ej.sheet_title_for_mr("BUG-352", "4103"), "BUG-352 MR!4103")
        self.assertEqual(
            ej.sheet_title_for_mr("BUG-502", "4100"), "BUG-502 MR!4100")
        self.assertEqual(
            ej.sheet_title_for_mr("RLZ-56748", "4099"), "RLZ-56748 MR!4099")
        self.assertEqual(
            ej.sheet_title_for_mr("UIA-413515", "4098"), "UIA-413515 MR!4098")

    def test_missing_jira_falls_back_to_mr(self):
        self.assertEqual(ej.sheet_title_for_mr("—", "4090"), "MR!4090")
        self.assertEqual(ej.sheet_title_for_mr("…", "4090"), "MR!4090")
        self.assertEqual(ej.sheet_title_for_mr("", "4090"), "MR!4090")
        self.assertEqual(ej.sheet_title_for_mr(None, "4090"), "MR!4090")

    def test_strips_mr_prefix_and_hash(self):
        self.assertEqual(
            ej.sheet_title_for_mr("BUG-352", "MR!4103"), "BUG-352 MR!4103")
        self.assertEqual(
            ej.sheet_title_for_mr("BUG-352", "#4103"), "BUG-352 MR!4103")

    def test_both_missing(self):
        self.assertEqual(ej.sheet_title_for_mr("", ""), "Source")


class TestSheetTitleForScan(unittest.TestCase):

    def test_prefers_task_name(self):
        self.assertEqual(
            ej.sheet_title_for_scan("UNS release_26-3-3 02", "06f77242-aaaa"),
            "UNS release_26-3-3 02")
        self.assertEqual(
            ej.sheet_title_for_scan("LOC-25241 UNS 26.4", "abc"),
            "LOC-25241 UNS 26.4")

    def test_empty_name_falls_back_to_scan_uuid(self):
        self.assertEqual(
            ej.sheet_title_for_scan("", "75040f78-aaaa-bbbb"),
            "Scan 75040f78")
        self.assertEqual(
            ej.sheet_title_for_scan("—", "75040f78-aaaa-bbbb"),
            "Scan 75040f78")
        self.assertEqual(
            ej.sheet_title_for_scan(None, "75040f78"),
            "Scan 75040f78")

    def test_both_missing(self):
        self.assertEqual(ej.sheet_title_for_scan("", ""), "Source")
        self.assertEqual(ej.sheet_title_for_scan(None, None), "Source")


class TestSanitizeExcelSheetTitle(unittest.TestCase):

    def test_exclamation_mark_is_kept(self):
        used = set()
        self.assertEqual(
            ej.sanitize_excel_sheet_title("BUG-352 MR!4103", used),
            "BUG-352 MR!4103")

    def test_invalid_chars_and_uniqueness(self):
        used = set()
        a = ej.sanitize_excel_sheet_title("A:B/C", used)
        b = ej.sanitize_excel_sheet_title("A:B/C", used)
        self.assertEqual(a, "A-B-C")
        self.assertEqual(b, "A-B-C (1)")

    def test_truncates_to_31_chars(self):
        used = set()
        long_name = "VERY-LONG-JIRA-TICKET-NAME MR!99999"
        title = ej.sanitize_excel_sheet_title(long_name, used)
        self.assertLessEqual(len(title), 31)


@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
class TestWriteSourceXlsx(unittest.TestCase):

    def test_three_columns_header_and_sheet_name(self):
        json_name = (
            "mr_pipeline_MR4100_bd6bba88_all_2026-08-31_08-03-01.json")
        rows = ej.source_rows_from_payload(
            {"translations": SEG_ROWS}, task_name=json_name)
        title = ej.sheet_title_for_mr("BUG-502", "4100")
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "source.xlsx")
            saved = ej.write_source_xlsx(
                [{"title": title, "rows": rows}], path)
            self.assertEqual(saved, path)
            wb = load_workbook(path)
            self.assertEqual(wb.sheetnames, ["BUG-502 MR!4100"])
            ws = wb.active
            self.assertEqual(
                [cell.value for cell in ws[1]],
                ["Key", "en-US Value", "task name"])
            self.assertEqual(
                ws.cell(2, 1).value,
                "common.uns.announcementsOnlyLoginInfo__email_html__3460:::seg:::10")
            self.assertEqual(ws.cell(2, 2).value, "Click on login")
            self.assertEqual(ws.cell(2, 3).value, json_name)
            # task name stays in a single cell (not split on underscores).
            self.assertIsNone(ws.cell(2, 4).value)
            self.assertEqual(ws.freeze_panes, "A2")

    def test_multiple_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "multi.xlsx")
            ej.write_source_xlsx(
                [
                    {"title": "BUG-352 MR!4103",
                     "rows": [("k1", "Hello", "a.json")]},
                    {"title": "BUG-502 MR!4100",
                     "rows": [("k2", "World", "b.json")]},
                ],
                path,
            )
            wb = load_workbook(path)
            self.assertEqual(
                wb.sheetnames, ["BUG-352 MR!4103", "BUG-502 MR!4100"])
            self.assertEqual(wb["BUG-352 MR!4103"]["A2"].value, "k1")
            self.assertEqual(wb["BUG-502 MR!4100"]["B2"].value, "World")

    def test_save_source_xlsx_returns_path(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "out.xlsx")
            saved = ej.save_source_xlsx(
                [{"title": "MR!1", "rows": [("k", "v", "t.json")]}], path)
            self.assertEqual(saved, path)
            self.assertTrue(os.path.isfile(saved))


class TestI18n(unittest.TestCase):

    def test_button_strings_in_both_languages(self):
        from export_gui import STRINGS
        for lang in ("en", "zh"):
            self.assertIn("mr_source_xlsx", STRINGS[lang])
            self.assertIn("mr_source_xlsx_need_selection", STRINGS[lang])
            self.assertTrue(STRINGS[lang]["mr_source_xlsx"].strip())
            self.assertIn("scan_source_xlsx_need_selection", STRINGS[lang])
            self.assertTrue(STRINGS[lang]["scan_source_xlsx_need_selection"].strip())


class TestSourceExportFilename(unittest.TestCase):

    def test_source_type_tag_and_json_companion_name(self):
        xlsx = _build(
            ".xlsx", mr_iid="4100", id_tag="bd6bba88", type_tag="source",
            created="2026-08-31 08:03:01", export_date="2026-09-02")
        json_name = _build(
            ".json", mr_iid="4100", id_tag="bd6bba88", type_tag="all",
            created="2026-08-31 08:03:01", export_date="2026-09-02")
        self.assertEqual(
            xlsx,
            "mr_pipeline_MR4100_bd6bba88_source_2026-08-31_08-03-01.xlsx")
        self.assertEqual(
            json_name,
            "mr_pipeline_MR4100_bd6bba88_all_2026-08-31_08-03-01.json")

    def test_scan_source_type_tag_and_json_companion_name(self):
        xlsx = _build_scan(
            ".xlsx", task_name="iva 260520", id_tag="75040f78",
            type_tag="source", created="2026-06-17 14:42:26",
            export_date="2026-06-18")
        json_name = _build_scan(
            ".json", task_name="iva 260520", id_tag="75040f78",
            type_tag="all", created="2026-06-17 14:42:26",
            export_date="2026-06-18")
        self.assertEqual(
            xlsx,
            "scan_task_iva_260520_75040f78_source_2026-06-17_14-42-26.xlsx")
        self.assertEqual(
            json_name,
            "scan_task_iva_260520_75040f78_all_2026-06-17_14-42-26.json")


class _FakeTree:
    def __init__(self, tags, values):
        self._tags = tags
        self._values = values

    def item(self, _iid, key):
        if key == "tags":
            return self._tags
        if key == "values":
            return self._values
        raise KeyError(key)


class _AfterParent:
    def after(self, _ms, fn, *args):
        fn(*args)


class _Status:
    def __init__(self):
        self.text = ""

    def configure(self, **kwargs):
        if "text" in kwargs:
            self.text = kwargs["text"]


class TestScanRowExportMeta(unittest.TestCase):

    def test_strips_post_edit_prefix_and_reads_created(self):
        tab = ScanTasksTab.__new__(ScanTasksTab)
        tab.scan_tree = _FakeTree(
            tags=("75040f78-aaaa-bbbb-cccc-dddddddddddd",),
            values=(
                1, tpe.POST_EDIT_PREFIX + "iva 260520", "common/uns",
                "base", "head", "completed", 10, "report_only",
                "2026-06-17 14:42:26 UTC+8", "3d",
            ),
        )
        meta = tab._scan_row_export_meta("row")
        self.assertEqual(meta["task_id"], "75040f78-aaaa-bbbb-cccc-dddddddddddd")
        self.assertEqual(meta["task_name"], "iva 260520")
        self.assertEqual(meta["created"], "2026-06-17 14:42:26 UTC+8")

    def test_plain_name_unchanged(self):
        tab = ScanTasksTab.__new__(ScanTasksTab)
        tab.scan_tree = _FakeTree(
            tags=("abc",),
            values=(1, "scheduled:integration-ips", "p", "b", "h",
                    "completed", 1, "create_mr", "2026-09-01", "today"),
        )
        meta = tab._scan_row_export_meta("row")
        self.assertEqual(meta["task_name"], "scheduled:integration-ips")

    def test_source_xlsx_requires_selection(self):
        tab = ScanTasksTab.__new__(ScanTasksTab)
        tab.lbl_scan_status_bar = _Status()
        tab.scan_tree = type("T", (), {"selection": staticmethod(lambda: ())})()
        tab._t = lambda k: k
        tab._on_export_source_xlsx()
        self.assertEqual(
            tab.lbl_scan_status_bar.text, "scan_source_xlsx_need_selection")


@unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
class TestScanSourceXlsxWorker(unittest.TestCase):

    def test_run_export_writes_scan_sheet_and_companion_json_name(self):
        tab = ScanTasksTab.__new__(ScanTasksTab)
        tab.parent = _AfterParent()
        tab.lbl_scan_status_bar = _Status()
        tab._enabled = []
        tab._set_scan_export_buttons_enabled = (
            lambda enabled, store=tab._enabled: store.append(enabled))
        tab._t = lambda k: {
            "status_exporting": "Exporting…",
            "status_saved": "✓ Saved: {filename}",
        }[k]

        payload = {"translations": SEG_ROWS}
        meta = {
            "task_id": "75040f78-aaaa-bbbb-cccc-dddddddddddd",
            "task_name": "iva 260520",
            "created": "2026-06-17 14:42:26",
        }
        xlsx_name = _build_scan(
            ".xlsx", task_name="iva 260520", id_tag="75040f78",
            type_tag="source", created="2026-06-17 14:42:26",
            export_date=date.today().isoformat())
        json_name = _build_scan(
            ".json", task_name="iva 260520", id_tag="75040f78",
            type_tag="all", created="2026-06-17 14:42:26",
            export_date=date.today().isoformat())
        with tempfile.TemporaryDirectory() as tmp, \
             mock.patch.object(scan_mod.mr_api, "fetch_scan_results",
                               return_value=payload) as fetch, \
             mock.patch.object(scan_mod, "export_output_dir",
                               return_value=tmp), \
             mock.patch.object(scan_mod, "reveal_in_folder") as reveal:
            tab._run_export_source_xlsx(meta)

            fetch.assert_called_once()
            self.assertEqual(fetch.call_args[0][0], meta["task_id"])
            saved = os.path.join(tmp, xlsx_name)
            self.assertTrue(os.path.isfile(saved), saved)
            from openpyxl import load_workbook
            wb = load_workbook(saved)
            self.assertEqual(wb.sheetnames, ["iva 260520"])
            ws = wb.active
            self.assertEqual(
                [cell.value for cell in ws[1]],
                ["Key", "en-US Value", "task name"])
            self.assertEqual(ws.cell(2, 3).value, json_name)
            reveal.assert_called_once_with(saved)
            self.assertEqual(tab._enabled, [True])
            self.assertIn(xlsx_name, tab.lbl_scan_status_bar.text)


if __name__ == "__main__":
    unittest.main()
